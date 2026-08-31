import os
from datetime import datetime
from flask import Blueprint, render_template, request, jsonify, send_file, current_app, session, redirect, url_for, flash
from app.services.database.queries import get_wo_summary_stats
from app.routes.auth import login_required

asp_bp = Blueprint("asp", __name__)


# ── Session helper ────────────────────────────────────────────────────────────

def _vendor_filter() -> str | None:
    """Return the labor_vendor_related value for the current ASP session, or None.

    asp_master: labor_vendor is None at login (they span the whole group), but
    may be scoped to a specific branch after switch_branch — so we still read
    the session value; it will be None (= unfiltered view) until a branch is chosen.
    asp_user: vendor filter is intentionally ignored — use _tech_id_filter() instead.
    """
    if session.get("role") in ("asp", "asp_master"):
        return session.get("labor_vendor") or None
    return None


def _tech_id_filter() -> str | None:
    """Return the tech_id for asp_user sessions only, or None for all other roles.

    When set, every WO query is narrowed to WOs assigned to this specific
    technician (wo_details.tech_id), so technicians cannot see each other's WOs.
    """
    if session.get("role") == "asp_user":
        return session.get("tech_id") or None
    return None


# ── Shared stat context ───────────────────────────────────────────────────────

def _stat_ctx() -> dict:
    """Stat counts only — no row data loaded on page request."""
    vf = _vendor_filter()
    tf = _tech_id_filter()
    s = get_wo_summary_stats(vendor_filter=vf, tech_id_filter=tf)
    return dict(
        total              = s["total"],
        total_closed       = s["closed"],
        total_open         = s["open"],
        total_part_hold    = s["part_hold"],
        total_part_transit = s["part_transit"],
        portal             = "asp",
    )


# ── Page routes ───────────────────────────────────────────────────────────────

@asp_bp.route("/asp/dashboard", methods=["GET"])
@login_required
def dashboard():
    ctx = _stat_ctx()
    ctx["active_page"] = "asp_dashboard"
    return render_template("asp/dashboard.html", **ctx)


@asp_bp.route("/asp/work-orders", methods=["GET"])
@login_required
def work_orders():
    ctx = _stat_ctx()
    tab = request.args.get("tab", "active")
    ctx["active_page"]  = {"active": "wo_active", "closed": "wo_closed",
                           "escalated": "wo_escalated", "pending": "wo_pending"}.get(tab, "wo_active")
    ctx["active_group"] = "work_orders"
    return render_template("asp/work_orders.html", **ctx)


@asp_bp.route("/asp/parts", methods=["GET"])
@login_required
def parts_management():
    ctx = _stat_ctx()
    tab = request.args.get("tab", "awaiting")
    ctx["active_page"]  = {"awaiting": "parts_awaiting", "received": "parts_received",
                            "return": "parts_return"}.get(tab, "parts_awaiting")
    ctx["active_group"] = "parts"
    return render_template("asp/parts_management.html", **ctx)


@asp_bp.route("/asp/reschedule", methods=["GET"])
@login_required
def reschedule():
    ctx = _stat_ctx()
    ctx["active_page"] = "reschedule"
    return render_template("asp/reschedule.html", **ctx)


@asp_bp.route("/asp/escalation", methods=["GET"])
@login_required
def escalation():
    ctx = _stat_ctx()
    ctx["active_page"] = "escalation"
    return render_template("asp/escalation.html", **ctx)


@asp_bp.route("/asp/switch-branch/<string:username>", methods=["GET"])
@login_required
def switch_branch(username: str):
    """Switch the current asp_master session context to a specific branch ASP.

    Only asp_master (or superadmin) may use this route.  Updates the session's
    labor_vendor and display_name so all pages filter correctly for the chosen
    branch.  Switching back to the master username (stored in original_username)
    clears the branch scope and returns to the unfiltered group view.
    """
    from app.services.database.db import get_db
    role         = session.get("role", "")
    own_username = session.get("username", "")

    if role not in ("superadmin", "asp_master"):
        flash("You do not have permission to switch offices.", "danger")
        return redirect(url_for("asp.dashboard"))

    db = get_db()

    # Fetch the target ASP
    target = db.execute(
        "SELECT username, service_provider, labor_vendor_related, parent_group, kota "
        "FROM asp_details WHERE username = ?",
        (username,),
    ).fetchone()

    if not target:
        flash("Branch office not found.", "danger")
        return redirect(url_for("asp.dashboard"))

    if role == "asp_master":
        # Verify the target belongs to this master's parent_group
        if target["parent_group"] != session.get("parent_group"):
            flash("You can only switch to offices in your own group.", "danger")
            return redirect(url_for("asp.dashboard"))

    # Preserve the original master identity on first switch
    if "original_username" not in session:
        session["original_username"]     = own_username
        session["original_display_name"] = session.get("display_name")
        session["original_labor_vendor"] = session.get("labor_vendor")
        session["original_office_kota"]  = session.get("office_kota", "")

    # Switching back to the master account: restore unscoped identity
    if username == session.get("original_username"):
        session["username"]     = session.pop("original_username")
        session["display_name"] = session.pop("original_display_name")
        session["labor_vendor"] = session.pop("original_labor_vendor")
        session["office_kota"]  = session.pop("original_office_kota", "")
    else:
        session["username"]     = target["username"]
        session["display_name"] = target["service_provider"] or target["username"]
        session["labor_vendor"] = target["labor_vendor_related"]
        session["office_kota"]  = target["kota"] or ""
        # is_hq_with_branches stays True — switcher must remain visible

    return redirect(url_for("asp.dashboard"))


@asp_bp.route("/asp/branch-office", methods=["GET"])
@login_required
def branch_office():
    """List all ASPs that share the same parent_group as the current asp_master user.
    Superadmin may pass ?parent_group=<name> to view any group."""
    from app.services.database.db import get_db
    role     = session.get("role", "")
    username = session.get("username", "")

    # Access control: only superadmin or asp_master
    if role not in ("superadmin", "asp_master"):
        flash("You do not have permission to access that page.", "danger")
        return redirect(url_for("asp.dashboard"))

    db = get_db()

    if role == "superadmin":
        # Superadmin may specify any parent_group via query string
        parent_group = request.args.get("parent_group", "").strip() or None
        current_asp  = None
    else:
        # asp_master: parent_group is stored directly in the session
        parent_group = session.get("parent_group")
        current_asp  = None

    if not parent_group:
        flash("No parent group configured for your account.", "warning")
        return redirect(url_for("asp.dashboard"))

    # Fetch all members of the group, HQ first then branches alphabetically
    members = db.execute(
        """
        SELECT username, service_provider, store_name, kota, island,
               vendor_code, labor_vendor_related, office_type,
               operational_status, operation_support, wo_count, phone_number
        FROM asp_details
        WHERE parent_group = ?
        ORDER BY
            CASE WHEN office_type = 'ASP HQ' THEN 0 ELSE 1 END,
            service_provider COLLATE NOCASE
        """,
        (parent_group,),
    ).fetchall()

    ctx = _stat_ctx()
    ctx.update(
        active_page  = "branch_office",
        parent_group = parent_group,
        members      = [dict(r) for r in members],
        current_asp  = current_asp,
    )
    return render_template("asp/branch_office.html", **ctx)


# ── API: tab data endpoints (server-side pagination) ─────────────────────────

def _int_arg(name: str, default: int) -> int:
    try:
        return max(1, int(request.args.get(name, default)))
    except (ValueError, TypeError):
        return default


def _bool_arg(name: str, default: bool) -> bool:
    value = request.args.get(name)
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


@asp_bp.route("/asp/api/all-wo", methods=["GET"])
@login_required
def api_all_wo():
    """All WO tab — filterable by status, WO type, and case status."""
    from app.services.database.queries import get_asp_all_wo_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_all_wo_page(
        search              = request.args.get("q", "").strip(),
        status_filter       = request.args.get("status", "").strip(),
        type_filter         = request.args.get("wo_type", "").strip(),
        case_status_filter  = request.args.get("case_status", "").strip(),
        page                = _int_arg("page", 1),
        page_size           = per_page,
        vendor_filter       = _vendor_filter(),
        tech_id_filter      = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/part-received", methods=["GET"])
@login_required
def api_part_received():
    """Part Received tab — WOs waiting for part / on part hold."""
    from app.services.database.queries import get_asp_part_received_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_part_received_page(
        search         = request.args.get("q", "").strip(),
        page           = _int_arg("page", 1),
        page_size      = per_page,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/cci-followup", methods=["GET"])
@login_required
def api_cci_followup():
    """CCI Follow-Up tab — all Carry-In WOs with computed followup_state."""
    from app.services.database.queries import get_asp_cci_followup_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_cci_followup_page(
        search         = request.args.get("q", "").strip(),
        followup_state = request.args.get("followup_state", "").strip(),
        page           = _int_arg("page", 1),
        page_size      = per_page,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/part-return", methods=["GET"])
@login_required
def api_part_return():
    """Part Return tab — closed / completed WOs."""
    from app.services.database.queries import get_asp_part_return_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_part_return_page(
        search         = request.args.get("q", "").strip(),
        page           = _int_arg("page", 1),
        page_size      = per_page,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/reschedule", methods=["GET"])
@login_required
def api_reschedule():
    """WO Reschedule tab — open WOs eligible for rescheduling."""
    from app.services.database.queries import get_asp_reschedule_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_reschedule_page(
        search         = request.args.get("q", "").strip(),
        page           = _int_arg("page", 1),
        page_size      = per_page,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/onsite-followup", methods=["GET"])
@login_required
def api_onsite_followup():
    """Onsite Follow-Up tab — all Onsite WOs with computed followup_state."""
    from app.services.database.queries import get_asp_onsite_followup_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_onsite_followup_page(
        search         = request.args.get("q", "").strip(),
        followup_state = request.args.get("followup_state", "").strip(),
        page           = _int_arg("page", 1),
        page_size      = per_page,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/wos-by-awb", methods=["GET"])
@login_required
def api_wos_by_awb():
    """Return all WOs sharing the given AWB number."""
    from app.services.database.queries import get_wos_by_awb
    awb = request.args.get("awb", "").strip()
    if not awb:
        return jsonify([])
    return jsonify(get_wos_by_awb(awb))


@asp_bp.route("/asp/api/wo-no-awb", methods=["GET"])
@login_required
def api_wo_no_awb():
    """Return open WOs for a given ASP (customer name) that have part lines with no AWB.
    If current_wo is supplied, that WO is always included even if it already has an AWB."""
    from app.services.database.queries import get_wo_no_awb_by_asp
    customer = request.args.get("customer", "").strip()
    if not customer:
        return jsonify([])
    current_wo = request.args.get("current_wo", "").strip()
    try:
        current_wo_id = int(current_wo) if current_wo else None
    except ValueError:
        current_wo_id = None
    return jsonify(get_wo_no_awb_by_asp(customer, current_wo_id=current_wo_id))


@asp_bp.route("/asp/api/return-part-same-asp", methods=["GET"])
@login_required
def api_return_part_same_asp():
    """Return closed WOs for a given ASP that have pending return_status (PENDING WITH PARTNER or PENDING FOR DC GENERATION)."""
    from app.services.database.queries import get_return_part_wos_by_asp
    customer = request.args.get("customer", "").strip()
    if not customer:
        return jsonify([])
    return jsonify(get_return_part_wos_by_asp(customer))


@asp_bp.route("/asp/api/wo-detail/<int:work_order_id>", methods=["GET"])
@login_required
def api_wo_detail(work_order_id: int):
    """Single WO full detail — wo_summary + wo_details joined."""
    from app.services.database.queries import get_wo_detail
    row = get_wo_detail(work_order_id)
    if not row:
        return jsonify({"error": "Not found"}), 404
    # asp_user: only their own assigned WOs
    tf = _tech_id_filter()
    if tf and row.get("tech_id") != tf:
        return jsonify({"error": "Not found"}), 404
    # asp / asp_master: vendor scope check
    vf = _vendor_filter()
    if vf and row.get("labor_vendor_related") != vf:
        return jsonify({"error": "Not found"}), 404
    return jsonify(row)


@asp_bp.route("/asp/api/wo-parts/<int:work_order_id>", methods=["GET"])
@login_required
def api_wo_parts(work_order_id: int):
    """All part-order lines for one WO from wo_product_detail."""
    from app.services.database.queries import get_parts_for_wo, get_wo_detail
    tf = _tech_id_filter()
    vf = _vendor_filter()
    if tf or vf:
        row = get_wo_detail(work_order_id)
        if tf and (not row or row.get("tech_id") != tf):
            return jsonify([])
        if vf and (not row or row.get("labor_vendor_related") != vf):
            return jsonify([])
    rows = get_parts_for_wo(work_order_id)
    return jsonify(rows)


@asp_bp.route("/asp/api/wo-related-serial/<int:work_order_id>", methods=["GET"])
@login_required
def api_wo_related_serial(work_order_id: int):
    """Return all WOs (including the current one) that share the same serial_number."""
    from app.services.database.queries import get_wo_detail, get_wo_by_serial
    detail = get_wo_detail(work_order_id)
    tf = _tech_id_filter()
    vf = _vendor_filter()
    # Only gate access to the current WO; history rows are unfiltered context.
    if tf and (not detail or detail.get("tech_id") != tf):
        return jsonify({"serial_number": None, "current_wo_id": work_order_id, "rows": []})
    if vf and (not detail or detail.get("labor_vendor_related") != vf):
        return jsonify({"serial_number": None, "current_wo_id": work_order_id, "rows": []})
    if not detail or not detail.get("serial_number"):
        return jsonify({"serial_number": None, "current_wo_id": work_order_id, "rows": []})
    rows = get_wo_by_serial(detail["serial_number"])
    return jsonify({"serial_number": detail["serial_number"], "current_wo_id": work_order_id, "rows": rows})


@asp_bp.route("/asp/api/wo-ticket-history/<int:work_order_id>", methods=["GET"])
@login_required
def api_wo_ticket_history(work_order_id: int):
    """Return all WOs (including the current one) that share the same case_number (ticket)."""
    from app.services.database.queries import get_wo_detail, get_wo_by_case_number
    detail = get_wo_detail(work_order_id)
    tf = _tech_id_filter()
    vf = _vendor_filter()
    # Only gate access to the current WO; history rows are unfiltered context.
    if tf and (not detail or detail.get("tech_id") != tf):
        return jsonify({"case_number": None, "current_wo_id": work_order_id, "rows": []})
    if vf and (not detail or detail.get("labor_vendor_related") != vf):
        return jsonify({"case_number": None, "current_wo_id": work_order_id, "rows": []})
    if not detail or not detail.get("case_number"):
        return jsonify({"case_number": None, "current_wo_id": work_order_id, "rows": []})
    rows = get_wo_by_case_number(detail["case_number"])
    return jsonify({"case_number": detail["case_number"], "current_wo_id": work_order_id, "rows": rows})


@asp_bp.route("/asp/api/working-hours", methods=["POST"])
@login_required
def api_save_working_hours():
    """Save updated working_hours string for the logged-in ASP user."""
    from app.services.database.db import get_db
    uid  = session.get("user_id")
    role = session.get("role", "")
    if role != "asp":
        return jsonify({"error": "Forbidden"}), 403
    data = request.get_json(silent=True) or {}
    working_hours = (data.get("working_hours") or "").strip()
    if not working_hours:
        return jsonify({"error": "working_hours is required"}), 400
    get_db().execute(
        "UPDATE asp_details SET working_hours = ? WHERE id = ?",
        (working_hours, uid)
    )
    get_db().commit()
    return jsonify({"ok": True, "working_hours": working_hours})


@asp_bp.route("/asp/api/operation-support", methods=["POST"])
@login_required
def api_save_operation_support():
    """Save updated operation_support value for the logged-in ASP user."""
    from app.services.database.db import get_db
    uid  = session.get("user_id")
    role = session.get("role", "")
    if role != "asp":
        return jsonify({"error": "Forbidden"}), 403
    data = request.get_json(silent=True) or {}
    value = (data.get("operation_support") or "").strip()
    allowed = {"CCI Only", "CCI & ONSITE"}
    if value not in allowed:
        return jsonify({"error": "Invalid value. Must be 'CCI Only' or 'CCI & ONSITE'."}), 400
    get_db().execute(
        "UPDATE asp_details SET operation_support = ? WHERE id = ?",
        (value, uid)
    )
    get_db().commit()
    return jsonify({"ok": True, "operation_support": value})


@asp_bp.route("/asp/api/request-password-change", methods=["POST"])
@login_required
def api_request_password_change():
    """Submit and auto-approve a password change for the logged-in ASP account."""
    from app.services.database.db import get_db
    uid      = session.get("user_id")
    username = session.get("username")
    role     = session.get("role", "")
    if role != "asp":
        return jsonify({"error": "Forbidden"}), 403
    data         = request.get_json(silent=True) or {}
    new_password = (data.get("new_password") or "").strip()
    if not new_password or len(new_password) < 8:
        return jsonify({"error": "New password must be at least 8 characters."}), 400
    db = get_db()
    # Cancel any existing pending request first
    db.execute(
        "UPDATE asp_pw_change_requests SET status='cancelled' "
        "WHERE asp_username=? AND status='pending'",
        (username,)
    )
    # Record the request and immediately mark it as approved (auto)
    db.execute(
        "INSERT INTO asp_pw_change_requests "
        "(asp_username, new_password, status, reviewed_by, reviewed_at) "
        "VALUES (?, ?, 'approved', 'auto', datetime('now'))",
        (username, new_password)
    )
    # Apply the new password directly
    db.execute(
        "UPDATE asp_details SET password=? WHERE username=?",
        (new_password, username)
    )
    db.commit()
    return jsonify({"ok": True, "message": "Password changed successfully."})


@asp_bp.route("/asp/api/location", methods=["PATCH"])
@login_required
def api_save_location():
    """Save updated location fields for the logged-in ASP account."""
    from app.services.database.db import get_db
    uid  = session.get("user_id")
    role = session.get("role", "")
    if role != "asp":
        return jsonify({"error": "Forbidden"}), 403
    data       = request.get_json(silent=True) or {}
    store_name = (data.get("store_name")    or "").strip() or None
    kota       = (data.get("kota")          or "").strip() or None
    island     = (data.get("island")        or "").strip() or None
    phone      = (data.get("phone_number")  or "").strip() or None
    address    = (data.get("address")       or "").strip() or None
    db = get_db()
    db.execute(
        "UPDATE asp_details SET store_name=?, kota=?, island=?, "
        "phone_number=?, address=?, updated_at=datetime('now') WHERE id=?",
        (store_name, kota, island, phone, address, uid)
    )
    db.commit()
    row = db.execute(
        "SELECT store_name, kota, island, phone_number, address "
        "FROM asp_details WHERE id=?", (uid,)
    ).fetchone()
    return jsonify({"ok": True, "location": dict(row)})


# ── ASP Users API ─────────────────────────────────────────────────────────────

def _asp_users_forbidden():
    """Only the parent ASP account may manage (create/edit) asp_users."""
    if session.get("role") != "asp":
        return jsonify({"error": "Forbidden"}), 403
    return None


@asp_bp.route("/asp/api/users", methods=["GET"])
@login_required
def api_list_asp_users():
    """Return all users belonging to the logged-in ASP."""
    err = _asp_users_forbidden()
    if err: return err
    from app.services.database.db import get_db
    labor_vendor = session.get("labor_vendor")
    rows = get_db().execute(
        "SELECT id, tech_id, full_name, email, password, phone_number, is_active, created_at "
        "FROM asp_users WHERE labor_vendor_related = ? ORDER BY id",
        (labor_vendor,)
    ).fetchall()
    return jsonify({"ok": True, "users": [dict(r) for r in rows]})


@asp_bp.route("/asp/api/users", methods=["POST"])
@login_required
def api_create_asp_user():
    """Create a new user under the logged-in ASP."""
    err = _asp_users_forbidden()
    if err: return err
    from app.services.database.db import get_db
    labor_vendor = session.get("labor_vendor")
    data      = request.get_json(silent=True) or {}
    full_name = (data.get("full_name")    or "").strip()
    email     = (data.get("email")        or "").strip()
    password  = (data.get("password")     or "").strip()
    phone     = (data.get("phone_number") or "").strip() or None
    tech_id   = (data.get("tech_id")      or "").strip() or None
    if not full_name:
        return jsonify({"error": "full_name is required"}), 400
    if not email:
        return jsonify({"error": "email is required"}), 400
    if not password or len(password) < 8:
        return jsonify({"error": "password must be at least 8 characters"}), 400
    db = get_db()
    # Uniqueness checks across the entire asp_users table
    dup_email = db.execute(
        "SELECT id FROM asp_users WHERE LOWER(email) = LOWER(?)",
        (email,)
    ).fetchone()
    if dup_email:
        return jsonify({"error": "That email address is already registered.", "field": "email"}), 409
    dup_name = db.execute(
        "SELECT id FROM asp_users WHERE LOWER(full_name) = LOWER(?)",
        (full_name,)
    ).fetchone()
    if dup_name:
        return jsonify({"error": "That full name is already registered.", "field": "full_name"}), 409
    cur = db.execute(
        "INSERT INTO asp_users (labor_vendor_related, tech_id, full_name, email, password, phone_number) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (labor_vendor, tech_id, full_name, email, password, phone)
    )
    db.commit()
    new_id = cur.lastrowid
    row = db.execute(
        "SELECT id, tech_id, full_name, email, password, phone_number, is_active, created_at "
        "FROM asp_users WHERE id = ?", (new_id,)
    ).fetchone()
    return jsonify({"ok": True, "user": dict(row)}), 201


@asp_bp.route("/asp/api/users/<int:user_id>", methods=["PUT"])
@login_required
def api_update_asp_user(user_id):
    """Update an existing ASP user (must belong to the logged-in ASP)."""
    err = _asp_users_forbidden()
    if err: return err
    from app.services.database.db import get_db
    labor_vendor = session.get("labor_vendor")
    db   = get_db()
    # Verify ownership
    existing = db.execute(
        "SELECT id FROM asp_users WHERE id = ? AND labor_vendor_related = ?",
        (user_id, labor_vendor)
    ).fetchone()
    if not existing:
        return jsonify({"error": "User not found"}), 404
    data      = request.get_json(silent=True) or {}
    # Fetch current values so partial updates (e.g. contact-only) don't wipe other fields
    current = db.execute(
        "SELECT full_name, email, password, phone_number FROM asp_users WHERE id = ?",
        (user_id,)
    ).fetchone()
    full_name = (data.get("full_name") or "").strip() or (current["full_name"] or "")
    email     = (data.get("email")     or "").strip() or (current["email"]     or "")
    password  = (data.get("password")  or "").strip()
    # phone_number key present → use it (even if empty string → None); key absent → keep current
    if "phone_number" in data:
        phone = (data.get("phone_number") or "").strip() or None
    else:
        phone = current["phone_number"]
    if not full_name:
        return jsonify({"error": "full_name is required"}), 400
    if not email:
        return jsonify({"error": "email is required"}), 400
    # Uniqueness checks across the entire asp_users table, excluding the current user
    dup_email = db.execute(
        "SELECT id FROM asp_users WHERE LOWER(email) = LOWER(?) AND id != ?",
        (email, user_id)
    ).fetchone()
    if dup_email:
        return jsonify({"error": "That email address is already registered.", "field": "email"}), 409
    dup_name = db.execute(
        "SELECT id FROM asp_users WHERE LOWER(full_name) = LOWER(?) AND id != ?",
        (full_name, user_id)
    ).fetchone()
    if dup_name:
        return jsonify({"error": "That full name is already registered.", "field": "full_name"}), 409
    # Only update password when a new one is supplied
    if password:
        if len(password) < 8:
            return jsonify({"error": "password must be at least 8 characters"}), 400
        db.execute(
            "UPDATE asp_users SET full_name=?, email=?, password=?, "
            "phone_number=?, updated_at=datetime('now') WHERE id=?",
            (full_name, email, password, phone, user_id)
        )
    else:
        db.execute(
            "UPDATE asp_users SET full_name=?, email=?, "
            "phone_number=?, updated_at=datetime('now') WHERE id=?",
            (full_name, email, phone, user_id)
        )
    db.commit()
    row = db.execute(
        "SELECT id, tech_id, full_name, email, password, phone_number, is_active, created_at "
        "FROM asp_users WHERE id = ?", (user_id,)
    ).fetchone()
    return jsonify({"ok": True, "user": dict(row)})


@asp_bp.route("/asp/api/users/<int:user_id>", methods=["DELETE"])
@login_required
def api_delete_asp_user(user_id):
    """Permanently delete an ASP user (must belong to the logged-in ASP)."""
    err = _asp_users_forbidden()
    if err: return err
    from app.services.database.db import get_db
    labor_vendor = session.get("labor_vendor")
    db = get_db()
    existing = db.execute(
        "SELECT id FROM asp_users WHERE id = ? AND labor_vendor_related = ?",
        (user_id, labor_vendor)
    ).fetchone()
    if not existing:
        return jsonify({"error": "User not found"}), 404
    db.execute("DELETE FROM asp_users WHERE id = ?", (user_id,))
    db.commit()
    return jsonify({"ok": True})


@asp_bp.route("/asp/api/users/<int:user_id>/status", methods=["PATCH"])
@login_required
def api_toggle_asp_user_status(user_id):
    """Toggle is_active for an ASP user (must belong to the logged-in ASP)."""
    err = _asp_users_forbidden()
    if err: return err
    from app.services.database.db import get_db
    labor_vendor = session.get("labor_vendor")
    db = get_db()
    existing = db.execute(
        "SELECT id, is_active FROM asp_users WHERE id = ? AND labor_vendor_related = ?",
        (user_id, labor_vendor)
    ).fetchone()
    if not existing:
        return jsonify({"error": "User not found"}), 404
    data       = request.get_json(silent=True) or {}
    new_status = 1 if data.get("is_active") else 0
    db.execute(
        "UPDATE asp_users SET is_active = ?, updated_at = datetime('now') WHERE id = ?",
        (new_status, user_id)
    )
    db.commit()
    row = db.execute(
        "SELECT id, tech_id, full_name, email, password, phone_number, is_active, created_at "
        "FROM asp_users WHERE id = ?", (user_id,)
    ).fetchone()
    return jsonify({"ok": True, "user": dict(row)})


@asp_bp.route("/asp/api/completed-last-30days", methods=["GET"])
@login_required
def api_completed_last_30days():
    """Completed Last 30 Days — WOs whose completion_date is within the past 30 days."""
    from app.services.database.queries import get_asp_completed_last_30_days
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_completed_last_30_days(
        search         = request.args.get("q", "").strip(),
        type_filter    = request.args.get("wo_type", "").strip(),
        no_awb         = _bool_arg("no_awb", False),
        page           = _int_arg("page", 1),
        page_size      = per_page,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/in-prepare", methods=["GET"])
@login_required
def api_in_prepare():
    """In-Prepare Follow-Up — WOs with part ordered but not yet shipped."""
    from app.services.database.queries import get_asp_in_prepare_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_in_prepare_page(
        search          = request.args.get("q", "").strip(),
        page            = _int_arg("page", 1),
        page_size       = per_page,
        vendor_filter   = _vendor_filter(),
        tech_id_filter  = _tech_id_filter(),
        prepare_filter  = request.args.get("prepare_filter", "").strip(),
    ))



@asp_bp.route("/asp/api/return-part", methods=["GET"])
@login_required
def api_return_part():
    """Return Part Follow-Up — WOs with is_exist_excel='yes' SOIDs; computed need_to_return / weekly_dc_report state."""
    from app.services.database.queries import get_asp_part_return_page
    per_page = min(_int_arg("per_page", 25), 100)
    return jsonify(get_asp_part_return_page(
        search         = request.args.get("q", "").strip(),
        followup_state = request.args.get("followup_state", "").strip(),
        page           = _int_arg("page", 1),
        page_size      = per_page,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    ))


@asp_bp.route("/asp/api/return-part/export", methods=["GET"])
@login_required
def api_return_part_export():
    """Export Return Part Follow-Up rows expanded by SOID (one row per part line)
    for every WO that exists on the Return Part tab (followup_state=need_to_return).
    Saves to files/report/ and streams the file back as a download."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.services.database.queries import get_asp_part_return_page
    from app.services.database.db import get_db

    # ── 1. Collect WOs in the Return Part tab ─────────────────────────────────
    result = get_asp_part_return_page(
        search         = request.args.get("q", "").strip(),
        followup_state = "need_to_return",
        page           = 1,
        page_size      = 9999,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    )
    wo_rows = result.get("rows", [])

    # Build a fast lookup: work_order_id → WO summary dict
    wo_map = {r["work_order_id"]: r for r in wo_rows}
    wo_ids = list(wo_map.keys())

    # ── 2. Fetch all SOID lines for those WOs ────────────────────────────────
    part_rows = []
    if wo_ids:
        conn = get_db()
        placeholders = ",".join("?" * len(wo_ids))
        part_rows = conn.execute(
            f"""
            SELECT
                p.work_order_id,
                p.soid,
                p.product,
                p.description,
                p.order_date,
                p.delivery_date,
                p.wo_product_status,
                p.return_status,
                p.awb,
                p.ship_pou_pod_time,
                p.dc_number
            FROM wo_product_detail p
            WHERE p.work_order_id IN ({placeholders})
              AND UPPER(TRIM(COALESCE(p.return_status,''))) IN (
                  'PENDING WITH PARTNER','PENDING FOR DC GENERATION'
              )
            ORDER BY p.work_order_id, p.soid
            """,
            wo_ids,
        ).fetchall()
        part_rows = [dict(r) for r in part_rows]

    # ── 3. Build export rows: one row per SOID ────────────────────────────────
    def _fmt_date(val):
        if not val:
            return ""
        s = str(val).strip()
        return s[:10] if len(s) >= 10 else s

    export_rows = []
    for p in part_rows:
        wo = wo_map.get(p["work_order_id"], {})
        export_rows.append({
            "work_order_id":                wo.get("work_order_id", ""),
            "created_on":                   _fmt_date(wo.get("created_on")),
            "work_order_type":              wo.get("work_order_type", ""),
            "case_desc":                    wo.get("case_desc", ""),
            "work_order_status":            wo.get("work_order_status", ""),
            "committed_delivery_date":      _fmt_date(wo.get("committed_delivery_date")),
            "actual_committed_onsite_date": _fmt_date(wo.get("actual_committed_onsite_date")),
            "contact_name":                 wo.get("contact_name", ""),
            "customer":                     wo.get("customer", ""),
            "soid":                         (str(int(p["soid"])) if p.get("soid") is not None else ""),
            "product":                      p.get("product", ""),
            "description":                  p.get("description", ""),
            "order_date":                   _fmt_date(p.get("order_date")),
            "delivery_date":                _fmt_date(p.get("delivery_date")),
            "wo_product_status":            p.get("wo_product_status", ""),
            "return_status":                p.get("return_status", ""),
            "awb":                          p.get("awb", ""),
            "ship_pou_pod_time":            _fmt_date(p.get("ship_pou_pod_time")),
            "dc_number":                    p.get("dc_number", ""),
        })

    # ── 4. Build workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Return Part - By SOID"

    headers = [
        "No.", "WO Number", "Created On", "WO Type", "Case",
        "WO Status", "Committed Delivery", "Actual Committed",
        "Contact Name", "ASP",
        "SOID", "Part Number", "Description", "Order Date", "Delivery Date",
        "Part Status", "Return Status", "AWB", "POD Date", "DC Number",
    ]
    col_keys = [
        None,
        "work_order_id", "created_on", "work_order_type", "case_desc",
        "work_order_status", "committed_delivery_date", "actual_committed_onsite_date",
        "contact_name", "customer",
        "soid", "product", "description", "order_date", "delivery_date",
        "wo_product_status", "return_status", "awb", "ship_pou_pod_time", "dc_number",
    ]
    col_widths = [6, 16, 16, 14, 32, 22, 20, 20, 22, 28, 14, 16, 30, 14, 14, 22, 12, 18, 14, 14]

    # Header style
    hdr_fill   = PatternFill("solid", fgColor="1F2328")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="E5E7EB")
    thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    # Row styles
    even_fill  = PatternFill("solid", fgColor="F7F8FA")
    data_font  = Font(size=11)
    data_align = Alignment(vertical="center")

    for ri, r in enumerate(export_rows, start=2):
        fill = even_fill if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(col_keys, start=1):
            value = (ri - 1) if key is None else (r.get(key) or "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border
            if key == "soid":
                cell.number_format = "@"
            if fill.fill_type:
                cell.fill = fill
        ws.row_dimensions[ri].height = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── 5. Save and stream ────────────────────────────────────────────────────
    report_dir = current_app.config["REPORT_DIR"]
    os.makedirs(report_dir, exist_ok=True)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename   = f"Return_Part_BySoid_{timestamp}.xlsx"
    filepath   = os.path.join(report_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ═══════════════════════════════════════════════════════════════════════════
#  In-Prepare Follow-Up — Export
# ═══════════════════════════════════════════════════════════════════════════

@asp_bp.route("/asp/api/in-prepare/export", methods=["GET"])
@login_required
def api_in_prepare_export():
    """Export In-Prepare Follow-Up — all rows, one per WO, as .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.services.database.queries import get_asp_in_prepare_page

    result = get_asp_in_prepare_page(
        search          = request.args.get("q", "").strip(),
        prepare_filter  = request.args.get("prepare_filter", "").strip(),
        page            = 1,
        page_size       = 9999,
        vendor_filter   = _vendor_filter(),
        tech_id_filter  = _tech_id_filter(),
    )
    wo_rows = result.get("rows", [])

    def _fd(val):
        if not val: return ""
        s = str(val).strip()
        return s[:16] if len(s) > 16 else s

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "In-Prepare Follow-Up"

    headers = [
        "No.", "WO Number", "Created On", "WO Type", "Case",
        "WO Status", "Part Product", "Part Description",
        "Part Order Date", "ETA tiba di YCH", "Part SOID",
        "On Hold", "Total Waiting Pickup",
        "Contact Name", "ASP",
    ]
    col_keys = [
        None,
        "work_order_id", "created_on", "work_order_type", "case_desc",
        "work_order_status", "part_product", "part_description",
        "part_order_date", "part_eta_wh", "part_soid",
        "part_on_hold_count", "part_waiting_pickup_count",
        "contact_name", "customer",
    ]
    col_widths = [6, 16, 18, 14, 34, 26, 20, 34, 18, 18, 14, 12, 18, 24, 32]

    hdr_fill   = PatternFill("solid", fgColor="1F2328")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="E5E7EB")
    thin_bdr   = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = thin_bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    even_fill  = PatternFill("solid", fgColor="F7F8FA")
    data_font  = Font(size=11)
    data_align = Alignment(vertical="center")

    for ri, r in enumerate(wo_rows, start=2):
        fill = even_fill if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(col_keys, start=1):
            if key is None:
                value = ri - 1
            elif key in ("created_on", "part_order_date", "part_eta_wh"):
                value = _fd(r.get(key))
            elif key == "part_soid":
                raw = r.get(key)
                value = str(int(raw)) if raw is not None else ""
            else:
                value = r.get(key) or ""
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = data_font; cell.alignment = data_align; cell.border = thin_bdr
            if fill.fill_type: cell.fill = fill
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"

    report_dir = current_app.config["REPORT_DIR"]
    os.makedirs(report_dir, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"InPrepare_FollowUp_{ts}.xlsx"
    filepath = os.path.join(report_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ═══════════════════════════════════════════════════════════════════════════
#  CCI Follow-Up — Export
# ═══════════════════════════════════════════════════════════════════════════

@asp_bp.route("/asp/api/cci-followup/export", methods=["GET"])
@login_required
def api_cci_followup_export():
    """Export CCI Follow-Up — all rows as .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.services.database.queries import get_asp_cci_followup_page

    result = get_asp_cci_followup_page(
        search         = request.args.get("q", "").strip(),
        followup_state = request.args.get("followup_state", "").strip(),
        page           = 1,
        page_size      = 9999,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    )
    wo_rows = result.get("rows", [])

    def _fd(val):
        if not val: return ""
        s = str(val).strip()
        return s[:16] if len(s) > 16 else s

    _state_labels = {
        "confirm_receipt": "Confirm AWB",
        "part_sla":        "Part SLA Overdue",
        "wo_sla":          "Escalate WO",
        "report_problem":  "WO SLA Follow-Up",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CCI Follow-Up"

    headers = [
        "No.", "WO Number", "Created On", "Follow-Up State",
        "Part Qty (In Transit)", "Target Fix/Sampai",
        "AWB", "WO Status", "Shipped On", "Case",
        "Contact Name", "ASP",
    ]
    col_keys = [
        None,
        "work_order_id", "created_on", "_followup_label",
        "part_qty", "part_eta",
        "part_awb", "work_order_status", "ship_pickup_time", "case_desc",
        "contact_name", "customer",
    ]
    col_widths = [6, 16, 18, 22, 18, 22, 20, 26, 20, 34, 24, 32]

    hdr_fill   = PatternFill("solid", fgColor="1F2328")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="E5E7EB")
    thin_bdr   = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = thin_bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    even_fill  = PatternFill("solid", fgColor="F7F8FA")
    data_font  = Font(size=11)
    data_align = Alignment(vertical="center")

    for ri, r in enumerate(wo_rows, start=2):
        fill = even_fill if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(col_keys, start=1):
            if key is None:
                value = ri - 1
            elif key == "_followup_label":
                value = _state_labels.get(r.get("followup_state", ""), r.get("followup_state", "") or "")
            elif key in ("created_on", "part_eta", "ship_pickup_time"):
                value = _fd(r.get(key))
            else:
                value = r.get(key) or ""
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = data_font; cell.alignment = data_align; cell.border = thin_bdr
            if fill.fill_type: cell.fill = fill
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"

    report_dir = current_app.config["REPORT_DIR"]
    os.makedirs(report_dir, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"CCI_FollowUp_{ts}.xlsx"
    filepath = os.path.join(report_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ═══════════════════════════════════════════════════════════════════════════
#  ONS (Onsite) Follow-Up — Export
# ═══════════════════════════════════════════════════════════════════════════

@asp_bp.route("/asp/api/onsite-followup/export", methods=["GET"])
@login_required
def api_onsite_followup_export():
    """Export ONS Follow-Up — all rows as .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.services.database.queries import get_asp_onsite_followup_page

    result = get_asp_onsite_followup_page(
        search         = request.args.get("q", "").strip(),
        followup_state = request.args.get("followup_state", "").strip(),
        page           = 1,
        page_size      = 9999,
        vendor_filter  = _vendor_filter(),
        tech_id_filter = _tech_id_filter(),
    )
    wo_rows = result.get("rows", [])

    def _fd(val):
        if not val: return ""
        s = str(val).strip()
        return s[:16] if len(s) > 16 else s

    _state_labels = {
        "wo_reschedule":  "ONS In-Transit",
        "part_sla":       "Part SLA Overdue",
        "wo_sla":         "Escalate WO",
        "report_problem": "WO SLA Follow-Up",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ONS Follow-Up"

    headers = [
        "No.", "WO Number", "Created On", "Follow-Up State",
        "Delivered On", "Part Qty (In Transit)", "Defer Date",
        "Target Fix/Sampai",
        "AWB", "WO Status", "Shipped On", "POD / Received On",
        "Case", "Contact Name", "ASP",
    ]
    col_keys = [
        None,
        "work_order_id", "created_on", "_followup_label",
        "part_pod_time", "part_qty", "customer_defer_date",
        "part_eta",
        "part_awb", "work_order_status", "ship_pickup_time", "part_pod_time",
        "case_desc", "contact_name", "customer",
    ]
    col_widths = [6, 16, 18, 20, 20, 18, 18, 22, 20, 26, 20, 20, 34, 24, 32]

    hdr_fill   = PatternFill("solid", fgColor="1F2328")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="E5E7EB")
    thin_bdr   = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = thin_bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    even_fill  = PatternFill("solid", fgColor="F7F8FA")
    data_font  = Font(size=11)
    data_align = Alignment(vertical="center")

    for ri, r in enumerate(wo_rows, start=2):
        fill = even_fill if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(col_keys, start=1):
            if key is None:
                value = ri - 1
            elif key == "_followup_label":
                value = _state_labels.get(r.get("followup_state", ""), r.get("followup_state", "") or "")
            elif key in ("created_on", "part_eta", "ship_pickup_time",
                         "part_pod_time", "customer_defer_date"):
                value = _fd(r.get(key))
            else:
                value = r.get(key) or ""
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = data_font; cell.alignment = data_align; cell.border = thin_bdr
            if fill.fill_type: cell.fill = fill
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"

    report_dir = current_app.config["REPORT_DIR"]
    os.makedirs(report_dir, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ONS_FollowUp_{ts}.xlsx"
    filepath = os.path.join(report_dir, filename)
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
