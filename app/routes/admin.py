import logging
import os
import re as _re
import sqlite3
import tempfile
import threading
import queue as _queue
import time as _time
import uuid as _uuid
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, current_app, send_file, jsonify,
)
from werkzeug.utils import secure_filename
from app.services.database.db import get_db, open_db
from app.services.upload.excel import allowed_excel, save_excel_upload, list_excel_uploads
from app.services.upload.upload_verification import verify_uploaded_file
from app.services.upload.meta_cache import (
    write_meta, read_meta, delete_meta, mark_upserted,
    read_active_open_wos, rebuild_active_open_wos,
    read_incomplete_prev_shipments, rebuild_incomplete_prev_shipments,
    write_wo_product_mismatch, read_wo_product_mismatch,
)
from app.services.upload.excel_to_df import (
    load_single_dataframe, DF_LABELS, _KEY_TO_DF,
)
from app.config.file_categories import FILE_CATEGORY_CONFIGS
from app.services.database.queries import get_wo_summary_stats
from app.services.database.upsert import dispatch_upsert

admin_bp = Blueprint("admin", __name__)

# Reverse map: file_category display string → category_key (e.g. "WOID")
_FILE_CATEGORY_TO_KEY = {
    cfg["file_category"]: key
    for key, cfg in FILE_CATEGORY_CONFIGS.items()
}.get


# ── Auth guard: all admin routes require admin role ───────────────────────────

@admin_bp.before_request
def _require_admin():
    from flask import session as _session
    if "user_id" not in _session:
        return redirect(url_for("auth.login", next=request.path))
    if _session.get("role") not in ("admin", "superadmin"):
        return redirect(url_for("auth.login"))


# ── Dashboard ────────────────────────────────────────────────────────────────

@admin_bp.route("/admin/dashboard", methods=["GET"])
def dashboard():
    stats = get_wo_summary_stats()
    return render_template("admin/dashboard.html",
                           portal="admin", active_page="admin_dashboard",
                           **stats)


# ── API: WO Summary (server-side pagination + search) ────────────────────────

@admin_bp.route("/admin/api/wo-summary", methods=["GET"])
def api_wo_summary():
    """
    JSON endpoint for the paginated WO Summary table.

    Query params:
        q        - free-text search (WO ID, serial, contact, customer, case)
        status   - status filter keyword
        wo_type  - work_order_type filter
        page     - 1-based page number (default 1)
        per_page - rows per page (default 25, max 100)
    """
    from app.services.database.queries import get_wo_summary_page
    per_page = min(int(request.args.get("per_page", 25)), 100)
    result   = get_wo_summary_page(
        search              = request.args.get("q", "").strip(),
        status_filter       = request.args.get("status", "").strip(),
        type_filter         = request.args.get("wo_type", "").strip(),
        case_status_filter  = request.args.get("case_status", "").strip(),
        page                = max(1, int(request.args.get("page", 1))),
        page_size           = per_page,
    )
    return jsonify(result)


# ── API: Dashboard overview stats ────────────────────────────────────────────

@admin_bp.route("/admin/api/dashboard-stats", methods=["GET"])
def api_dashboard_stats():
    """
    JSON endpoint returning lightweight stats for the admin dashboard overview:
      - open_esc_by_date : [{ day, count }] open escalations grouped by created date (all time),
                           where status is NOT Reject / Approved to Order / Complete
      - esc_open_total   : total open escalation count (same filter)
    """
    from app.services.database.db import get_db
    conn = get_db()

    _OPEN_FILTER = """
        LOWER(COALESCE(status, '')) NOT IN (
            'reject', 'approved to order', 'complete'
        )
    """

    # Total count per day (for bar height)
    open_by_date = conn.execute(f"""
        SELECT DATE(item_created_at) AS day, COUNT(*) AS count
        FROM technical_escalation
        WHERE {_OPEN_FILTER}
        GROUP BY day
        ORDER BY day ASC
    """).fetchall()

    # Per-status count per day (for colour segments and hover detail)
    open_by_date_status = conn.execute(f"""
        SELECT DATE(item_created_at) AS day,
               COALESCE(status, '—') AS status,
               COUNT(*) AS count
        FROM technical_escalation
        WHERE {_OPEN_FILTER}
        GROUP BY day, status
        ORDER BY day ASC, count DESC
    """).fetchall()

    esc_open_total = conn.execute(
        f"SELECT COUNT(*) FROM technical_escalation WHERE {_OPEN_FILTER}"
    ).fetchone()[0]

    return jsonify({
        "open_esc_by_date":        [dict(r) for r in open_by_date],
        "open_esc_by_date_status": [dict(r) for r in open_by_date_status],
        "esc_open_total":          esc_open_total,
    })


# ── API: WO Detail (single WO, on-demand) ────────────────────────────────────

@admin_bp.route("/admin/api/wo-detail/<int:work_order_id>", methods=["GET"])
def api_wo_detail(work_order_id: int):
    """
    JSON endpoint returning the full detail row for one WO
    (wo_summary + wo_details joined).
    """
    from app.services.database.queries import get_wo_detail
    row = get_wo_detail(work_order_id)
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify(row)


# ── API: Part Order Details for a single WO ──────────────────────────────────

@admin_bp.route("/admin/api/wo-parts/<int:work_order_id>", methods=["GET"])
def api_wo_parts(work_order_id: int):
    """Return all part-order lines for one WO from wo_product_detail."""
    from app.services.database.queries import get_parts_for_wo
    rows = get_parts_for_wo(work_order_id)
    return jsonify(rows)


# ── API: Related WOs by serial number ────────────────────────────────────────

@admin_bp.route("/admin/api/wo-related-serial/<int:work_order_id>", methods=["GET"])
def api_wo_related_serial(work_order_id: int):
    """Return all WOs (including the current one) that share the same serial_number."""
    from app.services.database.queries import get_wo_detail, get_wo_by_serial
    detail = get_wo_detail(work_order_id)
    if not detail or not detail.get("serial_number"):
        return jsonify({"serial_number": None, "current_wo_id": work_order_id, "rows": []})
    rows = get_wo_by_serial(detail["serial_number"])
    return jsonify({"serial_number": detail["serial_number"], "current_wo_id": work_order_id, "rows": rows})


# ── API: All WOs by serial number string (admin monday-data view) ─────────────

@admin_bp.route("/admin/api/sn-history/<path:serial_number>", methods=["GET"])
def api_sn_history(serial_number: str):
    """Return all WOs in wo_summary/wo_details that share the given serial_number."""
    from app.services.database.queries import get_wo_by_serial
    rows = get_wo_by_serial(serial_number.strip())
    return jsonify({"serial_number": serial_number.strip(), "rows": rows})


# ── API: Ticket history by case number ───────────────────────────────────────

@admin_bp.route("/admin/api/wo-ticket-history/<int:work_order_id>", methods=["GET"])
def api_wo_ticket_history(work_order_id: int):
    """Return all WOs (including the current one) that share the same case_number (ticket)."""
    from app.services.database.queries import get_wo_detail, get_wo_by_case_number
    detail = get_wo_detail(work_order_id)
    if not detail or not detail.get("case_number"):
        return jsonify({"case_number": None, "current_wo_id": work_order_id, "rows": []})
    rows = get_wo_by_case_number(detail["case_number"])
    return jsonify({"case_number": detail["case_number"], "current_wo_id": work_order_id, "rows": rows})


# ── API: Monday escalation records by serial number ──────────────────────────

@admin_bp.route("/admin/api/wo-monday-escalation/<int:work_order_id>", methods=["GET"])
def api_wo_monday_escalation(work_order_id: int):
    """Return all Monday technical_escalation rows that share the same serial_number as the WO."""
    import os as _os
    from app.services.database.queries import get_wo_detail

    detail = get_wo_detail(work_order_id)
    if not detail or not detail.get("serial_number"):
        return jsonify({"serial_number": None, "rows": []})

    sn = detail["serial_number"].strip()

    project_root = _os.path.normpath(_os.path.join(_os.path.dirname(__file__), "..", ".."))
    db_path = _os.path.join(project_root, "files", "lenovo_asp.db")
    if not _os.path.isfile(db_path):
        return jsonify({"serial_number": sn, "rows": []})

    edb = open_db(db_path)
    rows = []
    try:
        raw = edb.execute(
            """
            SELECT
                te.monday_item_id,
                te.board_id,
                te.asp_board,
                te.item_name,
                te.item_created_at,
                te.item_updated_at,
                te.status,
                te.work_order_type,
                te.wo_case_id,
                te.serial_number,
                te.ppsn_category,
                te.rrr_category,
                te.diag_datetime,
                te.diag_agent_ce,
                te.diag_model,
                te.diag_warranty,
                te.diag_problem,
                te.diag_esc_approval,
                te.diag_parts_request,
                te.diagnose_note,
                te.repair_note,
                (
                    SELECT COUNT(DISTINCT u2.update_id) + COUNT(DISTINCT r2.reply_id)
                    FROM item_updates u2
                    LEFT JOIN item_update_replies r2 ON u2.update_id = r2.update_id
                    WHERE u2.monday_item_id = te.monday_item_id
                ) AS disc_count,
                (
                    SELECT wd.case_number
                    FROM wo_details wd
                    WHERE CAST(wd.work_order_id AS TEXT) = TRIM(te.wo_case_id)
                    LIMIT 1
                ) AS case_number
            FROM technical_escalation te
            WHERE LOWER(TRIM(te.serial_number)) = LOWER(?)
            ORDER BY te.item_created_at ASC
            """,
            (sn,),
        ).fetchall()
        rows = [dict(r) for r in raw]
    except Exception as _e:
        current_app.logger.error("admin api_wo_monday_escalation query failed: %s", _e)
    finally:
        edb.close()

    return jsonify({"serial_number": sn, "rows": rows})


# ── Ticket Management ────────────────────────────────────────────────────────

@admin_bp.route("/admin/tickets", methods=["GET"])
def tickets():
    return render_template("admin/ticket_management.html",
                           portal="admin", active_page="admin_tickets")


# ── Data Import / Export ─────────────────────────────────────────────────────

@admin_bp.route("/admin/data-import", methods=["GET"])
def data_import():
    import traceback as _tb
    files = list_excel_uploads()
    upload_folder = current_app.config["EXCEL_UPLOAD_FOLDER"]
    meta_folder   = current_app.config["UPLOAD_META_FOLDER"]
    for f in files:
        f["modified_fmt"] = datetime.fromtimestamp(f["modified"]).strftime("%Y-%m-%d %H:%M")
        # Read metadata from the cheap JSON sidecar written at upload time.
        # Fall back to live verification only when the sidecar is missing
        # (e.g. files uploaded before this change was deployed).
        meta = read_meta(meta_folder, f["name"])
        if meta is None:
            try:
                filepath = os.path.join(upload_folder, f["name"])
                meta = verify_uploaded_file(filepath)
                # Backfill the sidecar so the next load is fast
                write_meta(meta_folder, f["name"], meta)
            except Exception:
                current_app.logger.warning(
                    "verify_uploaded_file failed for %s:\n%s", f["name"], _tb.format_exc()
                )
                meta = {}
        f["file_category"]     = meta.get("file_category") or ""
        f["source_file"]       = meta.get("source_file") or ""
        f["latest_date"]       = meta.get("latest_date") or ""
        f["days_range"]        = meta.get("days_range") or ""
        f["validation_status"] = meta.get("validation_status") or ""
        f["upserted"]          = bool(meta.get("upserted"))

    # Build a lookup: file_category → uploaded file dict (last one wins per category)
    uploaded_by_category = {}
    for f in files:
        if f["file_category"]:
            uploaded_by_category[f["file_category"]] = f

    # One row per known category — merged with uploaded file data if present
    category_rows = []
    for key, cfg in FILE_CATEGORY_CONFIGS.items():
        cat_name = cfg["file_category"]
        uploaded = uploaded_by_category.get(cat_name)
        category_rows.append({
            "category_key":  key,
            "file_category": cat_name,
            "source_file":   cfg["source_file"],
            # from uploaded file (empty strings when nothing uploaded yet)
            "filename":      uploaded["name"]                          if uploaded else "",
            "size_kb":       uploaded["size_kb"]                       if uploaded else "",
            "modified_fmt":  uploaded["modified_fmt"]                  if uploaded else "",
            "latest_date":   uploaded["latest_date"]                   if uploaded else "",
            "days_range":    uploaded["days_range"]                    if uploaded else "",
            "upserted":      bool(uploaded.get("upserted"))            if uploaded else False,
        })

    # ── Incomplete previous-month shipments — read from JSON cache ────────────
    incomplete_prev_shipments = read_incomplete_prev_shipments(meta_folder)

    # ── Active open WOs — read from JSON cache (no DB hit on GET) ────────────
    # The cache is rebuilt after every WOID upsert.  On first visit (before any
    # upsert has ever run) the cache file does not exist yet, so we fall back to
    # a one-time bootstrap that also filters by any currently uploaded WOID file.
    active_open_wos = read_active_open_wos(meta_folder)
    if not active_open_wos and not os.path.isfile(
        os.path.join(meta_folder, "active_open_wos.json")
    ):
        # Cache has never been written — bootstrap it now, honouring the
        # currently uploaded WOID file's WO IDs as the exclusion set.
        try:
            import io as _io_b, pandas as _pd_b
            from app.services.upload.upload_verification import (
                verify_uploaded_file as _vuf_b,
            )
            _woid_cfg = FILE_CATEGORY_CONFIGS.get("WOID", {})
            _woid_cat = _woid_cfg.get("file_category", "")
            _bootstrap_ids: set[int] = set()
            for _fname in os.listdir(upload_folder):
                from app.services.upload.excel import allowed_excel as _ae_b
                if not _ae_b(_fname):
                    continue
                _m = read_meta(meta_folder, _fname)
                if _m and _m.get("file_category") == _woid_cat:
                    try:
                        _fp = os.path.join(upload_folder, _fname)
                        _vr_b = _vuf_b(_fp)
                        _sn_b = _vr_b.get("sheet_name", "")
                        with open(_fp, "rb") as _fh_b:
                            _fb_b = _io_b.BytesIO(_fh_b.read())
                        _df_b = (
                            _pd_b.read_excel(_fb_b, sheet_name=_sn_b)
                            if _sn_b else _pd_b.read_excel(_fb_b)
                        )
                        _bootstrap_ids = {
                            int(float(str(v)))
                            for v in _df_b.get(
                                "Work Order ID", _pd_b.Series(dtype=object)
                            )
                            if v is not None
                            and str(v).strip() not in ("", "nan", "NaN")
                        }
                    except Exception:
                        pass
                    break  # only one WOID file
            active_open_wos = rebuild_active_open_wos(
                meta_folder,
                current_app.config["DATABASE_PATH"],
                _bootstrap_ids or None,
            )
        except Exception:
            current_app.logger.warning(
                "active_open_wos bootstrap failed:\n" + _tb.format_exc()
            )

    # ── WO-product mismatch — read from JSON cache ────────────────────────────
    wo_product_mismatch = read_wo_product_mismatch(meta_folder)

    return render_template("admin/data_import.html",
                           files=files,
                           category_rows=category_rows,
                           active_open_wos=active_open_wos,
                           incomplete_prev_shipments=incomplete_prev_shipments,
                           wo_product_mismatch=wo_product_mismatch,
                           portal="admin", active_page="data_import",
                           active_group="data_import_export")


@admin_bp.route("/admin/msd-wo-updates", methods=["GET"])
def msd_wo_updates():
    project_root = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    )
    script_path = os.path.join(
        project_root,
        "app",
        "scripts",
        "msd-auto-download",
        "msd-auto-download.py",
    )
    downloads_dir = os.path.join(project_root, "files", "msd-auto-download")

    files = []
    if os.path.isdir(downloads_dir):
        for name in sorted(
            os.listdir(downloads_dir),
            key=lambda item: os.path.getmtime(os.path.join(downloads_dir, item)),
            reverse=True,
        ):
            file_path = os.path.join(downloads_dir, name)
            if not os.path.isfile(file_path):
                continue
            files.append({
                "name": name,
                "size_kb": round(os.path.getsize(file_path) / 1024, 1),
                "modified_fmt": datetime.fromtimestamp(
                    os.path.getmtime(file_path)
                ).strftime("%Y-%m-%d %H:%M"),
            })

    with _msd_lock:
        is_running = _msd_thread is not None and _msd_thread.is_alive()

    return render_template(
        "admin/export-import/msd_wo_updates.html",
        portal="admin",
        active_page="msd_wo_updates",
        active_group="data_import_export",
        script_path=script_path,
        downloads_dir=downloads_dir,
        files=files,
        is_running=is_running,
        in_window=_msd_in_active_window(),
        boot_id=_MSD_BOOT_ID,
    )


# A unique ID stamped at process start — used by the frontend to detect server
# restarts and automatically flush stale sessionStorage log content.
_MSD_BOOT_ID: str = _uuid.uuid4().hex

_msd_log_queue: _queue.Queue = _queue.Queue(maxsize=2000)
# Ring buffer of the last 500 log records — replayed to new SSE clients so
# refreshing the page mid-run shows the full current-run log history.
import collections as _collections
_msd_log_history: _collections.deque = _collections.deque(maxlen=500)
_msd_thread: threading.Thread | None = None
_msd_lock = threading.Lock()

# Per-file upsert stats — keyed by filename, set by _msd_auto_upsert after each run.
# { filename: {"new_wo": int, "updated_wo": int, "new_users": int,
#              "status": "success"|"failed", "upsert_date": "YYYY-MM-DD"} }
_msd_upsert_stats: dict = {}

# Daily cumulative totals — resets automatically when the calendar date changes.
# { "YYYY-MM-DD": {"new_wo": int, "updated_wo": int, "new_users": int} }
_msd_daily_totals: dict = {}

# Path to the JSON sidecar that persists both dicts across server restarts.
# Stored alongside the other upload-meta JSON files for a consistent structure.
_MSD_STATS_FILE = os.path.join(
    os.path.normpath(os.path.join(os.path.dirname(__file__), "..")),
    "templates", "admin", "upload_meta", "_upsert_stats.json",
)


def _msd_stats_load() -> None:
    """Read persisted stats from disk into the in-memory dicts (called once at startup)."""
    global _msd_upsert_stats, _msd_daily_totals
    import json as _json
    try:
        if os.path.isfile(_MSD_STATS_FILE):
            with open(_MSD_STATS_FILE, "r", encoding="utf-8") as _f:
                _data = _json.load(_f)
            _msd_upsert_stats = _data.get("upsert_stats",  {})
            _msd_daily_totals = _data.get("daily_totals",  {})
    except Exception:
        pass   # corrupt / missing — start fresh


def _msd_stats_save() -> None:
    """Write both in-memory dicts to disk atomically (temp file + rename)."""
    import json as _json, tempfile as _tmp
    try:
        os.makedirs(os.path.dirname(_MSD_STATS_FILE), exist_ok=True)
        _payload = _json.dumps(
            {"upsert_stats": _msd_upsert_stats, "daily_totals": _msd_daily_totals},
            indent=2,
        )
        _dir = os.path.dirname(_MSD_STATS_FILE)
        with _tmp.NamedTemporaryFile("w", dir=_dir, delete=False,
                                     suffix=".tmp", encoding="utf-8") as _tf:
            _tf.write(_payload)
            _tmp_path = _tf.name
        os.replace(_tmp_path, _MSD_STATS_FILE)
    except Exception:
        pass   # non-fatal — next write will retry


# Load persisted stats immediately so the first GET /files returns real data.
_msd_stats_load()

# OTP handshake — script blocks on _msd_otp_queue.get(); route puts the code in
_msd_otp_queue:    _queue.Queue = _queue.Queue(maxsize=1)
_msd_otp_pending:  bool = False   # True while script is waiting for OTP
_msd_relogin_pending: bool = False  # True while script is waiting for re-login

class _OtpCancelledError(Exception):
    """Raised inside the MSD script thread when the user clicks Cancel on the OTP panel."""

# ── Startup auto-run ─────────────────────────────────────────────────────────
# Delay is read from Config.MSD_STARTUP_DELAY_SEC (app/config/settings.py).
# Default is 10 minutes; reduce it there for testing without touching this file.
_MSD_STARTUP_DELAY_SEC: int = 10 * 60   # fallback used before app context is ready
_msd_startup_run_at: float = _time.time() + _MSD_STARTUP_DELAY_SEC  # Unix ts


def _msd_startup_scheduler(app) -> None:
    """Wait MSD_STARTUP_DELAY_SEC then auto-trigger the MSD download exactly once."""
    delay = app.config.get("MSD_STARTUP_DELAY_SEC", _MSD_STARTUP_DELAY_SEC)
    _time.sleep(delay)
    global _msd_thread, _msd_startup_run_at
    _msd_startup_run_at = 0.0   # clear the countdown
    with _msd_lock:
        already_running = _msd_thread is not None and _msd_thread.is_alive()
    if already_running:
        return  # user already started it manually — skip
    if not _msd_in_active_window():
        return  # outside office hours — don't auto-start
    with app.app_context():
        with _msd_lock:
            if _msd_thread is not None and _msd_thread.is_alive():
                return
            _msd_thread = threading.Thread(
                target=_run_msd_download_task,
                args=(app, False),
                daemon=True,
                name="msd-auto-download-startup",
            )
            _msd_thread.start()


def _msd_start_startup_scheduler(app) -> None:
    global _msd_startup_run_at
    delay = app.config.get("MSD_STARTUP_DELAY_SEC", _MSD_STARTUP_DELAY_SEC)
    _msd_startup_run_at = _time.time() + delay   # sync UI countdown to configured delay
    t = threading.Thread(
        target=_msd_startup_scheduler,
        args=(app,),
        daemon=True,
        name="msd-startup-scheduler",
    )
    t.start()


def _msd_in_active_window() -> bool:
    """Mon–Fri 06:00–20:00 local time."""
    now = datetime.now()
    return now.weekday() < 5 and 6 <= now.hour < 20


_MSD_JAKARTA_TZ = ZoneInfo("Asia/Jakarta")


def _msd_has_file_today() -> bool:
    """Return whether the MSD output folder contains a file from today in Jakarta."""
    project_root = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    )
    downloads_dir = os.path.join(project_root, "files", "msd-auto-download")
    today = datetime.now(_MSD_JAKARTA_TZ).date()
    if not os.path.isdir(downloads_dir):
        return False
    return any(
        os.path.isfile(os.path.join(downloads_dir, name))
        and datetime.fromtimestamp(
            os.path.getmtime(os.path.join(downloads_dir, name)), _MSD_JAKARTA_TZ
        ).date() == today
        for name in os.listdir(downloads_dir)
    )


def _msd_930_watchdog(app) -> None:
    """At 09:30 Jakarta on weekdays, restart MSD only when today's file is absent."""
    global _msd_thread
    while True:
        now = datetime.now(_MSD_JAKARTA_TZ)
        run_at = now.replace(hour=9, minute=30, second=0, microsecond=0)
        if now >= run_at:
            run_at += timedelta(days=1)
        while run_at.weekday() >= 5:
            run_at += timedelta(days=1)
        _time.sleep((run_at - now).total_seconds())

        with app.app_context():
            if _msd_has_file_today():
                _msd_log("INFO", "09:30 Jakarta watchdog: today's MSD file exists; reset skipped.")
                continue
            _msd_log("WARNING", "09:30 Jakarta watchdog: no MSD file found for today; resetting and restarting.")
            _msd_reset_task_state()
            with _msd_lock:
                if _msd_thread is not None and _msd_thread.is_alive():
                    continue
                _msd_thread = threading.Thread(
                    target=_run_msd_download_task,
                    args=(app, False),
                    daemon=True,
                    name="msd-auto-download-0930-watchdog",
                )
                _msd_thread.start()


def _msd_start_930_watchdog(app) -> None:
    threading.Thread(
        target=_msd_930_watchdog,
        args=(app,),
        daemon=True,
        name="msd-0930-watchdog",
    ).start()


# Matches Werkzeug/Flask HTTP access log lines that leak via captured stderr.
# Pattern: '127.0.0.1 - - [DD/Mon/YYYY ...' or the formatted variant with timestamp prefix.
_WERKZEUG_LINE_RE = _re.compile(
    r'(?:^|\s)(?:\d{1,3}\.){3}\d{1,3}\s+-\s+-\s+\[',
)


class _MsdQueueHandler(logging.Handler):
    def emit(self, record: logging.LogRecord) -> None:  # type: ignore[override]
        # Only forward records that originated from the msd_auto_download logger
        if not record.name.startswith("msd_auto_download"):
            return
        # Drop Werkzeug HTTP access log lines that leak via captured stderr
        msg = record.getMessage()
        if _WERKZEUG_LINE_RE.search(msg):
            return
        rec = {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "level": record.levelname,
            "msg": msg,
        }
        # Always append to the history ring buffer so new SSE clients can replay
        _msd_log_history.append(rec)
        try:
            _msd_log_queue.put_nowait(rec)
        except _queue.Full:
            try:
                _msd_log_queue.get_nowait()
                _msd_log_queue.put_nowait(rec)
            except (_queue.Full, _queue.Empty):
                pass


_msd_queue_handler = _MsdQueueHandler()
_msd_queue_handler.setLevel(logging.INFO)


def _msd_log(level: str, message: str) -> None:
    """Publish a backend MSD message to the live SSE log."""
    rec = {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "level": level,
        "msg": message,
    }
    _msd_log_history.append(rec)
    try:
        _msd_log_queue.put_nowait(rec)
    except _queue.Full:
        try:
            _msd_log_queue.get_nowait()
            _msd_log_queue.put_nowait(rec)
        except (_queue.Full, _queue.Empty):
            pass


@admin_bp.route("/admin/msd-wo-updates/trigger", methods=["POST"])
def msd_wo_updates_trigger():
    global _msd_thread

    # Off-hours: allow trigger but run only once, then wait for office hours.
    # In-window: normal repeating loop.
    run_once = not _msd_in_active_window()

    with _msd_lock:
        if _msd_thread is not None and _msd_thread.is_alive():
            return jsonify({"ok": False, "error": "MSD auto-download is already running."}), 409

        app = current_app._get_current_object()
        _msd_thread = threading.Thread(
            target=_run_msd_download_task,
            args=(app, run_once),
            daemon=True,
            name="msd-auto-download",
        )
        _msd_thread.start()

    return jsonify({"ok": True, "run_once": run_once})


@admin_bp.route("/admin/msd-wo-updates/stream", methods=["GET"])
def msd_wo_updates_stream():
    def generate():
        import json as _json
        # Replay recent history so a freshly connected client (page refresh,
        # server restart mid-run) sees what has already been logged.
        # The first record carries history=True so the frontend knows to
        # replace (not append to) its stored log with the replayed content.
        history = list(_msd_log_history)
        for i, rec in enumerate(history):
            payload = dict(rec, history=True, history_first=(i == 0))
            yield f"data: {_json.dumps(payload)}\n\n"
        while True:
            try:
                rec = _msd_log_queue.get(timeout=15)
                yield f"data: {_json.dumps(rec)}\n\n"
            except _queue.Empty:
                yield "data: {\"keepalive\": true}\n\n"

    return current_app.response_class(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@admin_bp.route("/admin/msd-wo-updates/files", methods=["GET"])
def msd_wo_updates_files():
    project_root = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    )
    downloads_dir = os.path.join(project_root, "files", "msd-auto-download")
    files = []
    if os.path.isdir(downloads_dir):
        for name in sorted(
            os.listdir(downloads_dir),
            key=lambda item: os.path.getmtime(os.path.join(downloads_dir, item)),
            reverse=True,
        ):
            file_path = os.path.join(downloads_dir, name)
            if not os.path.isfile(file_path):
                continue
            stats = _msd_upsert_stats.get(name, {})
            files.append({
                "name":          name,
                "size_kb":       round(os.path.getsize(file_path) / 1024, 1),
                "modified_fmt":  datetime.fromtimestamp(
                    os.path.getmtime(file_path)
                ).strftime("%Y-%m-%d %H:%M"),
                "new_wo":        stats.get("new_wo",       None),
                "updated_wo":    stats.get("updated_wo",   None),
                "new_users":     stats.get("new_users",    None),
                "upsert_status": stats.get("status",       None),
                "upsert_date":   stats.get("upsert_date",  None),
            })
    # Include today's daily totals so the frontend can show the running counter
    from datetime import date as _d
    _today_key = str(_d.today())
    daily = _msd_daily_totals.get(_today_key, {})
    return jsonify({
        "ok": True,
        "files": files,
        "daily_date":       _today_key,
        "daily_new_wo":     daily.get("new_wo",     0),
        "daily_updated_wo": daily.get("updated_wo", 0),
        "daily_new_users":  daily.get("new_users",  0),
    })


@admin_bp.route("/admin/msd-wo-updates/relogin", methods=["POST"])
def msd_wo_updates_relogin():
    global _msd_relogin_pending
    if not _msd_relogin_pending:
        return jsonify({"ok": False, "error": "No re-login is currently pending."}), 409
    try:
        _msd_otp_queue.put_nowait("__RELOGIN_CONFIRMED__")
        _msd_relogin_pending = False
    except _queue.Full:
        return jsonify({"ok": False, "error": "Re-login already queued."}), 409
    return jsonify({"ok": True})


@admin_bp.route("/admin/msd-wo-updates/status", methods=["GET"])
def msd_wo_updates_status():
    with _msd_lock:
        is_running = _msd_thread is not None and _msd_thread.is_alive()
    # Expose startup countdown: positive Unix ts means auto-start pending
    startup_run_at = _msd_startup_run_at if _msd_startup_run_at > _time.time() else None
    return jsonify({
        "ok": True,
        "is_running": is_running,
        "otp_pending": _msd_otp_pending,
        "relogin_pending": _msd_relogin_pending,
        "in_window": _msd_in_active_window(),
        "startup_run_at": startup_run_at,
    })


def _msd_reset_task_state() -> str:
    """Hard-reset the MSD download task and return the Chrome cleanup result.

    Steps:
      1. Unblock any pending OTP / re-login queue so the script thread can
         exit cleanly (we send a cancel sentinel, same as the Cancel button).
      2. Kill every Chrome process that is using the Selenium profile dir —
         closes the stuck window without deleting the saved session/cookies.
      3. Clear the running-state flags so the UI and /status reflect idle.
      4. Return ok=True — the frontend will then restart a fresh run.

    The saved Chrome profile (cookies, Dynamics session) is intentionally
    preserved so the next run can skip the login page.
    """
    global _msd_otp_pending, _msd_relogin_pending, _msd_thread

    # ── Step 1: unblock the queue so the thread exits rather than hanging ───
    _msd_otp_pending    = False
    _msd_relogin_pending = False
    try:
        _msd_otp_queue.put_nowait("__OTP_CANCELLED__")
    except _queue.Full:
        pass

    # ── Step 2: kill Chrome PIDs on the Selenium profile ────────────────────
    import pathlib as _pl
    try:
        import psutil as _psutil
        script_dir = os.path.normpath(
            os.path.join(os.path.dirname(__file__), "..", "scripts",
                         "msd-auto-download")
        )
        # Read SELENIUM_PROFILE_DIR from the script's .env (same logic as script)
        env_path = _msd_env_path()
        env_vals = _read_env_file(env_path)
        profile_dir = env_vals.get(
            "SELENIUM_PROFILE_DIR",
            str(_pl.Path.home() / ".selenium_chrome_profile"),
        )
        target = os.path.normcase(os.path.normpath(profile_dir))
        killed = 0
        for proc in _psutil.process_iter(["pid", "name", "cmdline"]):
            try:
                name = (proc.info["name"] or "").lower()
                if "chrome" not in name:
                    continue
                cmdline = proc.info["cmdline"] or []
                for arg in cmdline:
                    if "--user-data-dir=" in arg:
                        arg_path = os.path.normcase(
                            os.path.normpath(arg.split("=", 1)[1])
                        )
                        if arg_path == target:
                            proc.terminate()
                            killed += 1
                            break
            except (_psutil.NoSuchProcess, _psutil.AccessDenied):
                pass

        import time as _t
        if killed:
            _t.sleep(1.5)

        # Remove stale Chrome lock files so the next launch starts cleanly
        _profile_path = _pl.Path(profile_dir)
        for _lk in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
            _lf = _profile_path / _lk
            if _lf.exists() or _lf.is_symlink():
                try:
                    _lf.unlink()
                except Exception:
                    pass

        killed_msg = f"Chrome closed ({killed} process(es) terminated)." if killed else "No Chrome process found on this profile."
    except ImportError:
        killed_msg = "psutil not installed — Chrome process not killed."
    except Exception as _exc:
        killed_msg = f"Chrome kill warning: {_exc}"

    # ── Step 3: mark thread as dead in our bookkeeping ──────────────────────
    # The thread will exit on its own once the OTP cancel sentinel unblocks it.
    # We don't join() here (would block the request); just null out the reference
    # so the next /trigger call treats it as idle.
    with _msd_lock:
        _msd_thread = None

    return killed_msg


@admin_bp.route("/admin/msd-wo-updates/reset", methods=["POST"])
def msd_wo_updates_reset():
    return jsonify({"ok": True, "msg": _msd_reset_task_state()})


@admin_bp.route("/admin/msd-wo-updates/otp", methods=["POST"])
def msd_wo_updates_otp():
    global _msd_otp_pending
    if not _msd_otp_pending:
        return jsonify({"ok": False, "error": "No OTP is currently expected."}), 409
    data = request.get_json(silent=True) or {}
    code = str(data.get("code", "")).strip()
    if not code:
        return jsonify({"ok": False, "error": "OTP code is required."}), 400
    try:
        _msd_otp_queue.put_nowait(code)
        _msd_otp_pending = False
    except _queue.Full:
        return jsonify({"ok": False, "error": "OTP already queued."}), 409
    return jsonify({"ok": True})


@admin_bp.route("/admin/msd-wo-updates/otp/cancel", methods=["POST"])
def msd_wo_updates_otp_cancel():
    """User dismissed the OTP panel without entering a code.
    Send the __OTP_CANCELLED__ sentinel so the script thread unblocks and
    closes Chrome cleanly — without wiping the saved Chrome session on disk.
    """
    global _msd_otp_pending
    if not _msd_otp_pending:
        return jsonify({"ok": False, "error": "No OTP is currently pending."}), 409
    try:
        _msd_otp_queue.put_nowait("__OTP_CANCELLED__")
        _msd_otp_pending = False
    except _queue.Full:
        return jsonify({"ok": False, "error": "OTP queue is full — try again."}), 409
    return jsonify({"ok": True})


# ── MSD credentials helpers ───────────────────────────────────────────────────
def _msd_env_path() -> str:
    """Absolute path to the .env file used by the MSD auto-download script."""
    return os.path.normpath(
        os.path.join(
            os.path.dirname(__file__), "..", "scripts",
            "msd-auto-download", ".env",
        )
    )


def _read_env_file(path: str) -> dict:
    """Parse KEY=VALUE lines from a .env file; comments and blanks are preserved
    as raw strings under key None in a list — but returned here only as a dict."""
    result = {}
    if not os.path.isfile(path):
        return result
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            if "=" in stripped:
                key, _, val = stripped.partition("=")
                result[key.strip()] = val.strip()
    return result


def _write_env_value(path: str, key: str, new_value: str) -> None:
    """Update a single KEY= line in the .env file in-place, preserving all other
    content (comments, blank lines, ordering).  Appends the key if not present."""
    lines = []
    if os.path.isfile(path):
        with open(path, encoding="utf-8") as fh:
            lines = fh.readlines()

    updated = False
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("#") or "=" not in stripped:
            continue
        k, _, _ = stripped.partition("=")
        if k.strip() == key:
            lines[i] = f"{key}={new_value}\n"
            updated = True
            break

    if not updated:
        lines.append(f"{key}={new_value}\n")

    with open(path, "w", encoding="utf-8") as fh:
        fh.writelines(lines)


@admin_bp.route("/admin/msd-wo-updates/credentials", methods=["GET"])
def msd_wo_updates_credentials_get():
    """Return the currently configured MSD email (masked password)."""
    env_path = _msd_env_path()
    env = _read_env_file(env_path)
    email = env.get("DYNAMICS_EMAIL", "")
    has_password = bool(env.get("DYNAMICS_PASSWORD", ""))
    return jsonify({"ok": True, "email": email, "has_password": has_password})


@admin_bp.route("/admin/msd-wo-updates/credentials", methods=["POST"])
def msd_wo_updates_credentials_post():
    """Save new DYNAMICS_EMAIL / DYNAMICS_PASSWORD to the .env file and reload
    the live environment so the next MSD run picks up the new values."""
    data = request.get_json(silent=True) or {}
    email    = str(data.get("email", "")).strip()
    password = str(data.get("password", "")).strip()

    if not email:
        return jsonify({"ok": False, "error": "Email is required."}), 400
    if not password:
        return jsonify({"ok": False, "error": "Password is required."}), 400

    env_path = _msd_env_path()
    try:
        _write_env_value(env_path, "DYNAMICS_EMAIL",    email)
        _write_env_value(env_path, "DYNAMICS_PASSWORD", password)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Failed to write .env: {exc}"}), 500

    # Reload into os.environ so the running process picks up the change immediately
    os.environ["DYNAMICS_EMAIL"]    = email
    os.environ["DYNAMICS_PASSWORD"] = password

    return jsonify({"ok": True})


@admin_bp.route("/admin/data-import/verify", methods=["POST"])
def data_import_verify():
    """
    Accepts a multipart file, saves it to a temp location, runs
    verify_uploaded_file(), and returns JSON — no permanent save is done here.
    """
    file = request.files.get("excel_file")
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "No file selected."})
    if not allowed_excel(file.filename):
        return jsonify({"ok": False, "error": "Invalid file type. Allowed: .xlsx, .xls, .csv"})

    safe_name = secure_filename(file.filename)
    suffix = "." + safe_name.rsplit(".", 1)[-1].lower()

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name
            file.save(tmp_path)
    except PermissionError:
        return jsonify({
            "ok": False,
            "filename": safe_name,
            "error": (
                f'The file "{safe_name}" is currently open in another program '
                "(e.g. Microsoft Excel or OneDrive sync). "
                "Please close it and try again."
            ),
        })

    try:
        result = verify_uploaded_file(tmp_path)
        result["filename"] = safe_name          # use the user-visible name
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return jsonify(result)


@admin_bp.route("/admin/data-import/upload", methods=["POST"])
def data_import_upload():
    import shutil, traceback as _tb
    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Please choose an Excel file to upload.", "danger")
        return redirect(url_for("admin.data_import"))
    if not allowed_excel(file.filename):
        flash("Invalid file type. Allowed: .xlsx, .xls, .csv", "danger")
        return redirect(url_for("admin.data_import"))

    safe_name = secure_filename(file.filename)
    upload_folder = current_app.config["EXCEL_UPLOAD_FOLDER"]
    meta_folder   = current_app.config["UPLOAD_META_FOLDER"]
    os.makedirs(upload_folder, exist_ok=True)

    # ── Step 1: save to a temp file and verify category ──────────────────────
    suffix = "." + safe_name.rsplit(".", 1)[-1].lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp_path = tmp.name
        file.save(tmp_path)

    try:
        new_result = verify_uploaded_file(tmp_path)
        new_category = new_result.get("file_category") or ""

        # ── Step 2: if category matches an existing file, compare dates ──────
        replaced_name = None
        if new_category:
            for existing_fname in list(os.listdir(upload_folder)):
                if not allowed_excel(existing_fname):
                    continue
                # Use the cheap sidecar instead of re-parsing the whole file
                ex_result = read_meta(meta_folder, existing_fname)
                if ex_result is None:
                    try:
                        ex_result = verify_uploaded_file(
                            os.path.join(upload_folder, existing_fname)
                        )
                    except Exception:
                        continue
                if ex_result.get("file_category") != new_category:
                    continue

                # Same category found — compare latest_date (format dd-mm-yyyy)
                def _parse_date(d):
                    try:
                        from datetime import datetime as _dt
                        return _dt.strptime(d, "%d-%m-%Y")
                    except Exception:
                        return None

                new_date = _parse_date(new_result.get("latest_date") or "")
                ex_date  = _parse_date(ex_result.get("latest_date") or "")

                if new_date and ex_date and new_date >= ex_date:
                    # New file is newer or equal — delete the old one and its sidecar
                    existing_path = os.path.join(upload_folder, existing_fname)
                    os.remove(existing_path)
                    delete_meta(meta_folder, existing_fname)
                    replaced_name = existing_fname
                elif new_date and ex_date and new_date < ex_date:
                    existing_path = os.path.join(upload_folder, existing_fname)
                    # Existing file is newer — reject the upload
                    flash(
                        f'Upload rejected: "{existing_fname}" already covers a '
                        f'later date ({ex_result["latest_date"]}) for category '
                        f'"{new_category}". Delete it first if you want to replace it.',
                        "warning",
                    )
                    return redirect(url_for("admin.data_import"))
                else:
                    # Can't compare dates — keep new, delete old
                    existing_path = os.path.join(upload_folder, existing_fname)
                    os.remove(existing_path)
                    delete_meta(meta_folder, existing_fname)
                    replaced_name = existing_fname
                break  # only one file per category allowed

        # ── Step 3: move temp file to upload folder ───────────────────────────
        dest_path = os.path.join(upload_folder, safe_name)
        shutil.move(tmp_path, dest_path)
        tmp_path = None  # prevent finally from deleting it again

        # ── Step 4: write sidecar so GET loads are instant ───────────────────
        write_meta(meta_folder, safe_name, new_result)

        if replaced_name:
            flash(
                f'"{safe_name}" uploaded and replaced "{replaced_name}" '
                f'(same category: {new_category}).',
                "success",
            )
        else:
            flash(f'File "{safe_name}" uploaded successfully.', "success")

    except Exception:
        current_app.logger.error("data_import_upload ERROR:\n" + _tb.format_exc())
        flash("Upload failed due to an internal error.", "danger")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass

    return redirect(url_for("admin.data_import"))


@admin_bp.route("/admin/data-import/upsert-preview/<category_key>", methods=["GET"])
def data_import_upsert_preview(category_key: str):
    """
    Return JSON preview of rows that will be impacted when the user clicks
    Upsert for *category_key*.

    Logic per category:
      SOID     — excel soid values NOT yet in wo_product_detail
                 (where work_order_id already exists in wo_summary)
      SHIPMENT — excel rows whose SO (work_order_id) matches wo_summary
                 AND whose SOID either does not exist in wo_product_detail
                 OR exists but has ship_pn IS NULL
                 (i.e. records that will receive shipment data for the first time)

    NOTE: WOID preview logic is not implemented yet. The Upsert button for
    Work Order Advance Find View will run the upsert directly without a preview.

    Response JSON:
      {
        "ok": true,
        "category_key": "SHIPMENT",
        "impacted_count": 42,
        "total_excel_rows": 150,
        "date_col": "Order Date",
        "latest_rows": [ { col: val, ... }, ... ],   // top-3 newest
        "oldest_rows": [ { col: val, ... }, ... ],   // top-3 oldest
        "preview_cols": ["col1", "col2", ...]
      }
    """
    import traceback as _tb
    category_key = category_key.upper()
    if category_key not in FILE_CATEGORY_CONFIGS:
        return jsonify({"ok": False, "error": f'Unknown category "{category_key}".'})

    upload_folder = current_app.config["EXCEL_UPLOAD_FOLDER"]
    meta_folder   = current_app.config["UPLOAD_META_FOLDER"]

    # Find the uploaded file for this category
    cat_name = FILE_CATEGORY_CONFIGS[category_key]["file_category"]
    target_file = None
    for fname in os.listdir(upload_folder):
        from app.services.upload.excel import allowed_excel as _allowed
        if not _allowed(fname):
            continue
        m = read_meta(meta_folder, fname)
        if m and m.get("file_category") == cat_name:
            target_file = fname
            break

    if not target_file:
        return jsonify({"ok": False, "error": f'No uploaded file found for category "{cat_name}".'})

    try:
        import io
        import pandas as pd
        from app.services.database.seed import _safe_int, _build_soid
        from app.services.upload.upload_verification import verify_uploaded_file

        filepath  = os.path.join(upload_folder, target_file)
        vresult   = verify_uploaded_file(filepath)
        sheet_name = vresult.get("sheet_name", "")
        ext = filepath.rsplit(".", 1)[-1].lower()

        # Read the file bytes into memory once so openpyxl's file handle
        # (held open by verify_uploaded_file on Windows) does not block the
        # subsequent pd.read_excel call.
        with open(filepath, "rb") as _fh:
            _file_bytes = io.BytesIO(_fh.read())

        if ext == "csv":
            df = pd.read_csv(_file_bytes)
        elif sheet_name:
            df = pd.read_excel(_file_bytes, sheet_name=sheet_name)
        else:
            df = pd.read_excel(_file_bytes)

        db_path = current_app.config["DATABASE_PATH"]
        db_conn = open_db(db_path)

        try:
            if category_key == "WOID":
                # Smart diff preview for WOID.
                # A row qualifies for upsert when ANY of the following is true:
                #   1. work_order_id is not yet in wo_summary or wo_details (NEW row)
                #   2. Any date column in Excel is strictly later than the DB value
                #      (or DB is NULL and Excel has a real value)
                #   3. Any tracked status column has a different value from the DB
                import math as _math
                from datetime import datetime as _dt_woid
                from app.services.database.seed import _to_iso as _woid_to_iso

                _woid_date_cols = {
                    "Created On":                       ("wo_summary", "created_on"),
                    "Committed Delivery Date":          ("wo_summary", "committed_delivery_date"),
                    "Actual Committed Onsite Date":     ("wo_summary", "actual_committed_onsite_date"),
                    "Release Date":                     ("wo_details", "release_date"),
                    "Original Committed Onsite Date":   ("wo_details", "original_committed_onsite_date"),
                    "Customer Defer Date":              ("wo_details", "customer_defer_date"),
                    "Completion Date":                  ("wo_details", "completion_date"),
                    "Closing Date":                     ("wo_details", "closing_date"),
                }
                _woid_status_cols = {
                    "Work Order Status":          ("wo_summary", "work_order_status"),
                    "Case Status (Case) (Case)":  ("wo_summary", "case_status"),
                    "Closing Code":               ("wo_details", "closing_code"),
                    "Repeat Repair Reason":       ("wo_details", "repeat_repair_reason"),
                    "WO Cancellation Reason":     ("wo_details", "wo_cancellation_reason"),
                }

                db_summary_snap = {
                    r[0]: {
                        "created_on":                   r[1],
                        "committed_delivery_date":      r[2],
                        "actual_committed_onsite_date": r[3],
                        "work_order_status":            r[4],
                        "case_status":                  r[5],
                    }
                    for r in db_conn.execute(
                        "SELECT work_order_id, created_on, committed_delivery_date, "
                        "actual_committed_onsite_date, work_order_status, case_status "
                        "FROM wo_summary"
                    ).fetchall()
                }
                db_details_snap = {
                    r[0]: {
                        "release_date":                   r[1],
                        "original_committed_onsite_date": r[2],
                        "customer_defer_date":            r[3],
                        "completion_date":                r[4],
                        "closing_date":                   r[5],
                        "closing_code":                   r[6],
                        "repeat_repair_reason":           r[7],
                        "wo_cancellation_reason":         r[8],
                    }
                    for r in db_conn.execute(
                        "SELECT work_order_id, release_date, original_committed_onsite_date, "
                        "customer_defer_date, completion_date, closing_date, "
                        "closing_code, repeat_repair_reason, wo_cancellation_reason "
                        "FROM wo_details"
                    ).fetchall()
                }

                def _has_woid_val(v) -> bool:
                    if v is None:
                        return False
                    if isinstance(v, float) and _math.isnan(v):
                        return False
                    s = str(v).strip()
                    return s not in ("", "nan", "nat", "none", "null", "NaT")

                def _date_str_woid(v):
                    if not _has_woid_val(v):
                        return None
                    raw = _woid_to_iso(v)
                    return raw[:10] if raw else None

                def _date_newer(excel_val, db_str):
                    ex = _date_str_woid(excel_val)
                    if ex is None:
                        return False
                    if db_str is None:
                        return True
                    try:
                        return _dt_woid.fromisoformat(ex) > _dt_woid.fromisoformat(db_str[:10])
                    except (ValueError, TypeError):
                        return False

                def _status_changed_woid(excel_val, db_val):
                    ex_s = str(excel_val).strip() if _has_woid_val(excel_val) else ""
                    db_s = str(db_val).strip()    if db_val is not None else ""
                    return ex_s != db_s

                def _reason_woid(row) -> str | None:
                    """Return a human-readable upsert reason, or None if the row is skipped."""
                    wo_id = _safe_int(row.get("Work Order ID"))
                    if wo_id is None:
                        return None
                    # Rule 1 — brand new WO
                    if wo_id not in db_summary_snap or wo_id not in db_details_snap:
                        return "New WO"
                    s_row = db_summary_snap[wo_id]
                    d_row = db_details_snap[wo_id]
                    reasons: list[str] = []
                    # Rule 2 — newer date
                    for excel_col, (tbl, db_col) in _woid_date_cols.items():
                        db_val = (s_row if tbl == "wo_summary" else d_row).get(db_col)
                        if _date_newer(row.get(excel_col), db_val):
                            ex_s  = _date_str_woid(row.get(excel_col)) or ""
                            db_s  = db_val[:10] if db_val else "null"
                            reasons.append(f"{excel_col}: {db_s} → {ex_s}")
                    # Rule 3 — status changed
                    for excel_col, (tbl, db_col) in _woid_status_cols.items():
                        db_val = (s_row if tbl == "wo_summary" else d_row).get(db_col)
                        if _status_changed_woid(row.get(excel_col), db_val):
                            ex_s = str(row.get(excel_col, "")).strip() if _has_woid_val(row.get(excel_col)) else "(empty)"
                            db_s = str(db_val).strip() if db_val is not None else "(empty)"
                            reasons.append(f"{excel_col}: \u201c{db_s}\u201d → \u201c{ex_s}\u201d")
                    return "; ".join(reasons) if reasons else None

                def _qualifies_woid(row):
                    return _reason_woid(row) is not None

                wo_id_col    = "Work Order ID"
                date_col     = "Modified On"
                _woid_reason_col = "Reason"
                preview_cols = [
                    _woid_reason_col,
                    "Modified On",
                    "Work Order ID", "Serial Number", "Created On",
                    "Work Order Status", "Case Status (Case) (Case)",
                    "Closing Code", "Repeat Repair Reason", "WO Cancellation Reason",
                    "Completion Date", "Closing Date", "Committed Delivery Date",
                ]

                new_df = df[df.apply(_qualifies_woid, axis=1)].copy() \
                    if wo_id_col in df.columns \
                    else pd.DataFrame()

                # Inject synthetic "Reason" column (first column — explains why each row qualifies)
                # Also rename the raw Excel column "(Do Not Modify) Modified On" → "Modified On"
                # so the display name is clean without changing the filter/sort logic.
                if not new_df.empty:
                    new_df[_woid_reason_col] = new_df.apply(_reason_woid, axis=1).fillna("")
                    if "(Do Not Modify) Modified On" in new_df.columns:
                        new_df = new_df.rename(columns={"(Do Not Modify) Modified On": "Modified On"})

                # ── Active WOs not present in this Excel file ──────────────────
                # Fetch all WOs from the DB that are still "open": no closing_date
                # AND no completion_date.  Then exclude any WO ID that appears in
                # the Excel file (whether impacted or not) — those are already
                # accounted for.  What remains are open WOs that were simply not
                # included in the upload, which may need attention.
                excel_wo_ids = set()
                if wo_id_col in df.columns:
                    excel_wo_ids = {
                        _safe_int(v)
                        for v in df[wo_id_col]
                        if _safe_int(v) is not None
                    }

                active_wo_rows = db_conn.execute(
                    "SELECT s.work_order_id, s.serial_number, s.created_on, "
                    "s.work_order_status, s.case_status, "
                    "d.completion_date, d.closing_date, d.closing_code "
                    "FROM wo_summary s "
                    "LEFT JOIN wo_details d ON d.work_order_id = s.work_order_id "
                    "WHERE (d.completion_date IS NULL OR d.completion_date = '') "
                    "  AND (d.closing_date IS NULL OR d.closing_date = '')"
                ).fetchall()

                def _is_cancelled(wo_status, case_status) -> bool:
                    """True when either status contains 'cancel' (case-insensitive)."""
                    for s in (wo_status or "", case_status or ""):
                        if "cancel" in str(s).lower():
                            return True
                    return False

                active_wo_not_in_excel = [
                    {
                        "Work Order ID":         str(r[0]),
                        "Serial Number":         r[1] or "",
                        "Created On":            r[2] or "",
                        "Work Order Status":     r[3] or "",
                        "Case Status":           r[4] or "",
                        "Completion Date":       r[5] or "",
                        "Closing Date":          r[6] or "",
                        "Closing Code":          r[7] or "",
                    }
                    for r in active_wo_rows
                    if r[0] not in excel_wo_ids
                    and not _is_cancelled(r[3], r[4])
                ]
                active_wo_cols = [
                    "Work Order ID", "Serial Number", "Created On",
                    "Work Order Status", "Case Status",
                    "Completion Date", "Closing Date", "Closing Code",
                ]

            elif category_key == "SOID":
                # Impacted rows qualify when ANY of the following is true for a
                # row whose work_order_id exists in wo_summary:
                #   1. SOID is brand-new (not yet in wo_product_detail)
                #   2. Acceptance Date / Shipment Date / Delivery Date in Excel
                #      is strictly later than the DB value (or DB is NULL and
                #      Excel has a real date)
                #   3. Work Order Product Status has a different value from the DB
                import math as _math_soid
                from datetime import datetime as _dt_soid
                from app.services.database.seed import _to_iso as _soid_to_iso

                valid_wo_ids = {
                    r[0] for r in db_conn.execute(
                        "SELECT work_order_id FROM wo_summary"
                    ).fetchall()
                }
                # Full state snapshot: soid → (acceptance_date, shipment_date,
                #                               delivery_date, wo_product_status)
                db_soid_snap = {
                    r[0]: {
                        "acceptance_date":   r[1],
                        "shipment_date":     r[2],
                        "delivery_date":     r[3],
                        "wo_product_status": r[4],
                    }
                    for r in db_conn.execute(
                        "SELECT soid, acceptance_date, shipment_date, "
                        "delivery_date, wo_product_status "
                        "FROM wo_product_detail"
                    ).fetchall()
                }

                wo_col   = "Work Order"
                line_col = "Line Order"
                date_col = "Modified On"
                _soid_reason_col = "Upsert Reason"
                preview_cols = [
                    _soid_reason_col,
                    "Modified On",
                    wo_col, line_col, "Product", "Description",
                    "Acceptance Date", "Shipment Date", "Delivery Date",
                    "Work Order Product Status",
                ]

                _soid_date_map = [
                    ("Acceptance Date",  "acceptance_date"),
                    ("Shipment Date",    "shipment_date"),
                    ("Delivery Date",    "delivery_date"),
                ]

                def _has_soid_val(v) -> bool:
                    if v is None:
                        return False
                    if isinstance(v, float) and _math_soid.isnan(v):
                        return False
                    s = str(v).strip()
                    return s not in ("", "nan", "nat", "none", "null", "NaT")

                def _soid_date_str(v):
                    if not _has_soid_val(v):
                        return None
                    raw = _soid_to_iso(v)
                    return raw[:10] if raw else None

                def _soid_date_newer(excel_val, db_str):
                    ex = _soid_date_str(excel_val)
                    if ex is None:
                        return False
                    if db_str is None:
                        return True
                    try:
                        return _dt_soid.fromisoformat(ex) > _dt_soid.fromisoformat(db_str[:10])
                    except (ValueError, TypeError):
                        return False

                def _soid_status_changed(excel_val, db_val):
                    ex_s = str(excel_val).strip() if _has_soid_val(excel_val) else ""
                    db_s = str(db_val).strip()    if db_val is not None else ""
                    return ex_s != db_s

                def _qualify_soid(row):
                    """Return (qualifies: bool, reason: str)."""
                    wo_id = _safe_int(row.get(wo_col))
                    ln    = _safe_int(row.get(line_col))
                    soid  = _build_soid(wo_id, ln)
                    if soid is None or wo_id not in valid_wo_ids:
                        return False, ""
                    if soid not in db_soid_snap:
                        return True, "New SOID"
                    db_row = db_soid_snap[soid]
                    reasons = []
                    for excel_col, db_col in _soid_date_map:
                        if _soid_date_newer(row.get(excel_col), db_row.get(db_col)):
                            reasons.append(f"Newer {excel_col}")
                    if _soid_status_changed(
                        row.get("Work Order Product Status"),
                        db_row.get("wo_product_status"),
                    ):
                        reasons.append("Status changed")
                    if reasons:
                        return True, "; ".join(reasons)
                    return False, ""

                if wo_col in df.columns and line_col in df.columns:
                    _soid_quals = df.apply(_qualify_soid, axis=1)
                    _soid_mask  = _soid_quals.apply(lambda x: x[0])
                    new_df      = df[_soid_mask].copy()
                    new_df[_soid_reason_col] = _soid_quals[_soid_mask].apply(lambda x: x[1])
                    # Rename the raw Excel column to a clean display name
                    if "(Do Not Modify) Modified On" in new_df.columns:
                        new_df = new_df.rename(
                            columns={"(Do Not Modify) Modified On": "Modified On"}
                        )
                else:
                    new_df = pd.DataFrame()

                # ── WO-ID mismatch: Excel rows whose Work Order is not in either
                # wo_summary or wo_details.  These rows are silently skipped by the
                # upsert — surface them so the operator can investigate.
                # "valid_wo_ids" is built from wo_summary above; cross-check against
                # wo_details as well for completeness.
                _detail_ids = {
                    r[0] for r in db_conn.execute(
                        "SELECT work_order_id FROM wo_details"
                    ).fetchall()
                }
                _soid_mismatch_rows: list[dict] = []
                if wo_col in df.columns:
                    _seen_mismatch: set[int] = set()
                    for _, _mr in df.iterrows():
                        _mwo = _safe_int(_mr.get(wo_col))
                        if _mwo is None:
                            continue
                        if _mwo not in valid_wo_ids and _mwo not in _detail_ids:
                            if _mwo not in _seen_mismatch:
                                _seen_mismatch.add(_mwo)
                                # Normalise Created On to a plain date string
                                _created_raw = _mr.get("Created On")
                                try:
                                    import pandas as _pd_mm
                                    _created_str = str(_pd_mm.to_datetime(_created_raw).date()) \
                                        if _created_raw is not None and str(_created_raw).strip() not in ("", "nan", "NaT") \
                                        else ""
                                except Exception:
                                    _created_str = str(_created_raw or "").strip()[:10]
                                _soid_mismatch_rows.append({
                                    "Reason":            "WO Not Found",
                                    "Created On":        _created_str,
                                    "Work Order ID":     str(_mwo),
                                    "Line Order":        str(_mr.get(line_col) or ""),
                                    "Product":           str(_mr.get("Product") or ""),
                                    "Description":       str(_mr.get("Description") or ""),
                                    "WO Product Status": str(_mr.get("Work Order Product Status") or ""),
                                })
                # NOTE: _soid_mismatch_rows is NOT written to the JSON cache here.
                # Writing only happens inside data_import_upsert() after the user
                # confirms upsert — so cancelling the modal or refreshing the page
                # leaves the previous cache intact and the page card unchanged.

            elif category_key == "SHIPMENT":
                # Impacted rows = Excel rows where:
                #   1. SO (work_order_id) exists in wo_summary, AND
                #   2. For at least one of (ship_pn, awb, ship_pou_pod_time):
                #        - the DB value is NULL, AND
                #        - the Excel value is non-empty
                # If the Excel value for a NULL DB column is also empty/NaT,
                # there is nothing to write — the row is excluded from the preview.
                import math as _math

                valid_wo_ids = {
                    r[0] for r in db_conn.execute(
                        "SELECT work_order_id FROM wo_summary"
                    ).fetchall()
                }
                # Fetch current DB state for the three key columns, keyed by soid
                db_shipment = {
                    r[0]: {"ship_pn": r[1], "awb": r[2], "ship_pou_pod_time": r[3]}
                    for r in db_conn.execute(
                        "SELECT soid, ship_pn, awb, ship_pou_pod_time FROM wo_product_detail"
                    ).fetchall()
                }
                soid_col = "SOID"
                so_col   = "SO"
                date_col = "Order Date"
                _ship_reason_col = "Upsert Reason"
                preview_cols = [
                    _ship_reason_col,
                    so_col, soid_col, "Ship PN", "Ship PN Desc",
                    date_col, "Company Name", "AWB", "Ship POU POD Time", "SLA",
                ]

                # Excel column name → DB column name for the three tracked fields
                _excel_to_db = {
                    "Ship PN":           "ship_pn",
                    "AWB":               "awb",
                    "Ship POU POD Time": "ship_pou_pod_time",
                }

                def _has_value(v) -> bool:
                    """True when v is a non-empty, non-NaT, non-NaN value."""
                    if v is None:
                        return False
                    if isinstance(v, float) and _math.isnan(v):
                        return False
                    s = str(v).strip()
                    return s != "" and s.lower() not in ("nat", "nan", "none", "null")

                def _qualify_shipment(row):
                    """Return (qualifies: bool, reason: str)."""
                    soid  = _safe_int(row.get(soid_col))
                    wo_id = _safe_int(row.get(so_col))
                    if soid is None or wo_id not in valid_wo_ids:
                        return False, ""
                    db_row = db_shipment.get(soid)  # None if SOID not yet in DB
                    reasons = []
                    for excel_col, db_col in _excel_to_db.items():
                        db_val    = db_row[db_col] if db_row is not None else None
                        excel_val = row.get(excel_col)
                        if db_val is None and _has_value(excel_val):
                            reasons.append(f"New {excel_col}")
                    if reasons:
                        return True, "; ".join(reasons)
                    return False, ""

                if soid_col in df.columns and so_col in df.columns:
                    _ship_quals = df.apply(_qualify_shipment, axis=1)
                    _ship_mask  = _ship_quals.apply(lambda x: x[0])
                    new_df      = df[_ship_mask].copy()
                    new_df[_ship_reason_col] = _ship_quals[_ship_mask].apply(lambda x: x[1])
                else:
                    new_df = pd.DataFrame()

                # ── Previous-month SOIDs with incomplete shipment data ─────────
                # Uses the cached incomplete list (built from the latest-known anchor
                # month).  Cross-references against the current Excel to split into:
                #   A. filled_by_excel  — cached incomplete SOIDs present in this
                #      Excel file with a non-empty value for the missing field(s)
                #      → will be fixed by this upsert.
                #   B. still_incomplete — cached incomplete SOIDs NOT covered by
                #      this Excel (or still missing after it).
                from datetime import datetime as _dt_ship
                import math as _math_ship

                _pickup_col = "Ship PickUp Time"
                _incomplete_soid_cols = [
                    "Missing Fields",
                    "SOID", "SO (Work Order ID)", "Ship PN",
                    "wo_product_status", "ship_pickup_time",
                    "AWB", "Ship POU POD Time",
                ]

                # Step 1 — derive dominant month from this Excel's pickup dates
                _excel_month: str | None = None   # "YYYY-MM"
                if _pickup_col in df.columns:
                    _months: dict[str, int] = {}
                    for _v in df[_pickup_col]:
                        if not _has_value(_v):
                            continue
                        try:
                            _iso = str(pd.to_datetime(_v).date())[:7]
                            _months[_iso] = _months.get(_iso, 0) + 1
                        except Exception:
                            pass
                    if _months:
                        _excel_month = max(_months, key=lambda k: _months[k])

                # Step 2 — load the cached incomplete list (anchor = latest known month)
                from app.services.upload.meta_cache import read_incomplete_prev_shipments as _read_inc
                _cached_incomplete = _read_inc(meta_folder)

                # Step 3 — build a lookup of what this Excel provides, keyed by SOID
                #   excel_soid_vals[soid] = {"awb": val_or_None, "ship_pou_pod_time": val_or_None}
                _excel_soid_vals: dict[int, dict] = {}
                if soid_col in df.columns:
                    for _, _er in df.iterrows():
                        _esoid = _safe_int(_er.get(soid_col))
                        if _esoid is None:
                            continue
                        _excel_soid_vals[_esoid] = {
                            "awb":               _er.get("AWB"),
                            "ship_pou_pod_time":  _er.get("Ship POU POD Time"),
                            "ship_pn":           _er.get("Ship PN"),
                        }

                # Step 4 — split cached list into filled vs still-incomplete
                _filled_by_excel:   list[dict] = []
                _still_incomplete:  list[dict] = []

                for _ci in _cached_incomplete:
                    _csoid = _safe_int(_ci.get("SOID"))
                    if _csoid is None:
                        _still_incomplete.append(_ci)
                        continue
                    _ex = _excel_soid_vals.get(_csoid)
                    if _ex is None:
                        # SOID not in this Excel at all
                        _still_incomplete.append(_ci)
                        continue
                    # Check whether Excel fills each missing field
                    _fills_awb = (
                        "AWB" in _ci.get("Missing Fields", "")
                        and _has_value(_ex.get("awb"))
                    )
                    _fills_pod = (
                        "Ship POU POD Time" in _ci.get("Missing Fields", "")
                        and _has_value(_ex.get("ship_pou_pod_time"))
                    )
                    _needs_awb = "AWB" in _ci.get("Missing Fields", "")
                    _needs_pod = "Ship POU POD Time" in _ci.get("Missing Fields", "")

                    if (_needs_awb and not _fills_awb) or (_needs_pod and not _fills_pod):
                        # Excel doesn't fully fill this row — stays incomplete
                        _still_incomplete.append(_ci)
                    else:
                        # All missing fields will be filled by this Excel
                        _filled = dict(_ci)
                        # Show what Excel will write
                        if _fills_awb:
                            _filled["AWB"] = str(_ex["awb"])
                        if _fills_pod:
                            _filled["Ship POU POD Time"] = str(_ex["ship_pou_pod_time"])
                        _filled_by_excel.append(_filled)

                # _incomplete_soids shown in modal = still-incomplete after this upsert
                _incomplete_soids: list[dict] = _still_incomplete

            elif category_key == "PARTONHOLD":
                # Impacted rows = Excel rows where:
                #   1. SOID exists in wo_product_detail AND
                #   2. wo_product_status = 'On Hold - Part Hold' AND
                #   3. SO ETA is non-empty AND
                #      (DB eta_parthold_backlog IS NULL OR Excel SO ETA is strictly newer)
                import math as _math

                def _has_val_ph(v) -> bool:
                    if v is None:
                        return False
                    if isinstance(v, float) and _math.isnan(v):
                        return False
                    s = str(v).strip()
                    return s not in ("", "nan", "nat", "none", "null", "NaT")

                # DB state: soid → eta_parthold_backlog for On Hold - Part Hold rows only
                db_partonhold = {
                    r[0]: r[1]  # soid → eta_parthold_backlog
                    for r in db_conn.execute(
                        "SELECT soid, eta_parthold_backlog FROM wo_product_detail "
                        "WHERE LOWER(COALESCE(wo_product_status, '')) = 'on hold - part hold'"
                    ).fetchall()
                }

                soid_col = "SOID"
                eta_col  = "SO ETA"
                date_col = eta_col
                preview_cols = [
                    soid_col, "Service Order ID", "Part Number", "PN Desc",
                    "ETA", eta_col, "Status", "Category",
                ]

                def _is_impacted_partonhold(row):
                    soid = _safe_int(row.get(soid_col))
                    if soid is None or soid not in db_partonhold:
                        return False
                    eta_val = row.get(eta_col)
                    if not _has_val_ph(eta_val):
                        return False
                    eta_iso = str(eta_val).strip()[:10]
                    db_eta = db_partonhold[soid]
                    if db_eta is None:
                        return True
                    # Overwrite only if Excel SO ETA is strictly newer
                    try:
                        from datetime import datetime as _dt2
                        db_dt  = _dt2.fromisoformat(db_eta[:10])
                        eta_dt = _dt2.fromisoformat(eta_iso)
                        return eta_dt > db_dt
                    except (ValueError, TypeError):
                        return False

                new_df = df[df.apply(_is_impacted_partonhold, axis=1)].copy() \
                    if (soid_col in df.columns and eta_col in df.columns) \
                    else pd.DataFrame()

            elif category_key == "GTAAP":
                # Impacted rows = Excel rows where SOID exists in wo_product_detail AND:
                #   Pass 1 — DB dc_number is NULL/'0' and Excel DC# is non-null
                #   Pass 2 — DB dc_number is NULL/'0', no DC#, Return Flag = N/No
                #   Pass 3 — Excel Status is a forward move in the status hierarchy
                #             (PENDING FOR DC GENERATION → PENDING WITH PARTNER → DC GENERATED)
                #             Rows blocked by the hierarchy appear in a separate skipped table.
                from app.services.database.upsert import (
                    _GTAAP_STATUS_RANK,
                    _gtaap_status_eligible,
                )
                import math as _math

                db_gtaap = {
                    r[0]: {"dc_number": r[1], "return_status": r[2], "work_order_id": r[3]}
                    for r in db_conn.execute(
                        "SELECT soid, dc_number, return_status, work_order_id FROM wo_product_detail"
                    ).fetchall()
                }
                db_dc = {soid: v["dc_number"] for soid, v in db_gtaap.items()}
                # SOIDs fully blocked from any write (return_status already DC GENERATED)
                _blocked_soids_preview: set = {
                    soid for soid, v in db_gtaap.items()
                    if str(v["return_status"] or "").strip() == "DC GENERATED"
                }

                soid_col          = "SOID"
                dc_col            = "DC#"
                rf_col            = "Return Flag"
                status_col        = "Status"
                dc_insert_col     = "DC# (will be inserted)"
                status_insert_col = "Return Status (will be inserted)"
                reason_col        = "Reason"
                skip_reason_col   = "Skip Reason"
                date_col          = "Labor Fix Date/time"
                preview_cols = [
                    reason_col,
                    soid_col, "WO#", status_col,
                    status_insert_col,
                    dc_insert_col,
                    rf_col,
                    date_col,
                ]
                skipped_cols = [
                    soid_col, "WO#", status_col,
                    skip_reason_col,
                    rf_col,
                    date_col,
                ]

                def _has_dc_val(v) -> bool:
                    if v is None:
                        return False
                    if isinstance(v, float) and _math.isnan(v):
                        return False
                    s = str(v).strip()
                    return s not in ("", "nan", "nat", "none", "null", "NaT")

                def _dc_eligible(current) -> bool:
                    """DB dc_number is eligible to be overwritten."""
                    return current is None

                def _dc_to_str(v):
                    """Normalise whole-number floats: 17731.0 → '17731'."""
                    if not _has_dc_val(v):
                        return ""
                    try:
                        f = float(v)
                        if f == int(f):
                            return str(int(f))
                    except (ValueError, TypeError):
                        pass
                    return str(v).strip()

                def _dc_will_write(row):
                    """True when a real DC# will be written to this row."""
                    soid = _safe_int(row.get(soid_col))
                    if soid is None or soid in _blocked_soids_preview:
                        return False
                    if not _dc_eligible(db_dc.get(soid)):
                        return False
                    return _has_dc_val(row.get(dc_col))

                def _status_skip_reason(row) -> str:
                    """Return a human-readable reason why the status is blocked,
                    or '' if the status transition is allowed."""
                    soid = _safe_int(row.get(soid_col))
                    if soid is None or soid not in db_gtaap:
                        return ""
                    excel_status = str(row.get(status_col) or "").strip()
                    db_status    = str(db_gtaap[soid]["return_status"] or "").strip()
                    if not excel_status:
                        return ""
                    if excel_status == db_status:
                        return ""  # same value — not blocked, just not changed
                    incoming_rank = _GTAAP_STATUS_RANK.get(excel_status, -1)
                    current_rank  = _GTAAP_STATUS_RANK.get(db_status, -1)
                    if incoming_rank < 0:
                        return f'Unrecognised status "{excel_status}"'
                    if current_rank >= incoming_rank:
                        return (
                            f'Blocked: current status "{db_status}" '
                            f'(rank {current_rank}) cannot be overwritten '
                            f'by "{excel_status}" (rank {incoming_rank})'
                        )
                    return ""

                def _is_impacted(row):
                    soid = _safe_int(row.get(soid_col))
                    if soid is None or soid not in db_gtaap:
                        return False
                    if soid in _blocked_soids_preview:   # fully blocked — no write at all
                        return False
                    if _dc_will_write(row):
                        return True
                    # Pass 3: status is a valid forward move
                    excel_status = str(row.get(status_col) or "").strip()
                    db_status    = db_gtaap[soid]["return_status"]
                    return _gtaap_status_eligible(db_status, excel_status)

                def _is_status_blocked(row) -> bool:
                    """True when the Excel Status exists in db but is blocked by hierarchy."""
                    soid = _safe_int(row.get(soid_col))
                    if soid is None or soid not in db_gtaap:
                        return False
                    return bool(_status_skip_reason(row))

                def _preview_dc(row):
                    """Value that will actually be written — real DC# only."""
                    dc_val = row.get(dc_col)
                    if _has_dc_val(dc_val):
                        return _dc_to_str(dc_val)
                    return ""

                def _preview_status(row):
                    """Status value that will be written to return_status."""
                    return str(row.get(status_col) or "").strip()

                def _preview_reason(row):
                    """Human-readable reason why this row was selected for upsert."""
                    soid = _safe_int(row.get(soid_col))
                    parts = []
                    if soid is not None and _dc_will_write(row):
                        parts.append(f"New DC# ({_dc_to_str(row.get(dc_col))})")
                    if soid is not None and soid in db_gtaap:
                        excel_status = str(row.get(status_col) or "").strip()
                        db_status    = db_gtaap[soid]["return_status"]
                        if _gtaap_status_eligible(db_status, excel_status):
                            parts.append(f"Status: {db_status or 'NULL'} → {excel_status}")
                    return "; ".join(parts) if parts else ""

                if soid_col in df.columns and dc_col in df.columns:
                    new_df = df[df.apply(_is_impacted, axis=1)].copy()
                    new_df[dc_insert_col]     = new_df.apply(_preview_dc, axis=1)
                    new_df[status_insert_col] = new_df.apply(_preview_status, axis=1)
                    new_df[reason_col]        = new_df.apply(_preview_reason, axis=1)

                    # Skipped rows: in db, not impacted, but blocked by hierarchy
                    _blocked_mask = df.apply(_is_status_blocked, axis=1)
                    skipped_df = df[_blocked_mask & ~df.apply(_is_impacted, axis=1)].copy()
                    skipped_df[skip_reason_col] = skipped_df.apply(
                        lambda r: _status_skip_reason(r), axis=1
                    )

                    # Absent rows (Pass 4 preview) — DB rows with an open status
                    # whose SOID and work_order_id are both absent from the Excel file.
                    _open_statuses = {"PENDING FOR DC GENERATION", "PENDING WITH PARTNER"}
                    _excel_soids  = {
                        _safe_int(r.get(soid_col))
                        for _, r in df.iterrows()
                        if _safe_int(r.get(soid_col)) is not None
                    }
                    _excel_wo_ids = {
                        _safe_int(r.get("WO#"))
                        for _, r in df.iterrows()
                        if _safe_int(r.get("WO#")) is not None
                    }
                    _absent_rows: list[dict] = []
                    for _soid, _dbrow in db_gtaap.items():
                        _db_rs = str(_dbrow["return_status"] or "").strip()
                        if _db_rs not in _open_statuses:
                            continue
                        _wo_id = _dbrow.get("work_order_id")
                        if _soid in _excel_soids:
                            continue
                        if _wo_id is not None and _safe_int(_wo_id) in _excel_wo_ids:
                            continue
                        _absent_rows.append({
                            "SOID":              str(_soid),
                            "Work Order ID":     str(_wo_id) if _wo_id else "—",
                            "Current Status":    _db_rs,
                            "Will be set to":    "UNKNOWN",
                        })
                else:
                    new_df     = pd.DataFrame()
                    skipped_df = pd.DataFrame()
                    _absent_rows = []

            elif category_key == "UNRETURN":
                # Impacted rows = Excel rows whose SOID is in wo_product_detail.
                # Preview shows exactly what will be written to each column and
                # the reason for each change (including date-gate decisions).
                import math as _math_ur
                from datetime import datetime as _dt_ur

                # ── Column name constants ────────────────────────────────────────
                soid_col             = "SOID"
                dc_col               = "DC/Collection Form"
                rs_col               = "Return Status"
                subdate_col          = "DC/Collection Form-Submitted Date"
                awb_col              = "AWB Number"
                note_col             = "Note"
                date_col             = "SO Completion Date"
                dc_write_col         = "dc_lenovo (will write)"
                awb_write_col        = "awb_return (will write)"
                lrs_write_col        = "lenovo_return_status (will write)"
                notes_write_col      = "awb_notes (will write)"
                subdate_gate_col     = "modify_date_dc_lenovo (gate)"
                is_exist_write_col   = "is_exist_excel (will write)"
                reason_col           = "Reason"
                no_match_col         = "No Match Reason"
                # rs_write_col is used internally for the reason string but NOT shown
                rs_write_col         = "return_status (will write)"
                preview_cols     = [
                    reason_col,
                    soid_col, dc_col, dc_write_col,
                    rs_col, lrs_write_col,
                    awb_col, awb_write_col,
                    note_col, notes_write_col,
                    "Vendor Name", date_col,
                    subdate_col, subdate_gate_col,
                ]
                skipped_cols     = [
                    no_match_col,
                    soid_col, dc_col, rs_col, awb_col, note_col,
                    "Vendor Name",
                ]

                # ── Helper functions ─────────────────────────────────────────────
                def _ur_has_val(v) -> bool:
                    if v is None:
                        return False
                    if isinstance(v, float) and _math_ur.isnan(v):
                        return False
                    s = str(v).strip()
                    return s not in ("", "nan", "nat", "none", "null", "NaT", "0")

                def _ur_to_str_or_none(v):
                    if v is None:
                        return None
                    if isinstance(v, float) and _math_ur.isnan(v):
                        return None
                    s = str(v).strip()
                    return None if s in ("", "nan", "nat", "none", "null", "NaT") else s

                def _ur_parse_date(v):
                    """Convert 'DD-MM-YYYY' → 'YYYY-MM-DD', or None if unparseable."""
                    raw = _ur_to_str_or_none(v)
                    if not raw:
                        return None
                    try:
                        return _dt_ur.strptime(raw, "%d-%m-%Y").strftime("%Y-%m-%d")
                    except ValueError:
                        return None

                # ── DB snapshot ──────────────────────────────────────────────────
                db_unreturn = {
                    r[0]: {
                        "dc_lenovo":             r[1],
                        "return_status":         r[2],
                        "awb_return":            r[3],
                        "lenovo_return_status":  r[4],
                        "awb_notes":             r[5],
                        "modify_date_dc_lenovo": r[6],
                        "is_exist_excel":        r[7],
                    }
                    for r in db_conn.execute(
                        """SELECT soid, dc_lenovo, return_status,
                                  awb_return, lenovo_return_status, awb_notes,
                                  modify_date_dc_lenovo, is_exist_excel
                           FROM wo_product_detail"""
                    ).fetchall()
                }

                # ── File-level version stamp (mirrors upsert logic) ──────────────
                # Compute max(DC/Collection Form-Submitted Date) across all Excel rows.
                _ur_max_submitted: str | None = None
                for _, _ur_r in df.iterrows():
                    _d = _ur_parse_date(_ur_r.get(subdate_col))
                    if _d and (_ur_max_submitted is None or _d > _ur_max_submitted):
                        _ur_max_submitted = _d

                # Single stored stamp = MAX(modify_date_dc_lenovo) already in DB
                _ur_stamp_row = db_conn.execute(
                    "SELECT MAX(modify_date_dc_lenovo) FROM wo_product_detail"
                    " WHERE modify_date_dc_lenovo IS NOT NULL"
                ).fetchone()
                _ur_stored_stamp: str | None = _ur_stamp_row[0] if _ur_stamp_row else None

                # file_gate_pass mirrors upsert_from_unreturn():
                #   True  → file is newer; write all new-col rows
                #   False → file not newer; all new-col writes blocked
                #   None  → no parseable dates; first-time fill only
                if _ur_max_submitted:
                    _ur_file_gate_pass = (_ur_stored_stamp is None) or (_ur_max_submitted > _ur_stored_stamp)
                else:
                    _ur_file_gate_pass = None

                # Stage dc_lenovo in-memory first (needed for return_status decision)
                staged_dc: dict[int, str | None] = {
                    soid: v["dc_lenovo"] for soid, v in db_unreturn.items()
                }
                for _, _r in df.iterrows():
                    _s = _safe_int(_r.get(soid_col))
                    if _s is not None and _s in db_unreturn:
                        _incoming_dc = _ur_to_str_or_none(_r.get(dc_col))
                        _current_dc  = _ur_to_str_or_none(db_unreturn[_s]["dc_lenovo"])
                        # Only stage if real AND different from current DB value
                        if _ur_has_val(_incoming_dc) and _incoming_dc != _current_dc:
                            staged_dc[_s] = _incoming_dc

                # ── Pre-compute Excel SOID set (for is_exist_excel preview) ─────
                _ur_excel_soids: set[int] = set()
                for _, _xe_r in df.iterrows():
                    _xe_s = _safe_int(_xe_r.get(soid_col))
                    if _xe_s is not None:
                        _ur_excel_soids.add(_xe_s)

                impacted_rows: list[dict] = []
                skipped_rows:  list[dict] = []

                # ── Gate guard: file not newer → null-fill for eligible rows ──────────
                # Mirrors upsert_from_unreturn(): when file_gate_pass is False,
                # awb_return/lenovo_return_status/awb_notes are written when:
                #   • awb_return and awb_notes are NULL in the DB, AND
                #   • lenovo_return_status is NULL or "Unreturned" (unlocked)
                if _ur_file_gate_pass is False:
                    _gate_null_reason = (
                        f"file max date ({_ur_max_submitted}) \u2264 stored ({_ur_stored_stamp})"
                        " — null/Unreturned-fill"
                    )
                    _gate_skip_reason = (
                        f"file max date ({_ur_max_submitted}) \u2264 stored ({_ur_stored_stamp})"
                        " — date-gated columns skipped (cols already set)"
                    )
                    for _, _gs_row in df.iterrows():
                        _gs_soid   = _safe_int(_gs_row.get(soid_col))
                        _gs_vendor = str(_gs_row.get("Vendor Name") or "").strip()
                        _gs_dc     = _ur_to_str_or_none(_gs_row.get(dc_col))
                        _gs_rs     = str(_gs_row.get(rs_col) or "").strip()
                        _gs_awb    = _ur_to_str_or_none(_gs_row.get(awb_col))
                        _gs_note   = _ur_to_str_or_none(_gs_row.get(note_col))
                        _gs_rs_clean = _ur_to_str_or_none(_gs_row.get(rs_col))

                        if _gs_soid is None or _gs_soid not in db_unreturn:
                            skipped_rows.append({
                                soid_col:      str(_gs_soid) if _gs_soid is not None else "—",
                                "Vendor Name": _gs_vendor,
                                dc_col:        str(_gs_dc or "—"),
                                rs_col:        _gs_rs,
                                awb_col:       str(_gs_awb or "—"),
                                note_col:      str(_gs_note or "—"),
                                no_match_col:  "SOID not found in wo_product_detail",
                            })
                            continue

                        _gs_db     = db_unreturn[_gs_soid]
                        _gs_db_lrs = _ur_to_str_or_none(_gs_db["lenovo_return_status"])
                        _lrs_unlocked = (
                            _gs_db_lrs is None
                            or _gs_db_lrs.strip().lower() == "unreturned"
                        )
                        _eligible = (
                            _ur_to_str_or_none(_gs_db["awb_return"]) is None
                            and _lrs_unlocked
                            and _ur_to_str_or_none(_gs_db["awb_notes"]) is None
                        )
                        _has_any = (_gs_awb is not None or _gs_rs_clean is not None or _gs_note is not None)

                        # dc_lenovo is written unconditionally (not gated) — check if it would change
                        _gs_current_dc = _ur_to_str_or_none(_gs_db["dc_lenovo"])
                        _gs_dc_write   = _gs_dc if (_ur_has_val(_gs_dc) and _gs_dc != _gs_current_dc) else None

                        if _eligible and _has_any:
                            _null_parts = []
                            if _gs_dc_write is not None:
                                _null_parts.append(f"dc_lenovo ← '{_gs_dc_write}'")
                            if _gs_awb is not None:
                                _null_parts.append(f"awb_return ← '{_gs_awb}'")
                            if _gs_rs_clean is not None:
                                # Skip if incoming LRS is identical to what's already in DB
                                _lrs_same = (
                                    _gs_db_lrs is not None
                                    and _gs_db_lrs.strip().lower() == _gs_rs_clean.strip().lower()
                                )
                                if not _lrs_same:
                                    _lrs_label = (
                                        f"lenovo_return_status: Unreturned → '{_gs_rs_clean}'"
                                        if _gs_db_lrs is not None
                                        else f"lenovo_return_status ← '{_gs_rs_clean}'"
                                    )
                                    _null_parts.append(_lrs_label)
                            if _gs_note is not None:
                                _null_parts.append(f"awb_notes ← '{_gs_note}'")
                            # Nothing actually changes — route to skipped
                            if not _null_parts:
                                skipped_rows.append({
                                    soid_col:      str(_gs_soid),
                                    "Vendor Name": _gs_vendor,
                                    dc_col:        str(_gs_dc or "—"),
                                    rs_col:        _gs_rs,
                                    awb_col:       str(_gs_awb or "—"),
                                    note_col:      str(_gs_note or "—"),
                                    no_match_col:  "no change — all values match DB",
                                })
                                continue
                            impacted_rows.append({
                                reason_col:       "; ".join(_null_parts),
                                soid_col:         str(_gs_soid),
                                dc_col:           str(_gs_dc or "—"),
                                dc_write_col:     str(_gs_dc_write or ""),
                                rs_col:           _gs_rs,
                                awb_col:          str(_gs_awb or "—"),
                                awb_write_col:    str(_gs_awb or ""),
                                lrs_write_col:    str(_gs_rs_clean or ""),
                                note_col:         str(_gs_note or "—"),
                                notes_write_col:  str(_gs_note or ""),
                                "Vendor Name":    _gs_vendor,
                                date_col:         "",
                                subdate_col:      "",
                                subdate_gate_col: _gate_null_reason,
                            })
                        elif _gs_dc_write is not None:
                            # Only dc_lenovo changes (no new-col write) — still an impacted row
                            impacted_rows.append({
                                reason_col:       f"dc_lenovo ← '{_gs_dc_write}'",
                                soid_col:         str(_gs_soid),
                                dc_col:           str(_gs_dc or "—"),
                                dc_write_col:     str(_gs_dc_write),
                                rs_col:           _gs_rs,
                                awb_col:          str(_gs_awb or "—"),
                                awb_write_col:    "",
                                lrs_write_col:    "",
                                note_col:         str(_gs_note or "—"),
                                notes_write_col:  "",
                                "Vendor Name":    _gs_vendor,
                                date_col:         "",
                                subdate_col:      "",
                                subdate_gate_col: _gate_null_reason,
                            })
                        else:
                            _skip_why = _gate_skip_reason
                            if not _lrs_unlocked:
                                _skip_why = f"lenovo_return_status locked (current: '{_gs_db_lrs}') — skipped"
                            skipped_rows.append({
                                soid_col:      str(_gs_soid),
                                "Vendor Name": _gs_vendor,
                                dc_col:        str(_gs_dc or "—"),
                                rs_col:        _gs_rs,
                                awb_col:       str(_gs_awb or "—"),
                                note_col:      str(_gs_note or "—"),
                                no_match_col:  _skip_why,
                            })
                for _, row in df.iterrows():
                    if _ur_file_gate_pass is False:
                        break  # already handled above
                    soid    = _safe_int(row.get(soid_col))
                    vendor  = str(row.get("Vendor Name") or "").strip()
                    dc_val  = _ur_to_str_or_none(row.get(dc_col))
                    rs_val  = str(row.get(rs_col) or "").strip()
                    so_comp = str(row.get(date_col) or "").strip()
                    awb_val = _ur_to_str_or_none(row.get(awb_col))
                    note_val = _ur_to_str_or_none(row.get(note_col))
                    raw_subdate = str(row.get(subdate_col) or "").strip()
                    submitted_iso = _ur_parse_date(row.get(subdate_col))

                    if soid is None or soid not in db_unreturn:
                        skipped_rows.append({
                            soid_col:      str(soid) if soid is not None else "—",
                            "Vendor Name": vendor,
                            dc_col:        str(dc_val or "—"),
                            rs_col:        rs_val,
                            awb_col:       str(awb_val or "—"),
                            note_col:      str(note_val or "—"),
                            no_match_col:  "SOID not found in wo_product_detail",
                        })
                        continue

                    # dc_is_real: incoming DC value is real AND differs from current DB value
                    current_dc = _ur_to_str_or_none(db_unreturn[soid]["dc_lenovo"])
                    dc_is_real = _ur_has_val(dc_val) and dc_val != current_dc

                    # ── dc_lenovo decision ───────────────────────────────────
                    # return_status is not modified by the Unreturn upsert.
                    dc_write = dc_val if dc_is_real else ""

                    # ── new-column date-gate decision (FILE-LEVEL) ───────────────
                    # Mirrors the new upsert_from_unreturn() gate:
                    #   _ur_file_gate_pass = True  → file max date > stored stamp → write
                    #   _ur_file_gate_pass = False → file not newer → block all new-col writes
                    #   _ur_file_gate_pass = None  → no dates in file → first-time fill only
                    db_awb_cur   = _ur_to_str_or_none(db_unreturn[soid]["awb_return"])
                    db_lrs_cur   = _ur_to_str_or_none(db_unreturn[soid]["lenovo_return_status"])
                    db_notes_cur = _ur_to_str_or_none(db_unreturn[soid]["awb_notes"])
                    cols_are_null = (db_awb_cur is None and db_lrs_cur is None and db_notes_cur is None)

                    new_cols_write = False
                    gate_reason    = ""
                    skip_reason    = ""
                    if _ur_file_gate_pass is True:
                        new_cols_write = True
                        if _ur_stored_stamp:
                            gate_reason = f"file max {_ur_max_submitted} > stored {_ur_stored_stamp} — write"
                        else:
                            gate_reason = f"file max {_ur_max_submitted} (first write)"
                    elif _ur_file_gate_pass is False:
                        gate_reason = f"file max {_ur_max_submitted} ≤ stored {_ur_stored_stamp} — skip"
                        skip_reason = f"file max date ({_ur_max_submitted}) ≤ stored ({_ur_stored_stamp}) — skip"
                    else:
                        # No parseable dates in file
                        if cols_are_null:
                            new_cols_write = True
                            gate_reason = "no dates in file — first-time fill"
                        else:
                            gate_reason = "no dates in file, cols already set — skip"
                            skip_reason = "no dates in file, cols already set — skip"

                    # ── build reasons list ───────────────────────────────────
                    reasons = []
                    if dc_is_real:
                        reasons.append(f"dc_lenovo ← '{dc_val}'")
                    if new_cols_write:
                        # ── Same rules as upsert Pass 3 ──────────────────────

                        db_awb_val   = _ur_to_str_or_none(db_unreturn[soid]["awb_return"])
                        db_lrs_val   = _ur_to_str_or_none(db_unreturn[soid]["lenovo_return_status"])
                        db_notes_val = _ur_to_str_or_none(db_unreturn[soid]["awb_notes"])

                        # lenovo_return_status lock rule (mirrors upsert Pass 3):
                        #   NULL or "Unreturned" → unlocked, Excel value wins
                        #   Any other value      → locked, keep DB value
                        _lrs_locked_prev = (
                            db_lrs_val is not None
                            and db_lrs_val.strip().lower() != "unreturned"
                        )
                        eff_lrs_val = db_lrs_val if _lrs_locked_prev else (rs_val if rs_val is not None else db_lrs_val)

                        eff_awb_val   = awb_val  if awb_val   is not None else db_awb_val
                        eff_notes_val = note_val if note_val  is not None else db_notes_val

                        # Skip if all effective values match what is already in the DB
                        if eff_awb_val == db_awb_val and eff_lrs_val == db_lrs_val and eff_notes_val == db_notes_val:
                            new_cols_write = False
                            gate_reason = gate_reason + " [no change — all values match DB]"
                            skip_reason  = "no change — all values match DB"
                        # Nothing to write at all
                        elif eff_awb_val is None and eff_lrs_val is None and eff_notes_val is None:
                            new_cols_write = False
                            gate_reason = gate_reason + " [nothing to write after filters]"
                            skip_reason  = "nothing to write after filters"
                        else:
                            col_notes = []
                            if eff_awb_val is not None and eff_awb_val != db_awb_val:
                                col_notes.append(f"awb_return ← '{eff_awb_val}'")
                            if eff_lrs_val is not None and eff_lrs_val != db_lrs_val:
                                _lrs_prefix = f"Unreturned → " if db_lrs_val and db_lrs_val.strip().lower() == "unreturned" else ""
                                col_notes.append(f"lenovo_return_status: {_lrs_prefix}'{eff_lrs_val}'")
                            if eff_notes_val is not None and eff_notes_val != db_notes_val:
                                col_notes.append(f"awb_notes ← '{eff_notes_val}'")
                            if col_notes:
                                reasons.append("; ".join(col_notes))
                            else:
                                # All effective values match DB — nothing real to write
                                new_cols_write = False
                                gate_reason = gate_reason + " [no change — all values match DB]"
                                skip_reason  = "no change — all values match DB"
                    else:
                        # Still read DB values so the preview can show "existing [skipped]"
                        db_awb_val   = _ur_to_str_or_none(db_unreturn[soid]["awb_return"])
                        db_lrs_val   = _ur_to_str_or_none(db_unreturn[soid]["lenovo_return_status"])
                        db_notes_val = _ur_to_str_or_none(db_unreturn[soid]["awb_notes"])
                        eff_awb_val = eff_lrs_val = eff_notes_val = None
                        eff_rs_val  = rs_val

                    # Skip rows with nothing at all to write
                    if not dc_is_real and not new_cols_write:
                        skipped_rows.append({
                            soid_col:      str(soid),
                            dc_col:        str(dc_val or "—"),
                            rs_col:        rs_val,
                            awb_col:       str(awb_val or "—"),
                            note_col:      str(note_val or "—"),
                            "Vendor Name": vendor,
                            no_match_col:  skip_reason or "DC/Collection Form is empty, 0, or — and no new-column write",
                        })
                        continue

                    impacted_rows.append({
                        reason_col:       "; ".join(reasons),
                        soid_col:         str(soid),
                        dc_col:           str(dc_val or "—"),
                        dc_write_col:     dc_write,
                        rs_col:           rs_val,
                        # rs_write_col not included — hidden from preview
                        awb_col:          str(awb_val or "—"),
                        awb_write_col:    str(eff_awb_val or "") if new_cols_write else (f"{db_awb_val} [skipped]"   if db_awb_val   else "[skipped]"),
                        lrs_write_col:    str(eff_lrs_val or "") if new_cols_write else (f"{db_lrs_val} [skipped]"   if db_lrs_val   else "[skipped]"),
                        note_col:         str(note_val or "—"),
                        notes_write_col:  str(eff_notes_val or "") if new_cols_write else (f"{db_notes_val} [skipped]" if db_notes_val else "[skipped]"),
                        "Vendor Name":    vendor,
                        date_col:         so_comp,
                        subdate_col:      raw_subdate,
                        subdate_gate_col: gate_reason,
                    })

                import pandas as _pd_ur
                new_df     = _pd_ur.DataFrame(impacted_rows) if impacted_rows else _pd_ur.DataFrame()
                skipped_df = _pd_ur.DataFrame(skipped_rows)  if skipped_rows  else _pd_ur.DataFrame()

            else:
                return jsonify({"ok": False, "error": "Category has no DB preview support."})

        finally:
            db_conn.close()

        impacted_count = len(new_df)

        # WOID and SOID: sort by Modified On ascending (oldest first).
        # All other categories: sort by date_col descending (newest first).
        _sort_asc = (category_key in ("WOID", "SOID"))

        def _all_rows_sorted(frame, col_list):
            if frame.empty or date_col not in frame.columns:
                return []
            tmp = frame.copy()
            tmp["_sort_date"] = pd.to_datetime(tmp[date_col], format="mixed", errors="coerce")
            tmp = tmp.sort_values("_sort_date", ascending=_sort_asc, na_position="last")
            tmp = tmp.drop(columns=["_sort_date"])
            cols = [c for c in col_list if c in tmp.columns]
            rows_out = []
            for _, r in tmp[cols].iterrows():
                rows_out.append({
                    k: ("" if (v is None or (isinstance(v, float) and __import__("math").isnan(v)))
                        else str(v))
                    for k, v in r.items()
                })
            return rows_out

        # Use new_df for WOID, GTAAP, SOID, SHIPMENT, and UNRETURN because they carry
        # synthetic columns ("Reason", "dc_lenovo (will write)", etc.) not in the raw df.
        _preview_col_source = new_df if (not new_df.empty and category_key in ("WOID", "GTAAP", "SOID", "SHIPMENT", "UNRETURN")) else df
        _resp = {
            "ok":               True,
            "category_key":     category_key,
            "filename":         target_file,
            "impacted_count":   impacted_count,
            "total_excel_rows": len(df),
            "date_col":         date_col,
            "preview_cols":     [c for c in preview_cols if c in _preview_col_source.columns],
            "all_rows":         _all_rows_sorted(new_df, preview_cols),
        }
        if category_key == "WOID":
            _resp["active_wo_not_in_excel"] = active_wo_not_in_excel
            _resp["active_wo_cols"]         = active_wo_cols

            # ── New technicians preview ───────────────────────────────────────
            _leap_col   = "LEAP ID (Technician ID) (Contact)"
            _tname_col  = "Technician ID"
            _vendor_col = "Labor Vendor Related"
            _new_tech_rows: list[dict] = []
            if _leap_col in df.columns:
                _seen_leaps: set[str] = set()
                for _, _tr in df.iterrows():
                    _leap = str(_tr.get(_leap_col) or "").strip()
                    if not _leap or _leap in _seen_leaps:
                        continue
                    _seen_leaps.add(_leap)
                    _tname  = str(_tr.get(_tname_col)  or "").strip()
                    _vendor = str(_tr.get(_vendor_col) or "").strip()
                    if _tname:
                        _new_tech_rows.append({
                            "LEAP ID":   _leap,
                            "Full Name": _tname,
                            "Vendor ID": _vendor,
                        })
                if _new_tech_rows:
                    _existing_leaps = {
                        r[0]
                        for r in db_conn.execute(
                            "SELECT tech_id FROM asp_users WHERE tech_id IS NOT NULL"
                        ).fetchall()
                    }
                    _new_tech_rows = [
                        t for t in _new_tech_rows
                        if t["LEAP ID"] not in _existing_leaps
                    ]
            _resp["new_technicians"]      = _new_tech_rows
            _resp["new_technicians_cols"] = ["LEAP ID", "Full Name", "Vendor ID"]
        if category_key == "SOID":
            _resp["wo_product_mismatch"]      = _soid_mismatch_rows
            _resp["wo_product_mismatch_cols"] = [
                "Reason", "Created On", "Work Order ID", "Line Order",
                "Product", "Description", "WO Product Status",
            ]
        if category_key == "GTAAP":
            _skipped_cols = [c for c in skipped_cols if c in skipped_df.columns] \
                            if not skipped_df.empty else skipped_cols
            _resp["gtaap_skipped_rows"]      = _all_rows_sorted(skipped_df, skipped_cols)
            _resp["gtaap_skipped_cols"]      = _skipped_cols
            _resp["gtaap_skipped_count"]     = len(skipped_df)
            _resp["gtaap_absent_rows"]       = _absent_rows
            _resp["gtaap_absent_cols"]       = ["SOID", "Work Order ID", "Current Status", "Will be set to"]
            _resp["gtaap_absent_count"]      = len(_absent_rows)
        if category_key == "SHIPMENT":
            _resp["incomplete_prev_soids"]      = _incomplete_soids
            _resp["incomplete_prev_soid_cols"]  = _incomplete_soid_cols
            _resp["excel_month"]                = _excel_month or ""
            _resp["filled_by_excel"]            = _filled_by_excel
            _resp["filled_by_excel_cols"]       = _incomplete_soid_cols
        if category_key == "UNRETURN":
            _ur_skipped_cols = [c for c in skipped_cols if c in skipped_df.columns] \
                               if not skipped_df.empty else skipped_cols
            _resp["unreturn_skipped_rows"]  = (
                skipped_df[_ur_skipped_cols].to_dict(orient="records")
                if not skipped_df.empty else []
            )
            _resp["unreturn_skipped_cols"]  = _ur_skipped_cols
            _resp["unreturn_skipped_count"] = len(skipped_df)
        return jsonify(_resp)

    except Exception:
        current_app.logger.error(
            "upsert_preview failed for %s:\n%s", category_key, _tb.format_exc()
        )
        return jsonify({"ok": False, "error": "Preview failed. Check server logs."})


@admin_bp.route("/admin/data-import/upsert/<category_key>", methods=["POST"])
def data_import_upsert(category_key: str):
    """Manual upsert: push the already-uploaded file for *category_key* into the DB."""
    import traceback as _tb
    category_key = category_key.upper()
    if category_key not in FILE_CATEGORY_CONFIGS:
        flash(f'Unknown category "{category_key}".', "danger")
        return redirect(url_for("admin.data_import"))

    upload_folder = current_app.config["EXCEL_UPLOAD_FOLDER"]
    meta_folder   = current_app.config["UPLOAD_META_FOLDER"]

    # Find the uploaded file for this category
    cat_name = FILE_CATEGORY_CONFIGS[category_key]["file_category"]
    target_file = None
    for fname in os.listdir(upload_folder):
        from app.services.upload.excel import allowed_excel as _allowed
        if not _allowed(fname):
            continue
        m = read_meta(meta_folder, fname)
        if m and m.get("file_category") == cat_name:
            target_file = fname
            break

    if not target_file:
        flash(f'No uploaded file found for category "{cat_name}".', "warning")
        return redirect(url_for("admin.data_import"))

    filepath = os.path.join(upload_folder, target_file)
    try:
        db_path = current_app.config["DATABASE_PATH"]
        db_conn = open_db(db_path)
        try:
            _upsert_result = dispatch_upsert(category_key, filepath, db_conn)
        finally:
            db_conn.close()
        if category_key == "WOID":
            n_new_wo, n_updated, n_new_users = _upsert_result
            n_rows = n_new_wo + n_updated
        elif category_key == "GTAAP":
            n_new_dc_ui, n_new_status_ui = _upsert_result
            n_rows = n_new_dc_ui + n_new_status_ui
            n_new_wo    = 0
            n_updated   = 0
            n_new_users = 0
        else:
            n_rows = _upsert_result
            n_new_wo    = 0
            n_updated   = 0
            n_new_users = 0
        mark_upserted(meta_folder, target_file)
        # ── SOID: rebuild the WO-product mismatch cache ───────────────────────
        # Only written here (on confirmed upsert), never during preview-only.
        # Cancel / refresh without confirming leaves the previous cache intact.
        if category_key == "SOID":
            try:
                import io as _io_soid
                import pandas as _pd_soid
                from app.services.upload.upload_verification import (
                    verify_uploaded_file as _vuf_soid,
                )
                from app.services.database.seed import _safe_int as _si_soid
                _vr_soid = _vuf_soid(filepath)
                _sn_soid = _vr_soid.get("sheet_name", "")
                with open(filepath, "rb") as _fh_soid:
                    _fb_soid = _io_soid.BytesIO(_fh_soid.read())
                _df_soid = (
                    _pd_soid.read_excel(_fb_soid, sheet_name=_sn_soid)
                    if _sn_soid else _pd_soid.read_excel(_fb_soid)
                )
                _db_soid = open_db(db_path)
                try:
                    _valid_ids_soid = {
                        r[0] for r in _db_soid.execute(
                            "SELECT work_order_id FROM wo_summary"
                        ).fetchall()
                    }
                    _detail_ids_soid = {
                        r[0] for r in _db_soid.execute(
                            "SELECT work_order_id FROM wo_details"
                        ).fetchall()
                    }
                finally:
                    _db_soid.close()
                _wo_col_soid  = "Work Order"
                _ln_col_soid  = "Line Order"
                _mismatch_soid: list[dict] = []
                _seen_soid: set[int] = set()
                if _wo_col_soid in _df_soid.columns:
                    for _, _mr_s in _df_soid.iterrows():
                        _mwo_s = _si_soid(_mr_s.get(_wo_col_soid))
                        if _mwo_s is None:
                            continue
                        if _mwo_s not in _valid_ids_soid and _mwo_s not in _detail_ids_soid:
                            if _mwo_s not in _seen_soid:
                                _seen_soid.add(_mwo_s)
                                _cr_raw = _mr_s.get("Created On")
                                try:
                                    _cr_str = str(_pd_soid.to_datetime(_cr_raw).date()) \
                                        if _cr_raw is not None and str(_cr_raw).strip() not in ("", "nan", "NaT") \
                                        else ""
                                except Exception:
                                    _cr_str = str(_cr_raw or "").strip()[:10]
                                _mismatch_soid.append({
                                    "Reason":            "WO Not Found",
                                    "Created On":        _cr_str,
                                    "Work Order ID":     str(_mwo_s),
                                    "Line Order":        str(_mr_s.get(_ln_col_soid) or ""),
                                    "Product":           str(_mr_s.get("Product") or ""),
                                    "Description":       str(_mr_s.get("Description") or ""),
                                    "WO Product Status": str(_mr_s.get("Work Order Product Status") or ""),
                                })
                write_wo_product_mismatch(meta_folder, _mismatch_soid)
            except Exception:
                current_app.logger.warning(
                    "write_wo_product_mismatch failed after upsert:\n" + _tb.format_exc()
                )
        # Rebuild the active-WO cache after a WOID upsert.
        # Pass the WO IDs from the Excel file so the cache only stores WOs
        # that were genuinely absent from this upload (matching the modal view).
        if category_key == "SHIPMENT":
            # Rebuild the incomplete-prev-shipments cache after a SHIPMENT upsert.
            # Derive the current Excel month from the uploaded file's pickup dates.
            try:
                import io as _io3
                import pandas as _pd3
                from app.services.upload.upload_verification import (
                    verify_uploaded_file as _vuf3,
                )
                _vr3 = _vuf3(filepath)
                _sn3 = _vr3.get("sheet_name", "")
                with open(filepath, "rb") as _fh3:
                    _fb3 = _io3.BytesIO(_fh3.read())
                _df3 = (
                    _pd3.read_excel(_fb3, sheet_name=_sn3)
                    if _sn3 else _pd3.read_excel(_fb3)
                )
                _months3: dict[str, int] = {}
                for _v3 in _df3.get("Ship PickUp Time", _pd3.Series(dtype=object)):
                    try:
                        _iso3 = str(_pd3.to_datetime(_v3).date())[:7]
                        _months3[_iso3] = _months3.get(_iso3, 0) + 1
                    except Exception:
                        pass
                # Use the MAXIMUM month across all months in the Excel so
                # uploading an older monthly report never regresses the anchor.
                _excel_month3 = max(_months3.keys()) if _months3 else ""
                if _excel_month3:
                    # Also compare against the existing cached anchor so we
                    # never move the anchor backwards.
                    from app.services.upload.meta_cache import (
                        read_incomplete_prev_shipments as _read_inc3,
                    )
                    _cur_cache3 = _read_inc3(meta_folder)
                    if _cur_cache3:
                        # Infer current anchor from the highest ship_pickup_time in cache
                        _cached_months3 = [
                            r.get("ship_pickup_time", "")[:7]
                            for r in _cur_cache3
                            if r.get("ship_pickup_time", "")[:7]
                        ]
                        if _cached_months3:
                            _anchor_from_cache = max(_cached_months3)
                            # advance anchor by one month so cached rows are included
                            import calendar as _cal3
                            _y3, _m3 = int(_anchor_from_cache[:4]), int(_anchor_from_cache[5:7])
                            _m3 += 1
                            if _m3 > 12:
                                _m3, _y3 = 1, _y3 + 1
                            _cache_anchor_next = f"{_y3:04d}-{_m3:02d}"
                            _excel_month3 = max(_excel_month3, _cache_anchor_next)
                    rebuild_incomplete_prev_shipments(meta_folder, db_path, _excel_month3)
            except Exception:
                current_app.logger.warning(
                    "rebuild_incomplete_prev_shipments failed after upsert:\n" + _tb.format_exc()
                )
        if category_key == "WOID":
            try:
                import io as _io2
                import pandas as _pd2
                from app.services.upload.upload_verification import (
                    verify_uploaded_file as _vuf,
                )
                _vr = _vuf(filepath)
                _sn = _vr.get("sheet_name", "")
                with open(filepath, "rb") as _fh2:
                    _fb2 = _io2.BytesIO(_fh2.read())
                _df2 = (
                    _pd2.read_excel(_fb2, sheet_name=_sn)
                    if _sn else _pd2.read_excel(_fb2)
                )
                _excel_ids: set[int] = {
                    int(float(str(v)))
                    for v in _df2.get("Work Order ID", _pd2.Series(dtype=object))
                    if v is not None and str(v).strip() not in ("", "nan", "NaN")
                }
                rebuild_active_open_wos(meta_folder, db_path, _excel_ids)
            except Exception:
                current_app.logger.warning(
                    "rebuild_active_open_wos failed after upsert:\n" + _tb.format_exc()
                )
        if category_key == "WOID":
            _parts = [f"{n_new_wo} new WO{'s' if n_new_wo != 1 else ''}"]
            if n_updated:
                _parts.append(f"{n_updated} updated")
            if n_new_users:
                _parts.append(f"{n_new_users} new technician{'s' if n_new_users != 1 else ''} added to asp_users")
            flash(
                f'"{target_file}" upserted successfully — ' + ", ".join(_parts) + ".",
                "success",
            )
        elif category_key == "GTAAP":
            flash(
                f'"{target_file}" upserted successfully — '
                f'{n_new_dc_ui} new DC#, {n_new_status_ui} status change{"s" if n_new_status_ui != 1 else ""}.',
                "success",
            )
        else:
            flash(
                f'"{target_file}" upserted successfully — {n_rows} row{"s" if n_rows != 1 else ""} processed.',
                "success",
            )
    except Exception:
        current_app.logger.error("manual upsert failed for %s:\n%s", target_file, _tb.format_exc())
        flash(f'Upsert failed for "{target_file}". Check server logs.', "danger")

    return redirect(url_for("admin.data_import"))


@admin_bp.route("/admin/data-import/delete/<filename>", methods=["POST"])
def data_import_delete(filename):
    safe = secure_filename(filename)
    folder      = current_app.config["EXCEL_UPLOAD_FOLDER"]
    meta_folder = current_app.config["UPLOAD_META_FOLDER"]
    path = os.path.join(folder, safe)
    if os.path.isfile(path):
        os.remove(path)
        delete_meta(meta_folder, safe)
        flash(f'File "{safe}" deleted.', "success")
    else:
        flash("File not found.", "danger")
    return redirect(url_for("admin.data_import"))


@admin_bp.route("/admin/data-import/reset", methods=["POST"])
def data_import_reset():
    import traceback as _tb, sys
    try:
        folder      = current_app.config["EXCEL_UPLOAD_FOLDER"]
        meta_folder = current_app.config["UPLOAD_META_FOLDER"]
        deleted = 0
        if os.path.isdir(folder):
            for fname in list(os.listdir(folder)):
                if not allowed_excel(fname):
                    continue
                fpath = os.path.join(folder, fname)
                if os.path.isfile(fpath):
                    os.remove(fpath)
                    delete_meta(meta_folder, fname)
                    deleted += 1
        flash(f"Reset complete — {deleted} file{'s' if deleted != 1 else ''} deleted.", "success")
    except Exception:
        tb = _tb.format_exc()
        print("=== data_import_reset ERROR ===", file=sys.stderr)
        print(tb, file=sys.stderr)
        flash(f"Reset failed: {tb.splitlines()[-1]}", "danger")
    return redirect(url_for("admin.data_import"))


# ── DB Maintenance helpers ───────────────────────────────────────────────────

@admin_bp.route("/admin/api/backfill-return-status", methods=["GET", "POST"])
def backfill_return_status():
    """Backfill return_status = 'DC GENERATED' for every wo_product_detail row
    where dc_number is not empty/null but return_status is still empty/null.

    Only dc_number drives DC GENERATED — dc_lenovo alone is never sufficient.

    GET  → dry-run: returns {"count": N} (how many rows would be updated).
    POST → applies the UPDATE and returns {"updated": N}.
    """
    from app.services.database.db import get_db
    conn = get_db()
    try:
        if request.method == "GET":
            row = conn.execute(
                """SELECT COUNT(*) FROM wo_product_detail
                    WHERE dc_number IS NOT NULL
                      AND TRIM(dc_number) NOT IN ('', '0')
                      AND (return_status IS NULL OR TRIM(return_status) = '')"""
            ).fetchone()
            return jsonify({"count": row[0] if row else 0})
        else:
            cur = conn.execute(
                """UPDATE wo_product_detail
                      SET return_status = 'DC GENERATED'
                    WHERE dc_number IS NOT NULL
                      AND TRIM(dc_number) NOT IN ('', '0')
                      AND (return_status IS NULL OR TRIM(return_status) = '')"""
            )
            conn.commit()
            return jsonify({"updated": cur.rowcount})
    except Exception as exc:
        import traceback as _tb
        current_app.logger.error("backfill_return_status failed:\n%s", _tb.format_exc())
        return jsonify({"error": str(exc)}), 500


# ── Validation Center ────────────────────────────────────────────────────────

@admin_bp.route("/admin/validation", methods=["GET"])
def validation():
    return render_template("admin/validation_center.html",
                           portal="admin", active_page="validation",
                           active_group="validation_center")


@admin_bp.route("/admin/validation/pou-unreturn", methods=["GET"])
def pou_unreturn_report():
    from app.services.database.queries import get_pou_unreturn_report

    upload_folder = current_app.config["EXCEL_UPLOAD_FOLDER"]
    meta_folder   = current_app.config["UPLOAD_META_FOLDER"]
    _POU_CATEGORY = "ID-IBM ID POU Unreturn"

    # Find the most recently uploaded POU Unreturn file
    pou_file_path = None
    pou_filename  = None
    for fname in sorted(os.listdir(upload_folder), reverse=True):
        if not allowed_excel(fname):
            continue
        meta = read_meta(meta_folder, fname)
        if meta and meta.get("file_category") == _POU_CATEGORY:
            pou_file_path = os.path.join(upload_folder, fname)
            pou_filename  = fname
            break

    if pou_file_path is None:
        return render_template(
            "admin/verification_center/pou_unreturn_report.html",
            rows=[], no_file=True,
            portal="admin", active_page="pou_unreturn",
            active_group="validation_center",
        )

    # Query DB rows with is_exist_excel = 'yes'.
    rows = get_pou_unreturn_report()

    # Build unique filter option lists server-side (avoids Jinja list-append quirks)
    rs_options  = sorted({r["return_status"]         for r in rows if r["return_status"]})
    lrs_options = sorted({r["lenovo_return_status"]  for r in rows if r["lenovo_return_status"]})

    # ── Summary counts ────────────────────────────────────────────────────────
    # Total SOID — all rows in the report
    summary_total = len(rows)

    # DC + AWB Filled — at least one of dc_number/dc_lenovo AND at least one of awb_resolv/awb_return
    def _has_val(r, key):
        v = r.get(key)
        return bool(v and str(v).strip() not in ("", "—"))

    summary_dc_awb = sum(
        1 for r in rows
        if (_has_val(r, "dc_number") or _has_val(r, "dc_lenovo"))
        and (_has_val(r, "awb_resolv") or _has_val(r, "awb_return"))
    )

    # Pending Partner / DC Generate — return_status is PENDING WITH PARTNER, PENDING FOR DC GENERATION,
    # UNKNOWN, or empty AND both dc_number (DC Resolve) and dc_lenovo (DC Lenovo) are empty
    _PENDING_STATUSES = {"PENDING WITH PARTNER", "PENDING FOR DC GENERATION", "UNKNOWN", ""}
    summary_pending = sum(
        1 for r in rows
        if (r.get("return_status") or "").strip().upper() in _PENDING_STATUSES
        and not _has_val(r, "dc_number")
        and not _has_val(r, "dc_lenovo")
    )

    # Missing AWB — both awb_resolv and awb_return are NULL or empty
    summary_no_awb = sum(
        1 for r in rows
        if not _has_val(r, "awb_resolv") and not _has_val(r, "awb_return")
    )

    return render_template(
        "admin/verification_center/pou_unreturn_report.html",
        rows=rows, no_file=False,
        pou_filename=pou_filename,
        rs_options=rs_options,
        lrs_options=lrs_options,
        summary_total=summary_total,
        summary_dc_awb=summary_dc_awb,
        summary_pending=summary_pending,
        summary_no_awb=summary_no_awb,
        portal="admin", active_page="pou_unreturn",
        active_group="validation_center",
    )



@admin_bp.route("/admin/validation/pou-unreturn/awb-notes", methods=["PATCH"])
def pou_unreturn_update_awb_notes():
    """Update awb_notes for a single SOID in wo_product_detail."""
    from app.services.database.connection import get_db
    data  = request.get_json(silent=True) or {}
    soid  = (data.get("soid") or "").strip()
    notes = (data.get("awb_notes") or "").strip() or None
    if not soid:
        return jsonify({"ok": False, "error": "soid is required."}), 400
    conn = get_db()
    cur  = conn.execute(
        "UPDATE wo_product_detail SET awb_notes = ? WHERE soid = ?",
        (notes, soid),
    )
    conn.commit()
    if cur.rowcount == 0:
        return jsonify({"ok": False, "error": "SOID not found."}), 404
    return jsonify({"ok": True, "soid": soid, "awb_notes": notes})



@admin_bp.route("/admin/validation/pou-unreturn/generate-report", methods=["GET"])
def pou_unreturn_generate_report():
    """Clone the source POU Unreturn Excel file and fill columns Q/R/S/T
    (DC/Collection Form, AWB Number, Return Status, Note) from the database,
    mirroring exactly what the report page displays.

    DC/Collection Form  → dc_number  (green = has value) else dc_lenovo (green)
    AWB Number          → awb_resolv (green = has value) else awb_return (green)
    Return Status       → Computed Lenovo Return Status (same logic as page JS)
    Note                → awb_notes
    """
    import io
    import math as _math
    from datetime import date as _date, datetime as _datetime
    from openpyxl import load_workbook
    from app.services.database.queries import get_pou_unreturn_report

    upload_folder = current_app.config["EXCEL_UPLOAD_FOLDER"]
    meta_folder   = current_app.config["UPLOAD_META_FOLDER"]
    _POU_CATEGORY = "ID-IBM ID POU Unreturn"

    # ── Find the source file ──────────────────────────────────────────────────
    pou_file_path = None
    pou_filename  = None
    for fname in sorted(os.listdir(upload_folder), reverse=True):
        if not allowed_excel(fname):
            continue
        meta = read_meta(meta_folder, fname)
        if meta and meta.get("file_category") == _POU_CATEGORY:
            pou_file_path = os.path.join(upload_folder, fname)
            pou_filename  = fname
            break

    if pou_file_path is None:
        return jsonify({"error": "No POU Unreturn file uploaded yet."}), 404

    # ── DB rows ───────────────────────────────────────────────────────────────
    rows = get_pou_unreturn_report()

    def _has(v) -> bool:
        if v is None:
            return False
        if isinstance(v, float) and _math.isnan(v):
            return False
        return str(v).strip() not in ("", "—", "0")

    # ── Replicate the JS overrideLRS() logic in Python ────────────────────────
    def _compute_display_lrs(r: dict) -> str:
        """Mirror the client-side Lenovo Return Status override logic exactly."""
        today = _date.today()
        dc_num_val  = str(r.get("dc_number")  or "").strip()
        dc_len_val  = str(r.get("dc_lenovo")  or "").strip()
        awb_res_val = str(r.get("awb_resolv") or "").strip()
        awb_ret_val = str(r.get("awb_return") or "").strip()
        lrs_db      = str(r.get("lenovo_return_status") or "").strip()
        wo_date_str = str(r.get("completion_date") or r.get("closing_date") or "").strip()
        dc_date_str = str(r.get("dc_generate_date") or "").strip()

        def _is_hardclose(v: str) -> bool:
            return bool(v) and "hardclose" in v.lower()

        def _is_real(v: str) -> bool:
            return bool(v) and not _is_hardclose(v)

        # Rule 1 — hardclose
        if _is_hardclose(dc_len_val) or _is_hardclose(awb_ret_val):
            return "Required Hard Close"

        # Rule 2 — DC + AWB both filled, WO Complete > 1 month ago
        dc_filled  = _is_real(dc_num_val)  or _is_real(dc_len_val)
        awb_filled = _is_real(awb_res_val) or _is_real(awb_ret_val)
        if dc_filled and awb_filled and wo_date_str:
            try:
                wo_date = _datetime.strptime(wo_date_str[:10], "%Y-%m-%d").date()
                m = wo_date.month + 1
                y = wo_date.year + (1 if m > 12 else 0)
                m = m if m <= 12 else 1
                one_month_after = wo_date.replace(year=y, month=m)
                if today > one_month_after:
                    return "Picked Up by Logistics"
            except (ValueError, OverflowError):
                pass

        # Rules 3 & 4 — DC Generate Date present
        if dc_date_str:
            try:
                dc_date = _datetime.strptime(dc_date_str[:10], "%Y-%m-%d").date()
                return "Pending Pickup" if (today - dc_date).days < 7 else "Picked Up by Logistics"
            except (ValueError, OverflowError):
                pass

        # No override — return DB value
        return lrs_db

    # ── Build SOID → display-values lookup ───────────────────────────────────
    db_lookup: dict = {}
    for r in rows:
        soid = r.get("soid")
        # DC: prefer dc_number (green when filled), fallback to dc_lenovo
        dc_val    = str(r.get("dc_number") or "").strip() \
                    if _has(r.get("dc_number")) \
                    else (str(r.get("dc_lenovo") or "").strip() or None)
        # AWB: prefer awb_resolv (green when filled), fallback to awb_return
        awb_val   = str(r.get("awb_resolv") or "").strip() \
                    if _has(r.get("awb_resolv")) \
                    else (str(r.get("awb_return") or "").strip() or None)
        lrs_val   = _compute_display_lrs(r) or None
        notes_val = str(r.get("awb_notes") or "").strip() or None
        db_lookup[str(soid)] = {
            "dc":    dc_val or None,
            "awb":   awb_val,
            "lrs":   lrs_val,
            "notes": notes_val,
        }

    # ── Clone the workbook in-memory and fill Q/R/S/T ─────────────────────────
    wb = load_workbook(pou_file_path)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

    # Discover SOID column index dynamically
    soid_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col_idx).value or "").strip().upper() == "SOID":
            soid_col_idx = col_idx
            break
    if soid_col_idx is None:
        return jsonify({"error": "SOID column not found in source Excel."}), 500

    # Discover target columns dynamically by header name
    target_cols = {
        "DC/Collection Form": None,
        "AWB Number":         None,
        "Return Status":      None,
        "Note":               None,
    }
    for col_idx in range(1, ws.max_column + 1):
        hdr = str(ws.cell(row=1, column=col_idx).value or "").strip()
        if hdr in target_cols:
            target_cols[hdr] = col_idx

    col_dc    = target_cols["DC/Collection Form"]
    col_awb   = target_cols["AWB Number"]
    col_rs    = target_cols["Return Status"]
    col_notes = target_cols["Note"]

    if not all([col_dc, col_awb, col_rs, col_notes]):
        missing = [k for k, v in target_cols.items() if v is None]
        return jsonify({"error": f"Missing columns in source Excel: {missing}"}), 500

    # Write values row-by-row matching on SOID
    for row_idx in range(2, ws.max_row + 1):
        raw_soid = ws.cell(row=row_idx, column=soid_col_idx).value
        if raw_soid is None:
            continue
        soid_str = str(int(raw_soid)) if isinstance(raw_soid, (int, float)) else str(raw_soid).strip()
        entry = db_lookup.get(soid_str)
        if entry is None:
            continue
        if entry["dc"]    is not None:
            ws.cell(row=row_idx, column=col_dc).value    = entry["dc"]
        if entry["awb"]   is not None:
            ws.cell(row=row_idx, column=col_awb).value   = entry["awb"]
        if entry["lrs"]   is not None:
            ws.cell(row=row_idx, column=col_rs).value    = entry["lrs"]
        if entry["notes"] is not None:
            ws.cell(row=row_idx, column=col_notes).value = entry["notes"]

    # ── Stream to response ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stem = os.path.splitext(pou_filename)[0]
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{stem}_Generated.xlsx",
    )


@admin_bp.route("/admin/validation/pou-unreturn/export", methods=["GET"])
def pou_unreturn_export():
    """Generate a styled .xlsx export of the PoU Unreturn Report."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from app.services.database.queries import get_pou_unreturn_report

    rows = get_pou_unreturn_report()

    # ── Palette ───────────────────────────────────────────────────────────────
    CLR_HEADER_BG  = "1F2328"   # dark header row
    CLR_HEADER_FG  = "FFFFFF"
    CLR_GREEN_CELL = "DCFCE7"
    CLR_RED_CELL   = "FEE2E2"
    CLR_AMBER_CELL = "FEF9C3"
    CLR_BLUE_CELL  = "DBEAFE"
    CLR_GREY_CELL  = "E5E7EB"
    CLR_MUTED_CELL = "F7F8FA"
    CLR_SUMM_BG    = "F7F8FA"

    def _fill(hex_col):
        return PatternFill("solid", fgColor=hex_col)

    def _border():
        thin = Side(style="thin", color="E5E7EB")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ══ Sheet 1 — Report data ═════════════════════════════════════════════════
    ws = wb.active
    ws.title = "POU Unreturn Report"

    col_headers = [
        "#", "SOID", "Completion / Closing Date", "Return Status",
        "DC Resolve (GTAAP)", "DC Lenovo", "AWB Return",
        "AWB Notes", "Lenovo Return Status", "Exist on Excel?"
    ]
    col_widths = [6, 18, 24, 22, 20, 16, 20, 20, 24, 16]

    # Header row
    ws.row_dimensions[1].height = 28
    for ci, (hdr, w) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill      = _fill(CLR_HEADER_BG)
        cell.font      = Font(bold=True, color=CLR_HEADER_FG, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    for ri, r in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 18
        date_val    = r.get("completion_date") or r.get("closing_date") or ""
        rs          = r.get("return_status") or ""
        dc_number   = r.get("dc_number") or ""
        dc_lenovo   = r.get("dc_lenovo") or ""
        awb_return  = r.get("awb_return") or ""
        awb_notes   = r.get("awb_notes") or ""
        lrs         = r.get("lenovo_return_status") or ""
        is_exist    = r.get("is_exist_excel") or ""
        rs_up       = rs.upper()

        row_data = [
            ri - 1,
            r.get("soid") or "",
            date_val[:10] if date_val else "",
            rs, dc_number, dc_lenovo,
            awb_return, awb_notes, lrs, is_exist,
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(size=10)

            # Column-specific colour rules
            if ci == 4:   # Return Status
                if "PENDING" in rs_up or rs_up == "UNKNOWN":
                    cell.fill = _fill(CLR_AMBER_CELL)
                elif rs:
                    cell.fill = _fill(CLR_BLUE_CELL)
                else:
                    cell.fill = _fill(CLR_MUTED_CELL)
            elif ci == 5: # DC Resolve
                cell.fill = _fill(CLR_GREEN_CELL) if dc_number else _fill(CLR_RED_CELL)
            elif ci == 6: # DC Lenovo
                if dc_number:
                    cell.fill = _fill(CLR_GREY_CELL)
                elif dc_lenovo:
                    cell.fill = _fill(CLR_GREEN_CELL)
                else:
                    cell.fill = _fill(CLR_MUTED_CELL)
            elif ci == 7: # AWB Return
                cell.fill = _fill(CLR_GREEN_CELL) if awb_return else _fill(CLR_RED_CELL)
            elif ci == 8: # AWB Notes
                cell.fill = _fill(CLR_GREEN_CELL) if awb_notes else _fill(CLR_MUTED_CELL)

        # Alternate row background for readability (very light)
        if ri % 2 == 0:
            for ci in range(1, len(row_data) + 1):
                c = ws.cell(row=ri, column=ci)
                if c.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                    c.fill = _fill("F9FAFB")

    # Freeze header row
    ws.freeze_panes = "A2"

    # ══ Sheet 2 — Summary ═════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14

    def _summ_hdr(row, label):
        c = ws2.cell(row=row, column=1, value=label)
        c.font      = Font(bold=True, size=11, color="1F2328")
        c.fill      = _fill(CLR_SUMM_BG)
        c.alignment = Alignment(vertical="center")
        c.border    = _border()
        ws2.row_dimensions[row].height = 22

    def _summ_val(row, val, fill_hex=CLR_SUMM_BG):
        c = ws2.cell(row=row, column=2, value=val)
        c.font      = Font(bold=True, size=11, color="1F2328")
        c.fill      = _fill(fill_hex)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()

    total_soid    = len(rows)
    total_dc_awb  = sum(
        1 for r in rows
        if (r.get("awb_return") or "").strip() not in ("", "—")
        and (r.get("dc_number") or r.get("dc_lenovo"))
    )
    total_pending = sum(
        1 for r in rows
        if (r.get("return_status") or "").strip().upper() != "DC GENERATED"
    )
    total_no_awb  = sum(
        1 for r in rows
        if not (r.get("awb_return") or "").strip()
        or (r.get("awb_return") or "").strip() == "—"
    )

    # Title
    ws2.merge_cells("A1:B1")
    title_cell = ws2["A1"]
    title_cell.value     = "PoU Unreturn Report — Summary"
    title_cell.font      = Font(bold=True, size=13, color=CLR_HEADER_FG)
    title_cell.fill      = _fill(CLR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border    = _border()
    ws2.row_dimensions[1].height = 30

    _summ_hdr(2, "Total SOID");        _summ_val(2, total_soid)
    _summ_hdr(3, "DC + AWB Filled");   _summ_val(3, total_dc_awb,  "DCFCE7")
    _summ_hdr(4, "Pending Partner / DC Generate"); _summ_val(4, total_pending, "FEF9C3")
    _summ_hdr(5, "Missing AWB");       _summ_val(5, total_no_awb,  "FEE2E2")

    # ── Stream to response ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="POU_Unreturn_Report.xlsx",
    )


# ── User & ASP Management ────────────────────────────────────────────────────

@admin_bp.route("/admin/users", methods=["GET"])
def users():
    return render_template("admin/user_management.html",
                           portal="admin", active_page="user_mgmt", active_group="user_mgmt")


# Kota → granular Java region mapping.
# Anything on Java that doesn't match stays as "Jawa".
_JABODETABEK = {
    "jakarta pusat", "jakarta barat", "jakarta selatan", "jakarta utara",
    "jakarta timur", "bogor", "depok", "tangerang", "bekasi", "banten",
    "karawang", "cikarang",
}
_JAWA_BARAT = {
    "bandung", "cirebon", "tasikmalaya", "garut", "sukabumi",
    "cianjur", "purwakarta", "subang", "indramayu",
}
_JAWA_TENGAH = {
    "semarang", "kudus", "pati", "tegal", "surakarta", "solo",
    "magelang", "purwokerto", "salatiga", "klaten", "wonosobo",
    "banyumas", "cilacap",
}
_JAWA_TIMUR = {
    "surabaya", "malang", "kediri", "jember", "madiun", "mojokerto",
    "pasuruan", "probolinggo", "blitar", "banyuwangi",
}
_DIY = {"yogyakarta", "sleman", "bantul", "gunung kidul", "kulonprogo"}


def _resolve_region(kota: str | None, island: str | None) -> str:
    """Return a granular region label for a given kota / island pair."""
    if kota:
        k = kota.strip().lower()
        if k in _JABODETABEK:
            return "Jabodetabek"
        if k in _JAWA_BARAT:
            return "Jawa Barat"
        if k in _JAWA_TENGAH:
            return "Jawa Tengah"
        if k in _JAWA_TIMUR:
            return "Jawa Timur"
        if k in _DIY:
            return "DIY"
    # Fall back to the island column (strip + title-case)
    if island and island.strip():
        raw = island.strip()
        # Normalise casing inconsistencies (e.g. "jawa" → "Jawa")
        return raw[0].upper() + raw[1:]
    return "—"


@admin_bp.route("/admin/users/asp-directory", methods=["GET"])
def asp_directory():
    db_path = current_app.config["DATABASE_PATH"]
    conn = open_db(db_path)

    # Refresh wo_count for all ASPs in one UPDATE pass
    conn.execute(
        """
        UPDATE asp_details
        SET wo_count = (
            SELECT COUNT(*)
            FROM wo_details
            WHERE wo_details.labor_vendor_related = asp_details.labor_vendor_related
        )
        """
    )
    conn.commit()

    rows = conn.execute(
        "SELECT * FROM asp_details ORDER BY id"
    ).fetchall()
    conn.close()

    # Attach computed region to each row as a plain dict so the template
    # can access asp.region without modifying the DB schema.
    asps = []
    for r in rows:
        d = dict(r)
        d["region"] = _resolve_region(d.get("kota"), d.get("island"))
        asps.append(d)

    regions = sorted({a["region"] for a in asps if a["region"] != "—"})

    return render_template(
        "admin/user_management/asp_directory.html",
        asps=asps,
        regions=regions,
        portal="admin",
        active_page="asp_directory",
        active_group="user_mgmt",
    )


@admin_bp.route("/admin/users/asp-directory/<int:asp_id>/edit", methods=["POST"])
def asp_directory_edit(asp_id):
    fields = [
        "username", "password", "vendor_code", "service_provider",
        "parent_group", "labor_vendor_related", "customer_partner",
        "store_name", "kota", "address", "lat_long", "link_map",
        "phone_number", "island", "working_hours", "operational_status",
        "future_status", "operation_support", "office_type",
    ]
    values = {f: request.form.get(f, "").strip() or None for f in fields}
    set_clause = ", ".join(f"{f} = ?" for f in fields)
    params = [values[f] for f in fields] + [asp_id]

    db_path = current_app.config["DATABASE_PATH"]
    conn = open_db(db_path)
    try:
        conn.execute(
            f"UPDATE asp_details SET {set_clause} WHERE id = ?", params
        )
        conn.commit()
        flash(f"ASP record #{asp_id} updated successfully.", "success")
    finally:
        conn.close()
    return redirect(url_for("admin.asp_directory"))


@admin_bp.route("/admin/users/asp-directory/create", methods=["POST"])
def asp_directory_create():
    fields = [
        "username", "password", "vendor_code", "service_provider",
        "parent_group", "labor_vendor_related", "customer_partner",
        "store_name", "kota", "address", "lat_long", "link_map",
        "phone_number", "island", "working_hours", "operational_status",
        "future_status", "operation_support", "office_type",
    ]
    values = {f: request.form.get(f, "").strip() or None for f in fields}

    # Require username and labor_vendor_related
    if not values.get("username") or not values.get("labor_vendor_related"):
        flash("Username and ASP ID are required to create a new ASP.", "danger")
        return redirect(url_for("admin.asp_directory"))

    cols   = ", ".join(fields)
    placeholders = ", ".join("?" for _ in fields)
    params = [values[f] for f in fields]

    db_path = current_app.config["DATABASE_PATH"]
    conn = open_db(db_path)
    try:
        conn.execute(
            f"INSERT INTO asp_details ({cols}) VALUES ({placeholders})", params
        )
        conn.commit()
        flash(f"ASP \"{values['username']}\" created successfully.", "success")
    except sqlite3.IntegrityError:
        flash(f"Username \"{values['username']}\" already exists in asp_details.", "danger")
    finally:
        conn.close()
    return redirect(url_for("admin.asp_directory"))


# ── ASP Users (admin read) ───────────────────────────────────────────────────

@admin_bp.route("/admin/users/asp-directory/<labor_vendor_related>/users", methods=["GET"])
def asp_directory_users(labor_vendor_related):
    """Return the asp_users list for a given ASP as JSON (admin use)."""
    from app.services.database.db import get_db
    db = get_db()
    rows = db.execute(
        """SELECT id, tech_id, full_name, email, phone_number, is_active, created_at
           FROM asp_users
           WHERE labor_vendor_related = ?
           ORDER BY id""",
        (labor_vendor_related,)
    ).fetchall()
    return jsonify({"ok": True, "users": [dict(r) for r in rows]})


# ── ASP Password Change Requests ─────────────────────────────────────────────

@admin_bp.route("/admin/users/pw-change-requests", methods=["GET"])
def pw_change_requests():
    """List the auto-approved ASP password change history."""
    from app.services.database.db import get_db
    db = get_db()
    history_rows = db.execute(
        """SELECT r.id, r.asp_username, r.requested_at, r.status,
                  r.new_password, r.reviewed_by, r.reviewed_at,
                  d.service_provider
           FROM asp_pw_change_requests r
           LEFT JOIN asp_details d ON d.username = r.asp_username
           ORDER BY r.requested_at DESC"""
    ).fetchall()
    return render_template(
        "admin/user_management/pw_change_requests.html",
        history=[dict(r) for r in history_rows],
        portal="admin",
        active_page="pw_change_requests",
        active_group="user_mgmt",
    )


@admin_bp.route("/admin/users/pw-change-requests/<int:req_id>/approve", methods=["POST"])
def pw_change_request_approve(req_id):
    """Approve a password change request: apply the ASP's requested password."""
    from app.services.database.db import get_db
    from flask import session as _session
    db = get_db()
    row = db.execute(
        "SELECT asp_username, new_password FROM asp_pw_change_requests "
        "WHERE id=? AND status='pending'",
        (req_id,)
    ).fetchone()
    if not row:
        flash("Request not found or already reviewed.", "danger")
        return redirect(url_for("admin.pw_change_requests"))
    asp_username = row["asp_username"]
    new_password = row["new_password"]
    if not new_password:
        flash("No password was submitted with this request.", "danger")
        return redirect(url_for("admin.pw_change_requests"))
    reviewed_by = _session.get("username", "admin")
    db.execute(
        "UPDATE asp_details SET password=? WHERE username=?",
        (new_password, asp_username)
    )
    db.execute(
        "UPDATE asp_pw_change_requests "
        "SET status='approved', reviewed_by=?, reviewed_at=datetime('now') "
        "WHERE id=?",
        (reviewed_by, req_id)
    )
    db.commit()
    flash(f"Password for {asp_username} approved and applied successfully.", "success")
    return redirect(url_for("admin.pw_change_requests"))


@admin_bp.route("/admin/users/pw-change-requests/<int:req_id>/deny", methods=["POST"])
def pw_change_request_deny(req_id):
    """Deny a password change request."""
    from app.services.database.db import get_db
    from flask import session as _session
    db = get_db()
    reviewed_by = _session.get("username", "admin")
    db.execute(
        "UPDATE asp_pw_change_requests "
        "SET status='denied', reviewed_by=?, reviewed_at=datetime('now') "
        "WHERE id=? AND status='pending'",
        (reviewed_by, req_id)
    )
    db.commit()
    flash("Request denied.", "warning")
    return redirect(url_for("admin.pw_change_requests"))


# ── System Archive ───────────────────────────────────────────────────────────

@admin_bp.route("/admin/archive", methods=["GET"])
def archive():
    files = list_excel_uploads()
    for f in files:
        f["modified_fmt"] = datetime.fromtimestamp(f["modified"]).strftime("%Y-%m-%d %H:%M")
    return render_template("admin/system_archive.html",
                           masterfiles=[],
                           uploaded_files=files,
                           portal="admin", active_page="archive")


@admin_bp.route("/admin/archive/download/masterfile/<filename>", methods=["GET"])
def archive_download(filename):
    from werkzeug.utils import secure_filename
    safe     = secure_filename(filename)
    filepath = os.path.join(current_app.config["EXCELS_DIR"], safe)
    if not os.path.isfile(filepath):
        flash("File not found.", "danger")
        return redirect(url_for("admin.archive"))
    return send_file(filepath, as_attachment=True, download_name=safe,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Superadmin Users (admin_users table, role = 'superadmin') ────────────────

@admin_bp.route("/admin/api/superadmin-users", methods=["GET"])
def api_superadmin_users_list():
    """Return all superadmin accounts (role = 'superadmin') for the profile page."""
    from flask import session as _session
    if _session.get("role") != "superadmin":
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    from app.services.database.db import get_db
    db = get_db()
    rows = db.execute(
        "SELECT id, username, full_name, email, is_active, created_at "
        "FROM admin_users WHERE role = 'superadmin' ORDER BY id"
    ).fetchall()
    return jsonify({"ok": True, "users": [dict(r) for r in rows]})


@admin_bp.route("/admin/api/superadmin-users", methods=["POST"])
def api_superadmin_users_create():
    """Create a new superadmin account in admin_users."""
    from flask import session as _session
    if _session.get("role") != "superadmin":
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    from app.services.database.db import get_db
    data     = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    fullname = (data.get("full_name") or "").strip()
    email    = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()

    if not username:
        return jsonify({"ok": False, "error": "Username is required.", "field": "username"})
    if not password or len(password) < 8:
        return jsonify({"ok": False, "error": "Password must be at least 8 characters.", "field": "password"})

    db = get_db()
    existing = db.execute(
        "SELECT id FROM admin_users WHERE LOWER(username) = LOWER(?)", (username,)
    ).fetchone()
    if existing:
        return jsonify({"ok": False, "error": "Username already exists.", "field": "username"})

    db.execute(
        "INSERT INTO admin_users (username, password, full_name, email, role, is_active) "
        "VALUES (?, ?, ?, ?, 'superadmin', 1)",
        (username, password, fullname or None, email or None),
    )
    db.commit()
    row = db.execute(
        "SELECT id, username, full_name, email, is_active, created_at "
        "FROM admin_users WHERE LOWER(username) = LOWER(?)", (username,)
    ).fetchone()
    return jsonify({"ok": True, "user": dict(row)})


@admin_bp.route("/admin/api/superadmin-users/<int:uid>", methods=["PUT"])
def api_superadmin_users_update(uid):
    """Update username / full_name / email / password for a superadmin account."""
    from flask import session as _session
    if _session.get("role") != "superadmin":
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    # Prevent a superadmin from editing themselves via this endpoint (use profile page instead)
    if _session.get("user_id") == uid:
        return jsonify({"ok": False, "error": "Cannot edit your own account here."})
    from app.services.database.db import get_db
    data     = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    fullname = (data.get("full_name") or "").strip()
    email    = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()

    if not username:
        return jsonify({"ok": False, "error": "Username is required.", "field": "username"})

    db = get_db()
    # Confirm target exists and is superadmin
    target = db.execute(
        "SELECT id FROM admin_users WHERE id = ? AND role = 'superadmin'", (uid,)
    ).fetchone()
    if not target:
        return jsonify({"ok": False, "error": "User not found."})

    # Uniqueness check (excluding self)
    clash = db.execute(
        "SELECT id FROM admin_users WHERE LOWER(username) = LOWER(?) AND id != ?",
        (username, uid)
    ).fetchone()
    if clash:
        return jsonify({"ok": False, "error": "Username already taken.", "field": "username"})

    if password:
        if len(password) < 8:
            return jsonify({"ok": False, "error": "Password must be at least 8 characters.", "field": "password"})
        db.execute(
            "UPDATE admin_users SET username=?, full_name=?, email=?, password=? WHERE id=?",
            (username, fullname or None, email or None, password, uid),
        )
    else:
        db.execute(
            "UPDATE admin_users SET username=?, full_name=?, email=? WHERE id=?",
            (username, fullname or None, email or None, uid),
        )
    db.commit()
    row = db.execute(
        "SELECT id, username, full_name, email, is_active, created_at FROM admin_users WHERE id=?", (uid,)
    ).fetchone()
    return jsonify({"ok": True, "user": dict(row)})


@admin_bp.route("/admin/api/superadmin-users/<int:uid>/status", methods=["PATCH"])
def api_superadmin_users_status(uid):
    """Toggle is_active for a superadmin account."""
    from flask import session as _session
    if _session.get("role") != "superadmin":
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    if _session.get("user_id") == uid:
        return jsonify({"ok": False, "error": "Cannot change your own account status."})
    from app.services.database.db import get_db
    data      = request.get_json(silent=True) or {}
    is_active = 1 if data.get("is_active") else 0
    db = get_db()
    target = db.execute(
        "SELECT id FROM admin_users WHERE id = ? AND role = 'superadmin'", (uid,)
    ).fetchone()
    if not target:
        return jsonify({"ok": False, "error": "User not found."})
    db.execute("UPDATE admin_users SET is_active=? WHERE id=?", (is_active, uid))
    db.commit()
    return jsonify({"ok": True, "is_active": is_active})


@admin_bp.route("/admin/api/superadmin-users/<int:uid>", methods=["DELETE"])
def api_superadmin_users_delete(uid):
    """Permanently delete a superadmin account."""
    from flask import session as _session
    if _session.get("role") != "superadmin":
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    if _session.get("user_id") == uid:
        return jsonify({"ok": False, "error": "Cannot delete your own account."})
    from app.services.database.db import get_db
    db = get_db()
    target = db.execute(
        "SELECT id FROM admin_users WHERE id = ? AND role = 'superadmin'", (uid,)
    ).fetchone()
    if not target:
        return jsonify({"ok": False, "error": "User not found."})
    db.execute("DELETE FROM admin_users WHERE id=?", (uid,))
    db.commit()
    return jsonify({"ok": True})


# ── Escalation Center (Monday.com Sync Manager) ──────────────────────────────
# Background sync state — module-level so it persists across requests

SYNC_INTERVAL_SEC = 30 * 60   # 30 minutes
SYNC_TZ_OFFSET    = 7 * 3600  # WIB = UTC+7
SYNC_HOUR_START   = 6         # 06:00 local
SYNC_HOUR_END     = 20        # 20:00 local (exclusive)

_sync_thread: threading.Thread | None = None
_sync_stop        = threading.Event()
_log_queue: _queue.Queue = _queue.Queue(maxsize=2000)
_sync_lock        = threading.Lock()
_scheduler_thread: threading.Thread | None = None
_scheduler_stop   = threading.Event()
_next_run_at: float | None = None      # Unix ts of next scheduled sync run
_window_opens_at: float | None = None  # Unix ts of next active-window open (when idle/off-hours)


def _local_now():
    """Current time as a datetime in WIB (UTC+7), timezone-naive."""
    from datetime import datetime, timezone, timedelta
    tz_wib = timezone(timedelta(seconds=SYNC_TZ_OFFSET))
    return datetime.now(tz_wib).replace(tzinfo=None)


def _is_active_window(dt=None):
    """Return True if dt (or now) is a weekday between SYNC_HOUR_START and SYNC_HOUR_END."""
    if dt is None:
        dt = _local_now()
    return dt.weekday() < 5 and SYNC_HOUR_START <= dt.hour < SYNC_HOUR_END


def _next_window_open_ts():
    """Return Unix timestamp (UTC) of when the next active window begins."""
    from datetime import datetime, timezone, timedelta
    tz_wib = timezone(timedelta(seconds=SYNC_TZ_OFFSET))
    now_local = _local_now()
    # Start from tomorrow if today's window already passed or it's a weekend
    candidate = now_local.replace(hour=SYNC_HOUR_START, minute=0, second=0, microsecond=0)
    if candidate <= now_local:
        # Today's window start is in the past — move to next day
        candidate = candidate + timedelta(days=1)
    # Advance past weekends (Saturday=5, Sunday=6)
    while candidate.weekday() >= 5:
        candidate = candidate + timedelta(days=1)
    # Convert back to UTC Unix timestamp
    candidate_aware = candidate.replace(tzinfo=tz_wib)
    return candidate_aware.timestamp()


_queue_formatter = logging.Formatter(datefmt="%Y-%m-%d %H:%M:%S")


class _QueueHandler(logging.Handler):
    """Logging handler that pushes records into the SSE queue."""
    def emit(self, record: logging.LogRecord) -> None:  # type: ignore[override]
        # Only forward records that originated from the monday_sync logger
        if not record.name.startswith("monday_sync"):
            return
        try:
            _log_queue.put_nowait({
                "ts":    _queue_formatter.formatTime(record, "%Y-%m-%d %H:%M:%S"),
                "level": record.levelname,
                "msg":   record.getMessage(),
            })
        except _queue.Full:
            # Queue full — drop oldest entry to make room for this one
            try:
                _log_queue.get_nowait()
                _log_queue.put_nowait({
                    "ts":    _queue_formatter.formatTime(record, "%Y-%m-%d %H:%M:%S"),
                    "level": record.levelname,
                    "msg":   record.getMessage(),
                })
            except (_queue.Full, _queue.Empty):
                pass


_queue_handler = _QueueHandler()
_queue_handler.setLevel(logging.DEBUG)


def _get_monday_sync_db():
    """Open a fresh SQLite connection to files/lenovo_asp.db.

    Also ensures the has_wo column exists on technical_escalation so the
    column can be queried immediately even before monday_sync.get_db() has
    run its own migration on this file.
    """
    import os as _os
    project_root = _os.path.normpath(
        _os.path.join(_os.path.dirname(__file__), "..", "..")
    )
    db_path = _os.path.join(project_root, "files", "lenovo_asp.db")
    if not _os.path.isfile(db_path):
        return None
    conn = open_db(db_path)
    # Ensure has_wo column exists (added in a later schema version)
    try:
        cols = {r[1] for r in conn.execute("PRAGMA table_info(technical_escalation)")}
        if "has_wo" not in cols:
            conn.execute("ALTER TABLE technical_escalation ADD COLUMN has_wo INTEGER DEFAULT NULL")
            conn.commit()
    except Exception:
        pass
    return conn


def _start_has_wo_backfill(app) -> None:
    """Spawn a one-shot daemon thread that stamps has_wo on all NULL rows.

    Runs once at startup — after that every row is already stamped, so the
    UPDATE touches zero rows and returns immediately.  Never called on every
    request; the sync scheduler also calls _stamp_has_wo after each sync.
    """
    import threading as _threading

    def _worker():
        import time as _time
        _time.sleep(5)          # let the app finish booting first
        try:
            from app.scripts.monday_sync import _stamp_has_wo
            conn = _get_monday_sync_db()
            if conn:
                try:
                    n = _stamp_has_wo(conn)
                    if n:
                        logging.getLogger("monday_sync").info(
                            "has_wo backfill: stamped %d row(s) at startup", n
                        )
                finally:
                    conn.close()
        except Exception as exc:
            logging.getLogger("monday_sync").warning("has_wo backfill failed: %s", exc)

    t = _threading.Thread(target=_worker, daemon=True, name="has_wo_backfill")
    t.start()







def _scheduler_loop(app) -> None:
    """
    Auto-sync loop — runs forever as a daemon thread.

    Guarantees:
      • Only runs during active window: Mon–Fri 06:00–20:00 WIB.
      • Outside the window it sleeps (interruptibly) until the window opens.
      • Never starts a new sync while the previous one is still running.
      • The 30-min cooldown starts ONLY after the previous sync finishes.
      • Responds to _scheduler_stop within 1 second for clean shutdown.
    """
    global _sync_thread, _next_run_at, _window_opens_at
    log = logging.getLogger("monday_sync")
    log.info("Auto-scheduler started — interval=%d min, window=%02d:00–%02d:00 WIB Mon–Fri",
             SYNC_INTERVAL_SEC // 60, SYNC_HOUR_START, SYNC_HOUR_END)

    # ── Startup delay: wait 15 minutes before first sync ────────────────────
    STARTUP_DELAY_SEC = 15 * 60
    log.info("Scheduler: startup delay — first sync in 15 min")
    _next_run_at = _time.time() + STARTUP_DELAY_SEC   # expose to status endpoint
    _scheduler_stop.wait(STARTUP_DELAY_SEC)
    _next_run_at = None
    if _scheduler_stop.is_set():
        log.info("Auto-scheduler stopped during startup delay.")
        return

    while not _scheduler_stop.is_set():

        # ── Step 0: enforce active-window gate ───────────────────────────────
        if not _is_active_window():
            opens_ts = _next_window_open_ts()
            wait_sec = max(0, opens_ts - _time.time())
            _window_opens_at = opens_ts
            _next_run_at     = None
            log.info("Scheduler: outside active window — sleeping %.0f min until window opens",
                     wait_sec / 60)
            _scheduler_stop.wait(wait_sec)
            _window_opens_at = None
            if _scheduler_stop.is_set():
                break
            continue   # re-check window at top of loop

        _window_opens_at = None   # we are inside the window

        # ── Step 1: wait for any in-progress sync (manual or previous scheduled)
        while True:
            with _sync_lock:
                running_thread = _sync_thread if (_sync_thread and _sync_thread.is_alive()) else None
            if running_thread is None:
                break
            log.info("Scheduler: sync already running, waiting for it to finish…")
            running_thread.join()

        if _scheduler_stop.is_set():
            break

        # ── Step 2: decide mode based on current DB contents
        edb = _get_monday_sync_db()
        item_count = 0
        if edb:
            try:
                item_count = edb.execute(
                    "SELECT COUNT(*) FROM technical_escalation"
                ).fetchone()[0]
            except Exception:
                pass
            finally:
                edb.close()
        mode = "incremental" if item_count > 0 else "full"

        # ── Step 3: claim the sync slot and start the thread
        with _sync_lock:
            if _sync_thread and _sync_thread.is_alive():
                continue
            _sync_stop.clear()
            t = threading.Thread(
                target=_run_sync_task,
                args=(app, mode, None),
                daemon=True,
                name="monday-sync-auto",
            )
            _sync_thread = t

        log.info("Scheduler: starting %s sync (%d items in DB)", mode, item_count)
        t.start()
        t.join()   # ← 30-min clock starts ONLY after this returns

        log.info("Scheduler: sync complete — next run in %d min", SYNC_INTERVAL_SEC // 60)

        # ── Step 4: interruptible sleep before next cycle
        # Only schedule next run if still inside the active window
        if _is_active_window():
            _next_run_at = _time.time() + SYNC_INTERVAL_SEC
            _scheduler_stop.wait(SYNC_INTERVAL_SEC)
            _next_run_at = None
        # If we've drifted outside the window, the loop will handle it on next iteration

    log.info("Auto-scheduler stopped.")


def start_sync_scheduler(app) -> None:
    """
    Start the background scheduler thread exactly once.
    Safe to call multiple times — subsequent calls are no-ops.
    """
    global _scheduler_thread
    if _scheduler_thread is not None and _scheduler_thread.is_alive():
        return
    _scheduler_stop.clear()
    _scheduler_thread = threading.Thread(
        target=_scheduler_loop,
        args=(app,),
        daemon=True,
        name="monday-scheduler",
    )
    _scheduler_thread.start()


def _run_sync_task(app, mode: str, board_id: str | None = None) -> None:
    """
    Background thread target: runs the Monday sync script.
    Attaches a _QueueHandler to the monday_sync logger so every log record
    is forwarded to the SSE queue.
    """
    import sys
    import os as _os

    project_root = _os.path.normpath(
        _os.path.join(_os.path.dirname(__file__), "..", "..")
    )
    scripts_dir = _os.path.join(project_root, "app", "scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)

    def _put(level: str, msg: str) -> None:
        """Put a log entry into the queue, dropping the oldest if full."""
        entry = {"ts": "", "level": level, "msg": msg}
        try:
            _log_queue.put_nowait(entry)
        except _queue.Full:
            try:
                _log_queue.get_nowait()
                _log_queue.put_nowait(entry)
            except (_queue.Full, _queue.Empty):
                pass

    try:
        import monday_sync as _ms
    except ImportError as exc:
        _put("ERROR", f"Cannot import monday_sync: {exc}")
        return

    # Wire queue handler into the sync script's logger
    _ms_logger = logging.getLogger("monday_sync")
    _ms_logger.propagate = False  # prevent leaking into root → _msd_log_queue
    _ms_logger.addHandler(_queue_handler)
    _ms_logger.setLevel(logging.DEBUG)

    xlsx_path = _os.path.join(project_root, "files", "source-db", "monday_link_map.xlsx")

    with app.app_context():
        try:
            conn  = _ms.get_db(_os.path.join(project_root, "files", "lenovo_asp.db"))
            state = _ms.load_state()
            boards = _ms.load_boards(xlsx_path)

            if board_id:
                boards = [b for b in boards if b["board_id"] == board_id]
                if not boards:
                    _put("ERROR", f"Board ID {board_id} not found in xlsx")
                    return

            if mode == "full":
                _ms.run_sync_all(conn, state, boards, force_full=True)
            elif mode == "incremental":
                _ms.run_sync_all(conn, state, boards, force_full=False)
            elif mode == "backfill":
                _ms._backfill_updates(conn)
            else:
                _put("ERROR", f"Unknown sync mode: {mode}")
                return

            _put("INFO", f"✓ Sync mode '{mode}' completed.")
        except Exception as exc:
            import traceback
            _put("ERROR", f"Sync error: {exc}")
            _put("ERROR", traceback.format_exc())
        finally:
            _ms_logger.removeHandler(_queue_handler)
            # Keep propagate=False permanently — restoring True would let Werkzeug
            # access log records leak into _msd_log_queue via the root logger.
            try:
                conn.close()
            except Exception:
                pass


def _run_msd_download_task(app, run_once: bool = False) -> None:
    import runpy
    import traceback
    from contextlib import redirect_stderr, redirect_stdout

    class _QueueWriter:
        def __init__(self, logger: logging.Logger, level: int) -> None:
            self._logger = logger
            self._level = level
            self._buffer = ""

        def write(self, value: str) -> int:
            self._buffer += value
            while "\n" in self._buffer:
                line, self._buffer = self._buffer.split("\n", 1)
                if line.strip():
                    self._logger.log(self._level, line)
            return len(value)

        def flush(self) -> None:
            if self._buffer.strip():
                self._logger.log(self._level, self._buffer.strip())
            self._buffer = ""

    project_root = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    )
    script_path = os.path.join(
        project_root,
        "app",
        "scripts",
        "msd-auto-download",
        "msd-auto-download.py",
    )

    logger = logging.getLogger("msd_auto_download")
    logger.setLevel(logging.INFO)
    logger.propagate = False  # prevent root-logger → stderr → _QueueWriter → logger loop
    queue_handler = _MsdQueueHandler()
    queue_handler.setLevel(logging.INFO)
    logger.addHandler(queue_handler)

    stdout_writer = _QueueWriter(logger, logging.INFO)
    stderr_writer = _QueueWriter(logger, logging.ERROR)

    # ── auto-upsert hook ─────────────────────────────────────────────────────
    # Called by the download script after a successful file move/copy.
    # Runs dispatch_upsert("WOID", ...) on the downloaded file and logs results.
    def _msd_auto_upsert(filepath: str, fname: str) -> None:
        try:
            from app.services.database.upsert import dispatch_upsert as _dispatch
            db_path = app.config["DATABASE_PATH"]
            logger.info("Auto-upsert: starting WOID upsert for %s …", fname)
            _db = open_db(db_path)
            try:
                n_new_wo, n_updated, n_usr = _dispatch("WOID", filepath, _db)
            finally:
                _db.close()
            from datetime import date as _date
            _today = str(_date.today())   # "YYYY-MM-DD"

            # Accumulate daily totals — reset when the date key changes
            if _today not in _msd_daily_totals:
                _msd_daily_totals.clear()   # drop all previous-day keys
                _msd_daily_totals[_today] = {"new_wo": 0, "updated_wo": 0, "new_users": 0}
            _msd_daily_totals[_today]["new_wo"]     += n_new_wo
            _msd_daily_totals[_today]["updated_wo"] += n_updated
            _msd_daily_totals[_today]["new_users"]  += n_usr

            # Store per-file stats (per-run, not cumulative)
            _msd_upsert_stats[fname] = {
                "new_wo":       n_new_wo,
                "updated_wo":   n_updated,
                "new_users":    n_usr,
                "status":       "success",
                "upsert_date":  _today,
            }
            _msd_stats_save()   # persist to disk immediately
            # Emit individual log lines — each appears in the live stream
            _day = _msd_daily_totals[_today]
            logger.info("Auto-upsert: ✅ Upsert complete for %s", fname)
            logger.info("Auto-upsert:    ↳ New WOs added      : %d  (today total: %d)", n_new_wo, _day["new_wo"])
            logger.info("Auto-upsert:    ↳ Existing WOs updated: %d (today total: %d)", n_updated, _day["updated_wo"])
            if n_usr:
                logger.info(
                    "Auto-upsert:    ↳ New technician%s added : %d",
                    "s" if n_usr != 1 else "", n_usr,
                )
        except Exception:
            import traceback as _tbk
            from datetime import date as _date_err
            _today_err = str(_date_err.today())
            _msd_upsert_stats[fname] = {
                "new_wo":       None,
                "updated_wo":   None,
                "new_users":    None,
                "status":       "failed",
                "upsert_date":  _today_err,
            }
            _msd_stats_save()   # persist failure status too
            logger.error("Auto-upsert: ❌ upsert failed for %s", fname)
            logger.error(_tbk.format_exc())

    # ── stdin shim ────────────────────────────────────────────────────────────
    # Replace builtins.input inside the script. Handles three cases:
    #   • OTP prompt      — shows OTP panel in browser, waits for /otp
    #   • OTP cancelled   — user clicked Cancel; close Chrome, preserve session
    #   • Re-login        — shows Re-login panel, waits for /relogin
    def _web_input(prompt: str = "") -> str:
        global _msd_otp_pending, _msd_relogin_pending
        if prompt == "__RELOGIN_WAIT__":
            # Session expired — emit sentinel, show Re-login panel, block
            logger.warning("__RELOGIN_REQUIRED__")  # picked up by SSE → re-login panel
            _msd_relogin_pending = True
            _msd_otp_queue.get()  # blocks until /relogin puts __RELOGIN_CONFIRMED__
            logger.info("Re-login confirmed by user.")
            return ""
        # Normal OTP flow
        if prompt:
            logger.info(prompt)
        logger.info("__OTP_REQUIRED__")   # sentinel picked up by SSE stream
        _msd_otp_pending = True
        code = _msd_otp_queue.get()       # blocks until /otp or /otp/cancel puts a value
        if code == "__OTP_CANCELLED__":
            # Raise so the except _OtpCancelledError block above closes Chrome cleanly.
            raise _OtpCancelledError("OTP cancelled by user.")
        logger.info("OTP code received from browser.")
        return code

    # Clear the history ring buffer so each new run starts with a clean log slate.
    _msd_log_history.clear()

    with app.app_context():
        try:
            logger.info("Starting MSD WO auto-download script...")
            original_argv = list(os.sys.argv)
            os.sys.argv = [script_path]
            with redirect_stdout(stdout_writer), redirect_stderr(stderr_writer):
                runpy.run_path(
                    script_path,
                    init_globals={
                        "input":              _web_input,
                        "_MSD_RUN_ONCE":      run_once,
                        "_MSD_INTERVAL_SEC":  app.config.get("MSD_INTERVAL_SEC", 30 * 60),
                        "_msd_auto_upsert":   _msd_auto_upsert,
                    },
                    run_name="__main__",
                )
            stdout_writer.flush()
            stderr_writer.flush()
        except _OtpCancelledError:
            # User cancelled OTP — close the Chrome window but keep the profile on disk
            # so the saved session (cookies) is preserved for the next run.
            logger.info("MSD auto-download cancelled by user (OTP dismissed).")
            try:
                # runpy executes the script in its own globals dict — the script's
                # `driver` variable lives there. We injected _web_input via init_globals
                # so we can reach the script module's globals through the exception
                # traceback frame.  Simplest safe approach: import the selenium driver
                # that was already instantiated at module level in the script by finding
                # it on the thread's current stack.
                import sys as _sys
                _frame = _sys._getframe()
                _drv = None
                while _frame is not None:
                    if "driver" in _frame.f_locals:
                        _candidate = _frame.f_locals["driver"]
                        # Make sure it's a Selenium WebDriver instance
                        try:
                            from selenium.webdriver.remote.webdriver import WebDriver as _WD
                            if isinstance(_candidate, _WD):
                                _drv = _candidate
                                break
                        except Exception:
                            pass
                    _frame = _frame.f_back
                if _drv is not None:
                    _drv.close()   # close window only — profile/cookies stay on disk
                    logger.info("Chrome window closed (session preserved).")
                else:
                    logger.warning("Could not locate Chrome driver to close window.")
            except Exception as _ce:
                logger.warning("Warning closing Chrome after OTP cancel: %s", _ce)
        except SystemExit as exc:
            code = exc.code if isinstance(exc.code, int) else 0
            if code == 0:
                logger.info("MSD WO auto-download finished.")
            else:
                logger.error("MSD WO auto-download exited with code %s.", code)
        except Exception:
            logger.error("MSD WO auto-download failed.")
            logger.error(traceback.format_exc())
        finally:
            global _msd_otp_pending, _msd_relogin_pending
            _msd_otp_pending = False
            _msd_relogin_pending = False
            os.sys.argv = original_argv
            logger.removeHandler(queue_handler)
            # Keep propagate=False permanently — this logger must never reach root

        # Trim files/msd-auto-download to the 5 most recent xlsx files
        _keep_latest_files(
            os.path.join(project_root, "files", "msd-auto-download"),
            keep=5,
            logger=logger,
        )


# ── Shared file-cleanup helper ────────────────────────────────────────────────

def _keep_latest_files(directory: str, keep: int = 5,
                       logger: logging.Logger | None = None) -> None:
    """Permanently delete the oldest .xlsx files in *directory*, keeping only
    the *keep* most recent ones.  Uses os.remove() — bypasses Recycle Bin."""
    if not os.path.isdir(directory):
        return
    files = sorted(
        [
            f for f in (
                os.path.join(directory, n) for n in os.listdir(directory)
            )
            if os.path.isfile(f) and f.lower().endswith(".xlsx")
        ],
        key=os.path.getmtime,
        reverse=True,   # newest first
    )
    for old_file in files[keep:]:
        try:
            os.remove(old_file)
            msg = f"[cleanup] Permanently deleted: {os.path.basename(old_file)}"
            if logger:
                logger.info(msg)
            else:
                print(msg)
        except OSError as exc:
            msg = f"[cleanup] Could not delete {os.path.basename(old_file)}: {exc}"
            if logger:
                logger.warning(msg)
            else:
                print(msg)


# ── RESOLV DC Updates ─────────────────────────────────────────────────────────

import collections as _col_resolve
_resolve_log_queue:   _queue.Queue       = _queue.Queue(maxsize=2000)
_resolve_log_history: _col_resolve.deque = _col_resolve.deque(maxlen=500)
_resolve_thread: threading.Thread | None = None
_resolve_lock    = threading.Lock()
_RESOLVE_BOOT_ID: str = _uuid.uuid4().hex

# Per-run upsert stats — keyed by AWB filename.
# { filename: {"new_dc": int, "new_awb": int, "status": "success"|"failed",
#              "upsert_date": "YYYY-MM-DD"} }
_resolve_upsert_stats: dict = {}

_RESOLVE_STATS_FILE = os.path.join(
    os.path.normpath(os.path.join(os.path.dirname(__file__), "..")),
    "templates", "admin", "upload_meta", "_resolve_upsert_stats.json",
)


def _resolve_stats_load() -> None:
    """Read persisted resolve stats from disk into the in-memory dict."""
    global _resolve_upsert_stats
    import json as _json
    try:
        if os.path.isfile(_RESOLVE_STATS_FILE):
            with open(_RESOLVE_STATS_FILE, "r", encoding="utf-8") as _f:
                _resolve_upsert_stats = _json.load(_f)
    except Exception:
        pass  # corrupt / missing — start fresh


def _resolve_stats_save() -> None:
    """Write the in-memory stats dict to disk atomically."""
    import json as _json, tempfile as _tmp
    try:
        os.makedirs(os.path.dirname(_RESOLVE_STATS_FILE), exist_ok=True)
        _payload = _json.dumps(_resolve_upsert_stats, indent=2)
        _dir = os.path.dirname(_RESOLVE_STATS_FILE)
        with _tmp.NamedTemporaryFile("w", dir=_dir, delete=False,
                                     suffix=".tmp", encoding="utf-8") as _tf:
            _tf.write(_payload)
            _tmp_path = _tf.name
        os.replace(_tmp_path, _RESOLVE_STATS_FILE)
    except Exception:
        pass  # non-fatal — next write will retry


_resolve_stats_load()


class _ResolveQueueHandler(logging.Handler):
    def emit(self, record: logging.LogRecord) -> None:  # type: ignore[override]
        if not record.name.startswith("resolve_auto_download"):
            return
        msg = record.getMessage()
        if _WERKZEUG_LINE_RE.search(msg):
            return
        rec = {
            "ts":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "level": record.levelname,
            "msg":   msg,
        }
        _resolve_log_history.append(rec)
        try:
            _resolve_log_queue.put_nowait(rec)
        except _queue.Full:
            try:
                _resolve_log_queue.get_nowait()
                _resolve_log_queue.put_nowait(rec)
            except (_queue.Full, _queue.Empty):
                pass


_resolve_queue_handler = _ResolveQueueHandler()
_resolve_queue_handler.setLevel(logging.INFO)


def _resolve_env_path() -> str:
    """Absolute path to the shared .env file (same file used by MSD)."""
    return _msd_env_path()


def _run_resolve_download_task(app) -> None:
    import runpy
    import traceback
    from contextlib import redirect_stderr, redirect_stdout

    class _QueueWriter:
        def __init__(self, logger: logging.Logger, level: int) -> None:
            self._logger = logger
            self._level  = level
            self._buffer = ""

        def write(self, value: str) -> int:
            self._buffer += value
            while "\n" in self._buffer:
                line, self._buffer = self._buffer.split("\n", 1)
                if line.strip():
                    self._logger.log(self._level, line)
            return len(value)

        def flush(self) -> None:
            if self._buffer.strip():
                self._logger.log(self._level, self._buffer.strip())
            self._buffer = ""

    project_root = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    )
    script_path = os.path.join(
        project_root, "app", "scripts", "resolve-auto-download",
        "lenovo_resolve_login.py",
    )
    awb_dir = os.path.join(project_root, "files", "resolve-auto-download", "Extract-AWB")
    dc_dir  = os.path.join(project_root, "files", "resolve-auto-download", "Extract-DC")
    os.makedirs(awb_dir, exist_ok=True)
    os.makedirs(dc_dir,  exist_ok=True)

    # Read credentials from the shared .env file
    env_vals  = _read_env_file(_resolve_env_path())
    username  = env_vals.get("RESOLVE_USERNAME", "")
    password  = env_vals.get("RESOLVE_PASSWORD", "")

    logger = logging.getLogger("resolve_auto_download")
    logger.setLevel(logging.INFO)
    logger.propagate = False
    logger.addHandler(_resolve_queue_handler)

    stdout_writer = _QueueWriter(logger, logging.INFO)
    stderr_writer = _QueueWriter(logger, logging.ERROR)

    _resolve_log_history.clear()

    # Add the script's own directory to sys.path so that
    # `from RecaptchaSolver import RecaptchaSolver` resolves correctly.
    script_dir = os.path.dirname(script_path)
    _resolve_path_inserted = script_dir not in os.sys.path
    if _resolve_path_inserted:
        os.sys.path.insert(0, script_dir)

    with app.app_context():
        try:
            logger.info("Starting RESOLV DC auto-download script...")
            original_argv = list(os.sys.argv)
            os.sys.argv = [script_path]
            with redirect_stdout(stdout_writer), redirect_stderr(stderr_writer):
                runpy.run_path(
                    script_path,
                    init_globals={
                        "USERNAME":         username,
                        "PASSWORD":         password,
                        "EXTRACT_AWB_DIR":  awb_dir,
                        "EXTRACT_DC_DIR":   dc_dir,
                    },
                    run_name="__main__",
                )
            stdout_writer.flush()
            stderr_writer.flush()
        except SystemExit as exc:
            code = exc.code if isinstance(exc.code, int) else 0
            if code == 0:
                logger.info("RESOLV DC auto-download finished.")
            else:
                logger.error("RESOLV DC auto-download exited with code %s.", code)
        except Exception:
            logger.error("RESOLV DC auto-download failed.")
            logger.error(traceback.format_exc())
        finally:
            os.sys.argv = original_argv
            # NOTE: handler removal is deferred to after post-run work so all
            # log lines (file copy, upsert stats) remain visible in the UI stream.

        # ── Post-run sync ────────────────────────────────────────────────────
        # The script always writes to its own Extract-AWB / Extract-DC dirs
        # (module-level assignments overwrite init_globals before they are used).
        # Copy any new files from the script dirs into the project-level
        # files/resolve-auto-download/ dirs so the web UI shows them correctly.
        import shutil as _shutil
        script_awb = os.path.join(script_dir, "Extract-AWB")
        script_dc  = os.path.join(script_dir, "Extract-DC")
        for src_folder, dst_folder in ((script_awb, awb_dir), (script_dc, dc_dir)):
            if not os.path.isdir(src_folder):
                continue
            os.makedirs(dst_folder, exist_ok=True)
            for fname in os.listdir(src_folder):
                if fname.startswith("."):          # skip .gitkeep etc.
                    continue
                src_file = os.path.join(src_folder, fname)
                dst_file = os.path.join(dst_folder, fname)
                if not os.path.isfile(src_file):
                    continue
                if not os.path.exists(dst_file):
                    try:
                        _shutil.copy2(src_file, dst_file)
                        logger.info("Copied %s → %s", fname, dst_folder)
                        # Remove from script-local dir after successful copy
                        os.remove(src_file)
                        logger.info("Removed source file: %s", fname)
                    except Exception as _cp_err:
                        logger.warning("Could not copy/remove %s: %s", fname, _cp_err)

        # Trim both project-level dirs to 5 most recent xlsx files
        for _dst in (awb_dir, dc_dir):
            _keep_latest_files(_dst, keep=5, logger=logger)

        # ── Upsert dc_number / return_status from the latest DC (GTAAP) excel ─
        # Applies the same hierarchy-enforced pass logic as the GTAAP Report
        # upsert button on data_import.html:
        #   Pass 1  — write real DC# to rows where db dc_number IS NULL
        #   Pass 1b — immediately promote return_status → DC GENERATED for those rows
        #   Pass 3  — forward-only status hierarchy write (PENDING WITH PARTNER →
        #             PENDING FOR DC GENERATION → DC GENERATED, locked)
        #   Pass 4  — absent open-status rows → UNKNOWN
        #   Pass 5a — whole-table promote: real dc_number but not DC GENERATED → promote
        #   Pass 5b — whole-table cleanup: DC GENERATED but no real dc_number → clear
        import glob as _g_dc
        _dc_pattern = os.path.join(dc_dir, "GTAAP_Report_export_*.xlsx")
        _dc_latest  = sorted(_g_dc.glob(_dc_pattern))
        _dc_fname   = os.path.basename(_dc_latest[-1]) if _dc_latest else None
        logger.info("[*] Upserting dc_number/return_status from latest DC excel...")
        if not _dc_fname:
            logger.warning("[!] No GTAAP_Report_export_*.xlsx found in %s — skipping DC upsert.", dc_dir)
        else:
            try:
                import pandas as _pd_dc
                from app.services.database.upsert import (
                    upsert_dc_from_gtaap as _upsert_dc_from_gtaap,
                )
                _dc_filepath = os.path.join(dc_dir, _dc_fname)
                _dc_df       = _pd_dc.read_excel(_dc_filepath, sheet_name="data")
                _db_path     = app.config["DATABASE_PATH"]
                _dc_conn     = open_db(_db_path)
                try:
                    _n_dc_new, _n_dc_status = _upsert_dc_from_gtaap(_dc_df, _dc_conn)
                finally:
                    _dc_conn.close()
                logger.info(
                    "[+] dc upsert complete: %d new DC#, %d status change(s).",
                    _n_dc_new, _n_dc_status,
                )
                from datetime import date as _date_dc
                _resolve_upsert_stats[_dc_fname] = {
                    "new_dc":        _n_dc_new,
                    "new_status":    _n_dc_status,
                    "status":        "success",
                    "upsert_date":   str(_date_dc.today()),
                }
                _resolve_stats_save()
            except Exception as _dc_exc:
                logger.error("[!] dc upsert failed: %s", _dc_exc)
                _resolve_upsert_stats[_dc_fname] = {
                    "new_dc": None, "new_status": None, "status": "failed",
                }
                _resolve_stats_save()

        # ── Upsert awb_resolv from the latest AWB excel ──────────────────────
        import glob as _g_awb
        _awb_pattern = os.path.join(awb_dir, "Generated_DCs_*.xlsx")
        _awb_latest  = sorted(_g_awb.glob(_awb_pattern))
        _awb_fname   = os.path.basename(_awb_latest[-1]) if _awb_latest else None
        logger.info("[*] Upserting awb_resolv from latest AWB excel...")
        try:
            from app.services.database.upsert import upsert_awb_resolv_from_awb_excel
            _db_path  = app.config["DATABASE_PATH"]
            _awb_conn = open_db(_db_path)
            try:
                _n_dc, _n_awb = upsert_awb_resolv_from_awb_excel(awb_dir, _awb_conn)
            finally:
                _awb_conn.close()
            logger.info(
                "[+] awb_resolv upsert complete: %d DC with new AWB / Date, %d row(s) updated.",
                _n_dc, _n_awb,
            )
            if _awb_fname:
                from datetime import date as _date_cls
                _resolve_upsert_stats[_awb_fname] = {
                    "new_dc":     _n_dc,
                    "new_awb":    _n_awb,
                    "status":     "success",
                    "upsert_date": str(_date_cls.today()),
                }
                _resolve_stats_save()
        except Exception as _awb_exc:
            logger.error("[!] awb_resolv upsert failed: %s", _awb_exc)
            if _awb_fname:
                _resolve_upsert_stats[_awb_fname] = {
                    "new_dc": None, "new_awb": None, "status": "failed",
                }
                _resolve_stats_save()

        # Remove queue handler now that all post-run logging is complete
        logger.removeHandler(_resolve_queue_handler)
        if _resolve_path_inserted and script_dir in os.sys.path:
            os.sys.path.remove(script_dir)


def _resolve_list_files(directory: str, limit: int = 5) -> list:
    """Return the *limit* most-recent file dicts for a directory."""
    result = []
    if not os.path.isdir(directory):
        return result
    for name in sorted(
        os.listdir(directory),
        key=lambda n: os.path.getmtime(os.path.join(directory, n)),
        reverse=True,
    ):
        if len(result) >= limit:
            break
        file_path = os.path.join(directory, name)
        if not os.path.isfile(file_path):
            continue
        if name.startswith("."):            # hide .gitkeep and other dot-files
            continue
        result.append({
            "name":         name,
            "size_kb":      round(os.path.getsize(file_path) / 1024, 1),
            "modified_fmt": datetime.fromtimestamp(
                os.path.getmtime(file_path)
            ).strftime("%Y-%m-%d %H:%M"),
        })
    return result


@admin_bp.route("/admin/resolve-dc-updates", methods=["GET"])
def resolve_dc_updates():
    project_root = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    )
    awb_dir = os.path.join(project_root, "files", "resolve-auto-download", "Extract-AWB")
    dc_dir  = os.path.join(project_root, "files", "resolve-auto-download", "Extract-DC")

    with _resolve_lock:
        is_running = _resolve_thread is not None and _resolve_thread.is_alive()

    # Enrich AWB file list with per-run upsert stats for initial page render
    awb_files_raw = _resolve_list_files(awb_dir)
    awb_files = []
    for f in awb_files_raw:
        stats = _resolve_upsert_stats.get(f["name"], {})
        awb_files.append({
            **f,
            "new_dc":        stats.get("new_dc",      None),
            "new_awb":       stats.get("new_awb",     None),
            "upsert_status": stats.get("status",      None),
            "upsert_date":   stats.get("upsert_date", None),
        })

    # Enrich DC file list with per-run upsert stats for initial page render
    dc_files_raw = _resolve_list_files(dc_dir)
    dc_files = []
    for f in dc_files_raw:
        stats = _resolve_upsert_stats.get(f["name"], {})
        dc_files.append({
            **f,
            "new_dc":        stats.get("new_dc",      None),
            "new_status":    stats.get("new_status",  None),
            "upsert_status": stats.get("status",      None),
            "upsert_date":   stats.get("upsert_date", None),
        })

    return render_template(
        "admin/export-import/resolve_dc_updates.html",
        portal="admin",
        active_page="resolve_dc_updates",
        active_group="data_import_export",
        awb_files=awb_files,
        dc_files=dc_files,
        is_running=is_running,
        boot_id=_RESOLVE_BOOT_ID,
    )


@admin_bp.route("/admin/resolve-dc-updates/trigger", methods=["POST"])
def resolve_dc_updates_trigger():
    global _resolve_thread
    with _resolve_lock:
        if _resolve_thread is not None and _resolve_thread.is_alive():
            return jsonify({"ok": False, "error": "RESOLV DC download is already running."}), 409
        app = current_app._get_current_object()
        _resolve_thread = threading.Thread(
            target=_run_resolve_download_task,
            args=(app,),
            daemon=True,
            name="resolve-auto-download",
        )
        _resolve_thread.start()
    return jsonify({"ok": True})


@admin_bp.route("/admin/resolve-dc-updates/stream", methods=["GET"])
def resolve_dc_updates_stream():
    def generate():
        import json as _json
        history = list(_resolve_log_history)
        for i, rec in enumerate(history):
            payload = dict(rec, history=True, history_first=(i == 0))
            yield f"data: {_json.dumps(payload)}\n\n"
        while True:
            try:
                rec = _resolve_log_queue.get(timeout=15)
                yield f"data: {_json.dumps(rec)}\n\n"
            except _queue.Empty:
                yield "data: {\"keepalive\": true}\n\n"

    return current_app.response_class(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@admin_bp.route("/admin/resolve-dc-updates/status", methods=["GET"])
def resolve_dc_updates_status():
    with _resolve_lock:
        is_running = _resolve_thread is not None and _resolve_thread.is_alive()
    return jsonify({"ok": True, "is_running": is_running})


@admin_bp.route("/admin/resolve-dc-updates/next-run", methods=["GET"])
def resolve_dc_updates_next_run():
    """Return seconds until next scheduled weekday slot (every 2h, 08:00–20:00 WIB)."""
    from datetime import datetime as _dt, timedelta as _td
    _next  = _next_resolve_slot()
    _now   = _dt.utcnow() + _td(hours=_SCHED_WIB_OFFSET)
    _secs  = max(0.0, (_next - _now).total_seconds())
    return jsonify({
        "ok":            True,
        "seconds_until": round(_secs),
        "next_run_wib":  _next.strftime("%Y-%m-%d %H:%M WIB"),
    })


@admin_bp.route("/admin/resolve-dc-updates/reset", methods=["POST"])
def resolve_dc_updates_reset():
    """Force-close Chrome on the resolve session profile and clear thread state."""
    global _resolve_thread

    import pathlib as _pl
    try:
        import psutil as _psutil
        session_dir = os.path.normpath(
            os.path.join(os.path.dirname(__file__), "..", "scripts",
                         "resolve-auto-download", "session")
        )
        target  = os.path.normcase(os.path.normpath(session_dir))
        killed  = 0
        for proc in _psutil.process_iter(["pid", "name", "cmdline"]):
            try:
                name = (proc.info["name"] or "").lower()
                if "chrome" not in name:
                    continue
                for arg in (proc.info["cmdline"] or []):
                    if "--user-data-dir=" in arg:
                        arg_path = os.path.normcase(
                            os.path.normpath(arg.split("=", 1)[1])
                        )
                        if arg_path == target:
                            proc.terminate()
                            killed += 1
                            break
            except (_psutil.NoSuchProcess, _psutil.AccessDenied):
                pass

        import time as _t
        if killed:
            _t.sleep(1.5)

        _profile_path = _pl.Path(session_dir)
        for _lk in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
            _lf = _profile_path / _lk
            if _lf.exists() or _lf.is_symlink():
                try:
                    _lf.unlink()
                except Exception:
                    pass

        killed_msg = (
            f"Chrome closed ({killed} process(es) terminated)."
            if killed else "No Chrome process found on this profile."
        )
    except ImportError:
        killed_msg = "psutil not installed — Chrome process not killed."
    except Exception as _exc:
        killed_msg = f"Chrome kill warning: {_exc}"

    with _resolve_lock:
        _resolve_thread = None

    return jsonify({"ok": True, "msg": killed_msg})


@admin_bp.route("/admin/resolve-dc-updates/credentials", methods=["GET"])
def resolve_dc_updates_credentials_get():
    env_path = _resolve_env_path()
    env      = _read_env_file(env_path)
    username = env.get("RESOLVE_USERNAME", "")
    has_pw   = bool(env.get("RESOLVE_PASSWORD", ""))
    return jsonify({"ok": True, "username": username, "has_password": has_pw})


@admin_bp.route("/admin/resolve-dc-updates/credentials", methods=["POST"])
def resolve_dc_updates_credentials_post():
    data     = request.get_json(silent=True) or {}
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", "")).strip()

    if not username:
        return jsonify({"ok": False, "error": "Username is required."}), 400
    if not password:
        return jsonify({"ok": False, "error": "Password is required."}), 400

    env_path = _resolve_env_path()
    try:
        _write_env_value(env_path, "RESOLVE_USERNAME", username)
        _write_env_value(env_path, "RESOLVE_PASSWORD", password)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Failed to write .env: {exc}"}), 500

    return jsonify({"ok": True})


@admin_bp.route("/admin/resolve-dc-updates/files", methods=["GET"])
def resolve_dc_updates_files():
    project_root = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..")
    )
    awb_dir = os.path.join(project_root, "files", "resolve-auto-download", "Extract-AWB")
    dc_dir  = os.path.join(project_root, "files", "resolve-auto-download", "Extract-DC")

    # Enrich AWB file list with per-run upsert stats
    awb_files_raw = _resolve_list_files(awb_dir)
    awb_files = []
    for f in awb_files_raw:
        stats = _resolve_upsert_stats.get(f["name"], {})
        awb_files.append({
            **f,
            "new_dc":        stats.get("new_dc",      None),
            "new_awb":       stats.get("new_awb",     None),
            "upsert_status": stats.get("status",      None),
            "upsert_date":   stats.get("upsert_date", None),
        })

    # Enrich DC file list with per-run upsert stats
    dc_files_raw = _resolve_list_files(dc_dir)
    dc_files = []
    for f in dc_files_raw:
        stats = _resolve_upsert_stats.get(f["name"], {})
        dc_files.append({
            **f,
            "new_dc":        stats.get("new_dc",      None),
            "new_status":    stats.get("new_status",  None),
            "upsert_status": stats.get("status",      None),
            "upsert_date":   stats.get("upsert_date", None),
        })

    return jsonify({
        "ok":        True,
        "awb_files": awb_files,
        "dc_files":  dc_files,
    })




# ── Resolve scheduler — weekdays 08:00–20:00 WIB, every 2 hours ─────────────
# Slots: 08:00 10:00 12:00 14:00 16:00 18:00 20:00

_resolve_scheduler_thread: threading.Thread | None = None
_resolve_scheduler_stop   = threading.Event()

# Scheduler constants (all in WIB = UTC+7)
_SCHED_WIB_OFFSET   = 7        # UTC+7
_SCHED_INTERVAL_H   = 2        # every 2 hours
_SCHED_WINDOW_START = 8        # 08:00 WIB (inclusive)
_SCHED_WINDOW_END   = 20       # 20:00 WIB (inclusive last slot)


def _next_resolve_slot() -> "datetime":
    """Return the next scheduled run datetime (in WIB, naive)."""
    from datetime import datetime as _dt, timedelta as _td
    now = _dt.utcnow() + _td(hours=_SCHED_WIB_OFFSET)
    # Build all slots for today and tomorrow, pick the first future weekday slot
    candidate = now.replace(minute=0, second=0, microsecond=0)
    # Step forward in 1-hour increments until we land on a valid slot
    for _ in range(7 * 24):  # safety ceiling: max 7 days ahead
        candidate += _td(hours=1)
        # Must be a weekday (Mon=0 … Fri=4)
        if candidate.weekday() >= 5:
            continue
        # Must be on an even slot within the window
        h = candidate.hour
        if h < _SCHED_WINDOW_START or h > _SCHED_WINDOW_END:
            continue
        if (h - _SCHED_WINDOW_START) % _SCHED_INTERVAL_H != 0:
            continue
        return candidate
    # Fallback — should never be reached
    return now + _td(hours=24)


def _resolve_scheduler_loop(app) -> None:
    """Loop: fire _run_resolve_download_task at every valid weekday slot."""
    global _resolve_thread  # declared here so the assignment below is valid
    from datetime import datetime as _dt, timedelta as _td
    _sched_logger = logging.getLogger("resolve_auto_download")
    _sched_logger.info(
        "[scheduler] Resolve scheduler started — weekdays %02d:00–%02d:00 WIB every %dh.",
        _SCHED_WINDOW_START, _SCHED_WINDOW_END, _SCHED_INTERVAL_H,
    )

    while not _resolve_scheduler_stop.is_set():
        next_run = _next_resolve_slot()
        now_wib  = _dt.utcnow() + _td(hours=_SCHED_WIB_OFFSET)
        wait_sec = max(0.0, (next_run - now_wib).total_seconds())
        _sched_logger.info(
            "[scheduler] Next Resolve auto-run at %s WIB (in %.0fs / %.1fh).",
            next_run.strftime("%Y-%m-%d %H:%M"),
            wait_sec,
            wait_sec / 3600,
        )

        # Sleep in 60-second ticks so the stop-event is checked regularly
        slept = 0.0
        while slept < wait_sec:
            if _resolve_scheduler_stop.is_set():
                return
            tick = min(60.0, wait_sec - slept)
            _resolve_scheduler_stop.wait(tick)
            slept += tick

        if _resolve_scheduler_stop.is_set():
            return

        # Skip if a manual run is already in flight
        with _resolve_lock:
            already = _resolve_thread is not None and _resolve_thread.is_alive()
        if already:
            _sched_logger.info("[scheduler] Skipping scheduled run — manual run already in progress.")
            continue

        _sched_logger.info(
            "[scheduler] Starting scheduled Resolve auto-run (%s WIB).",
            next_run.strftime("%H:%M"),
        )
        with _resolve_lock:
            if _resolve_thread is not None and _resolve_thread.is_alive():
                continue  # double-check inside the lock
            _resolve_thread = threading.Thread(
                target=_run_resolve_download_task,
                args=(app,),
                daemon=True,
                name="resolve-auto-download-scheduled",
            )
            _resolve_thread.start()

        # Wait for this run to finish before sleeping to the next slot
        _resolve_thread.join()
        _sched_logger.info("[scheduler] Scheduled Resolve run finished.")


def start_resolve_scheduler(app) -> None:
    """Start the background daily-scheduler thread (idempotent — safe to call twice)."""
    global _resolve_scheduler_thread
    if _resolve_scheduler_thread is not None and _resolve_scheduler_thread.is_alive():
        return
    _resolve_scheduler_stop.clear()
    _resolve_scheduler_thread = threading.Thread(
        target=_resolve_scheduler_loop,
        args=(app,),
        daemon=True,
        name="resolve-scheduler",
    )
    _resolve_scheduler_thread.start()


# ── Escalation Center page ────────────────────────────────────────────────────

@admin_bp.route("/admin/escalation-center", methods=["GET"])
def escalation_center():
    import os as _os
    project_root = _os.path.normpath(
        _os.path.join(_os.path.dirname(__file__), "..", "..")
    )
    xlsx_path = _os.path.join(project_root, "files", "source-db", "monday_link_map.xlsx")

    # Load board list from xlsx
    try:
        import sys
        scripts_dir = _os.path.join(project_root, "app", "scripts")
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)
        import monday_sync as _ms
        boards_raw = _ms.load_boards(xlsx_path)
    except Exception as exc:
        current_app.logger.warning("escalation_center: load_boards failed: %s", exc)
        boards_raw = []

    # Load sync state
    state_path = _os.path.join(project_root, "sync_state.json")
    try:
        import json as _json
        state = _json.loads(open(state_path, encoding="utf-8").read()) if _os.path.isfile(state_path) else {}
    except Exception:
        state = {}
    board_states = state.get("boards", {})

    board_meta = {}
    try:
        db = get_db()
        rows = db.execute(
            "SELECT monday_board_id, labor_vendor_related, customer_partner "
            "FROM asp_details WHERE monday_board_id IS NOT NULL"
        ).fetchall()
        board_meta = {
            str(row["monday_board_id"]): {
                "asp_id": row["labor_vendor_related"],
                "asp_name": row["customer_partner"],
            }
            for row in rows
        }
    except Exception:
        board_meta = {}

    # Enrich boards with per-board state
    boards = []
    for b in boards_raw:
        bid = str(b["board_id"])
        bs  = board_states.get(bid, {})
        meta = board_meta.get(bid, {})
        boards.append({
            "no":           b.get("no"),
            "asp_board":    b["asp_board"],
            "board_id":     bid,
            "asp_id":       meta.get("asp_id") or "—",
            "asp_name":     meta.get("asp_name") or "—",
            "last_sync":    bs.get("last_sync"),
            "total_synced": bs.get("total_synced", 0),
            "synced":       bool(bs.get("last_sync")),
        })

    # DB stats from lenovo_asp.db
    stats = {"items": 0, "updates": 0, "boards_synced": 0, "creators": 0}
    # latest_runs: dict keyed by board_id → most-recent sync_log row for that board
    latest_runs = {}
    edb = _get_monday_sync_db()
    if edb:
        try:
            stats["items"]        = edb.execute("SELECT COUNT(*) FROM technical_escalation").fetchone()[0]
            stats["updates"]      = edb.execute("SELECT COUNT(*) FROM item_updates").fetchone()[0]
            stats["boards_synced"] = edb.execute(
                "SELECT COUNT(DISTINCT board_id) FROM technical_escalation"
            ).fetchone()[0]
            stats["creators"]     = edb.execute("SELECT COUNT(*) FROM creators").fetchone()[0]
            # One latest row per board_id — run_at shifted to WIB (UTC+7)
            rows = edb.execute(
                "SELECT s1.board_id, s1.asp_board, s1.run_type, s1.items_found, "
                "s1.items_upserted, s1.updates_count, s1.duration_sec, "
                "datetime(s1.run_at, '+7 hours') AS run_at, "
                "COALESCE(t.total_items, 0) AS total_items "
                "FROM sync_log s1 "
                "LEFT JOIN ("
                "  SELECT board_id, COUNT(*) AS total_items"
                "  FROM technical_escalation GROUP BY board_id"
                ") t ON t.board_id = s1.board_id "
                "WHERE s1.id = (SELECT MAX(id) FROM sync_log s2 WHERE s2.board_id = s1.board_id) "
                "ORDER BY s1.run_at DESC"
            ).fetchall()
            latest_runs = {str(r["board_id"]): dict(r) for r in rows}
        except Exception:
            pass
        finally:
            edb.close()

    boards.sort(key=lambda b: (b["asp_board"] or "").lower())
    board_count = len(boards)
    synced_count = sum(1 for b in boards if b["synced"])

    with _sync_lock:
        is_running = _sync_thread is not None and _sync_thread.is_alive()

    return render_template(
        "admin/export-import/monday_case_update.html",
        portal="admin",
        active_page="monday_case_update",
        active_group="data_import_export",
        boards=boards,
        board_count=board_count,
        synced_count=synced_count,
        stats=stats,
        latest_runs=latest_runs,
        is_running=is_running,
    )


# ── Escalation Center: trigger sync ─────────────────────────────────────────

@admin_bp.route("/admin/escalation-center/trigger", methods=["POST"])
def escalation_center_trigger():
    global _sync_thread
    data     = request.get_json(silent=True) or {}
    mode     = data.get("mode", "incremental")
    board_id = data.get("board_id")  # optional — None means all boards

    with _sync_lock:
        if _sync_thread is not None and _sync_thread.is_alive():
            return jsonify({"ok": False, "error": "Sync already running"}), 409

        _sync_stop.clear()
        app = current_app._get_current_object()
        _sync_thread = threading.Thread(
            target=_run_sync_task,
            args=(app, mode, board_id),
            daemon=True,
            name="monday-sync",
        )
        _sync_thread.start()

    return jsonify({"ok": True, "mode": mode, "board_id": board_id})


# ── Escalation Center: stop sync ─────────────────────────────────────────────

@admin_bp.route("/admin/escalation-center/stop", methods=["POST"])
def escalation_center_stop():
    _sync_stop.set()
    try:
        _log_queue.put_nowait({"ts": "", "level": "WARNING",
                               "msg": "Stop signal sent — sync will halt after current item."})
    except _queue.Full:
        try:
            _log_queue.get_nowait()
            _log_queue.put_nowait({"ts": "", "level": "WARNING",
                                   "msg": "Stop signal sent — sync will halt after current item."})
        except (_queue.Full, _queue.Empty):
            pass
    return jsonify({"ok": True})


# ── Escalation Center: SSE log stream ─────────────────────────────────────────

@admin_bp.route("/admin/escalation-center/stream", methods=["GET"])
def escalation_center_stream():
    def generate():
        import json as _json
        while True:
            try:
                rec = _log_queue.get(timeout=15)
                yield f"data: {_json.dumps(rec)}\n\n"
            except _queue.Empty:
                yield "data: {\"keepalive\": true}\n\n"
    return current_app.response_class(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ── Escalation Center: status JSON ───────────────────────────────────────────

@admin_bp.route("/admin/escalation-center/status", methods=["GET"])
def escalation_center_status():
    import os as _os
    with _sync_lock:
        is_running = _sync_thread is not None and _sync_thread.is_alive()

    stats = {"items": 0, "updates": 0, "boards_synced": 0, "creators": 0}
    edb = _get_monday_sync_db()
    if edb:
        try:
            stats["items"]         = edb.execute("SELECT COUNT(*) FROM technical_escalation").fetchone()[0]
            stats["updates"]       = edb.execute("SELECT COUNT(*) FROM item_updates").fetchone()[0]
            stats["boards_synced"] = edb.execute(
                "SELECT COUNT(DISTINCT board_id) FROM technical_escalation"
            ).fetchone()[0]
            stats["creators"]      = edb.execute("SELECT COUNT(*) FROM creators").fetchone()[0]
        except Exception:
            pass
        finally:
            edb.close()

    return jsonify({
        "ok":             True,
        "is_running":     is_running,
        "in_window":      _is_active_window(),
        "next_run_at":    _next_run_at,      # Unix ts of next scheduled sync (in-window idle)
        "window_opens_at": _window_opens_at, # Unix ts of next window open (off-hours)
        "stats":          stats,
    })


# ── Escalation Center: boards + history refresh ──────────────────────────────

@admin_bp.route("/admin/escalation-center/boards", methods=["GET"])
def escalation_center_boards():
    import os as _os, json as _json
    project_root = _os.path.normpath(
        _os.path.join(_os.path.dirname(__file__), "..", "..")
    )
    xlsx_path  = _os.path.join(project_root, "files", "source-db", "monday_link_map.xlsx")
    state_path = _os.path.join(project_root, "sync_state.json")

    # Board list
    try:
        import sys
        scripts_dir = _os.path.join(project_root, "app", "scripts")
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)
        import monday_sync as _ms
        boards_raw = _ms.load_boards(xlsx_path)
    except Exception:
        boards_raw = []

    # Sync state
    try:
        state = _json.loads(open(state_path, encoding="utf-8").read()) if _os.path.isfile(state_path) else {}
    except Exception:
        state = {}
    board_states = state.get("boards", {})

    board_meta = {}
    try:
        db = get_db()
        rows = db.execute(
            "SELECT monday_board_id, labor_vendor_related, customer_partner "
            "FROM asp_details WHERE monday_board_id IS NOT NULL"
        ).fetchall()
        board_meta = {
            str(row["monday_board_id"]): {
                "asp_id": row["labor_vendor_related"],
                "asp_name": row["customer_partner"],
            }
            for row in rows
        }
    except Exception:
        board_meta = {}

    boards = []
    for b in boards_raw:
        bid = str(b["board_id"])
        bs  = board_states.get(bid, {})
        meta = board_meta.get(bid, {})
        boards.append({
            "no":           b.get("no"),
            "asp_board":    b["asp_board"],
            "board_id":     bid,
            "asp_id":       meta.get("asp_id") or "—",
            "asp_name":     meta.get("asp_name") or "—",
            "last_sync":    bs.get("last_sync"),
            "total_synced": bs.get("total_synced", 0),
            "synced":       bool(bs.get("last_sync")),
        })

    boards.sort(key=lambda b: (b["asp_board"] or "").lower())

    # Latest run per board
    latest_runs = {}
    edb = _get_monday_sync_db()
    if edb:
        try:
            rows = edb.execute(
                "SELECT s1.board_id, s1.asp_board, s1.run_type, s1.items_found, "
                "s1.items_upserted, s1.updates_count, s1.duration_sec, "
                "datetime(s1.run_at, '+7 hours') AS run_at, "
                "COALESCE(t.total_items, 0) AS total_items "
                "FROM sync_log s1 "
                "LEFT JOIN ("
                "  SELECT board_id, COUNT(*) AS total_items"
                "  FROM technical_escalation GROUP BY board_id"
                ") t ON t.board_id = s1.board_id "
                "WHERE s1.id = (SELECT MAX(id) FROM sync_log s2 WHERE s2.board_id = s1.board_id) "
                "ORDER BY s1.run_at DESC"
            ).fetchall()
            latest_runs = {str(r["board_id"]): dict(r) for r in rows}
        except Exception:
            pass
        finally:
            edb.close()

    return jsonify({"ok": True, "boards": boards, "latest_runs": latest_runs})


# ── Monday collector: daily totals persistence ───────────────────────────────

_MONDAY_DAILY_TOTALS_FILE = os.path.join(
    os.path.normpath(os.path.join(os.path.dirname(__file__), "..")),
    "templates", "admin", "upload_meta", "monday_daily_totals.json",
)


def _monday_daily_totals_load() -> dict:
    """Read the monday daily totals JSON from disk (returns {} on miss/error)."""
    import json as _j
    try:
        if os.path.isfile(_MONDAY_DAILY_TOTALS_FILE):
            with open(_MONDAY_DAILY_TOTALS_FILE, "r", encoding="utf-8") as f:
                return _j.load(f)
    except Exception:
        pass
    return {}


def _monday_daily_totals_save(data: dict) -> None:
    """Persist the monday daily totals JSON to disk."""
    import json as _j
    try:
        with open(_MONDAY_DAILY_TOTALS_FILE, "w", encoding="utf-8") as f:
            _j.dump(data, f, indent=2)
    except Exception:
        pass


@admin_bp.route("/admin/escalation-center/daily-totals", methods=["GET"])
def escalation_center_daily_totals_get():
    """Return the persisted monday daily totals JSON."""
    return jsonify({"ok": True, "data": _monday_daily_totals_load()})


@admin_bp.route("/admin/escalation-center/daily-totals", methods=["POST"])
def escalation_center_daily_totals_post():
    """Save updated monday daily totals sent from the browser."""
    payload = request.get_json(silent=True) or {}
    data = payload.get("data")
    if not isinstance(data, dict):
        return jsonify({"ok": False, "error": "Invalid payload"}), 400
    _monday_daily_totals_save(data)
    return jsonify({"ok": True})


# ── Monday token management ───────────────────────────────────────────────────

def _monday_env_path() -> str:
    """Absolute path to the .env file that holds MONDAY_TOKEN."""
    return _msd_env_path()   # reuse the same .env file as MSD credentials


@admin_bp.route("/admin/escalation-center/token", methods=["GET"])
def monday_token_get():
    """Return a masked view of the current Monday API token."""
    env      = _read_env_file(_monday_env_path())
    raw_tok  = env.get("MONDAY_TOKEN") or os.environ.get("MONDAY_TOKEN", "")
    if raw_tok:
        masked = raw_tok[:8] + "…" + raw_tok[-6:] if len(raw_tok) > 14 else "••••••••"
    else:
        masked = ""
    return jsonify({"ok": True, "token_set": bool(raw_tok), "token_masked": masked})


@admin_bp.route("/admin/escalation-center/token", methods=["POST"])
def monday_token_post():
    """Validate and save a new MONDAY_TOKEN to .env + os.environ."""
    data  = request.get_json(silent=True) or {}
    token = str(data.get("token", "")).strip()
    if not token:
        return jsonify({"ok": False, "error": "Token is required."}), 400

    # ── Verify token against Monday API before saving ────────────────────────
    import requests as _req
    try:
        resp = _req.post(
            "https://api.monday.com/v2",
            headers={
                "Authorization": token,
                "Content-Type":  "application/json",
                "API-Version":   "2024-01",
            },
            json={"query": "{ me { id name } }"},
            timeout=10,
        )
        if resp.status_code == 401:
            return jsonify({"ok": False, "error": "Token rejected by Monday.com (401 Unauthorized). Please check the token and try again."}), 400
        if resp.status_code == 403:
            return jsonify({"ok": False, "error": "Token rejected by Monday.com (403 Forbidden). The token may have expired or lack permissions."}), 400
        body = resp.json()
        if "errors" in body:
            msgs = "; ".join(e.get("message", str(e)) for e in body["errors"])
            return jsonify({"ok": False, "error": f"Monday API error: {msgs}"}), 400
        me = (body.get("data") or {}).get("me") or {}
        display_name = me.get("name") or ""
    except _req.exceptions.Timeout:
        return jsonify({"ok": False, "error": "Monday API timed out while verifying the token. Please try again."}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Could not verify token: {exc}"}), 400

    # ── Save to .env and reload into os.environ ──────────────────────────────
    env_path = _monday_env_path()
    try:
        _write_env_value(env_path, "MONDAY_TOKEN", token)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Failed to write .env: {exc}"}), 500

    os.environ["MONDAY_TOKEN"] = token
    return jsonify({"ok": True, "name": display_name})


@admin_bp.route("/admin/escalation-center/token/verify", methods=["GET"])
def monday_token_verify():
    """Quick-check whether the current token is valid.

    Resolution order (mirrors monday_sync.py):
      1. os.environ["MONDAY_TOKEN"]  (set at startup by load_dotenv or admin UI)
      2. MONDAY_TOKEN key in the .env file  (read fresh in case env isn't loaded)

    A network error or timeout is treated as "unknown" (not invalid) so the
    banner is never shown just because the Monday API is unreachable.
    """
    import requests as _req

    # Resolve the same token the sync script will actually use
    token = os.environ.get("MONDAY_TOKEN") or _read_env_file(_monday_env_path()).get("MONDAY_TOKEN", "")

    if not token:
        # No token in env or .env — skip check, don't show banner
        return jsonify({"ok": True, "valid": True, "name": "", "source": "none"})

    try:
        resp = _req.post(
            "https://api.monday.com/v2",
            headers={
                "Authorization": token,
                "Content-Type":  "application/json",
                "API-Version":   "2024-01",
            },
            json={"query": "{ me { id name } }"},
            timeout=10,
        )
        # Only a definitive auth rejection means "expired/invalid"
        if resp.status_code in (401, 403):
            return jsonify({"ok": True, "valid": False,
                            "reason": f"Monday.com rejected the token (HTTP {resp.status_code}). "
                                      f"The token may have expired — please update it."})
        body = resp.json()
        if "errors" in body:
            # Surface only auth-related GraphQL errors; others are not token issues
            auth_errors = [
                e for e in body["errors"]
                if "auth" in str(e).lower() or "token" in str(e).lower()
                   or "not authorized" in str(e).lower()
            ]
            if auth_errors:
                msgs = "; ".join(e.get("message", str(e)) for e in auth_errors)
                return jsonify({"ok": True, "valid": False, "reason": msgs})
            # Non-auth GraphQL errors — token is fine
            return jsonify({"ok": True, "valid": True, "name": ""})
        me   = (body.get("data") or {}).get("me") or {}
        name = me.get("name") or ""
        return jsonify({"ok": True, "valid": True, "name": name})
    except Exception:
        # Network error, timeout, etc. — can't determine validity; don't show banner
        return jsonify({"ok": True, "valid": True, "name": "", "source": "unreachable"})


# ── Monday Data page ──────────────────────────────────────────────────────────

@admin_bp.route("/admin/monday-data", methods=["GET"])
def monday_data():
    """Render the Monday Data page shell — data is loaded by the JS via API."""
    edb = _get_monday_sync_db()
    boards_list = []
    total_count = 0
    board_count = 0

    if edb:
        try:
            total_count = edb.execute(
                "SELECT COUNT(*) FROM technical_escalation"
            ).fetchone()[0]
            board_raw = edb.execute("""
                SELECT asp_board, board_id, COUNT(*) AS item_count
                FROM technical_escalation
                GROUP BY board_id
                ORDER BY asp_board
            """).fetchall()
            boards_list = [dict(r) for r in board_raw]
            board_count = len(boards_list)
        except Exception:
            pass
        finally:
            edb.close()

    return render_template(
        "admin/escalation_center/monday_data.html",
        portal="admin",
        active_page="monday_data",
        active_group="escalation_center",
        boards=boards_list,
        total_count=total_count,
        board_count=board_count,
    )


# ── Monday Data: meta API (boards + status counts — lightweight) ──────────────

@admin_bp.route("/admin/api/monday-data/meta", methods=["GET"])
def monday_data_meta():
    """Return board list and status counts only — no row data.
    Used to populate the sidebar and status dropdown without pulling all rows."""
    edb = _get_monday_sync_db()
    boards_list  = []
    status_list  = []
    total_count  = 0

    if edb:
        try:
            total_count = edb.execute(
                "SELECT COUNT(*) FROM technical_escalation"
            ).fetchone()[0]

            board_raw = edb.execute("""
                SELECT asp_board, board_id, COUNT(*) AS item_count
                FROM technical_escalation
                GROUP BY board_id
                ORDER BY asp_board
            """).fetchall()
            boards_list = [dict(r) for r in board_raw]

            status_raw = edb.execute("""
                SELECT COALESCE(status, '') AS status, COUNT(*) AS cnt
                FROM technical_escalation
                GROUP BY status
            """).fetchall()
            status_list = [dict(r) for r in status_raw]
        except Exception:
            pass
        finally:
            edb.close()

    board_names = {b["board_id"]: b["asp_board"] for b in boards_list}
    return jsonify({
        "total_count": total_count,
        "boards": boards_list,
        "board_names": board_names,
        "statuses": status_list,
    })


# ── Monday Data: JSON API (paginated, filtered) ────────────────────────────────

@admin_bp.route("/admin/api/monday-data", methods=["GET"])
def monday_data_api():
    """Return a paginated, filtered page of Monday sync rows.
    Query params:
      page     (int, default 1)
      per_page (int, default 50, max 200)
      board_id (str, optional)
      status   (str, optional — comma-separated group like 'in progress')
      q        (str, optional — search across item_name, wo_case_id, serial_number, status)
    """
    from flask import request as _req

    # ── parse params ──────────────────────────────────────────────
    try:
        page = max(1, int(_req.args.get("page", 1)))
    except (ValueError, TypeError):
        page = 1
    try:
        per_page = min(200, max(1, int(_req.args.get("per_page", 50))))
    except (ValueError, TypeError):
        per_page = 50

    board_id   = (_req.args.get("board_id") or "").strip()
    status_arg = (_req.args.get("status")   or "").strip()
    q          = (_req.args.get("q")        or "").strip()

    # Status grouping (mirrors the JS STATUS_GROUPS)
    STATUS_GROUPS = {
        "in progress": ["technical escalation", "progress", "qa result"],
    }

    # ── build WHERE clauses ───────────────────────────────────────
    where_parts  = []
    params       = []

    if board_id:
        where_parts.append("te.board_id = ?")
        params.append(board_id)

    if status_arg:
        members = STATUS_GROUPS.get(status_arg.lower())
        if members:
            placeholders = ",".join("?" * len(members))
            where_parts.append(f"LOWER(COALESCE(te.status,'')) IN ({placeholders})")
            params.extend(members)
        else:
            where_parts.append("LOWER(COALESCE(te.status,'')) = ?")
            params.append(status_arg.lower())

    if q:
        where_parts.append(
            "(te.item_name LIKE ? OR te.wo_case_id LIKE ? "
            "OR te.serial_number LIKE ? OR te.status LIKE ?)"
        )
        like = f"%{q}%"
        params.extend([like, like, like, like])

    where_sql = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    edb       = _get_monday_sync_db()
    rows_list = []
    total     = 0

    if edb:
        try:
            # ── total count for this filter ───────────────────────
            count_sql = f"SELECT COUNT(*) FROM technical_escalation te {where_sql}"
            total = edb.execute(count_sql, params).fetchone()[0]

            # ── disc_count subquery (no expensive GROUP-BY JOIN) ──
            offset = (page - 1) * per_page
            data_sql = f"""
                SELECT
                    te.monday_item_id,
                    te.board_id,
                    te.asp_board,
                    te.item_name,
                    te.item_created_at,
                    te.item_updated_at,
                    te.db_synced_at,
                    te.status,
                    te.work_order_type,
                    te.wo_case_id,
                    te.serial_number,
                    te.location,
                    te.ppsn_category,
                    te.rrr_category,
                    te.diag_datetime,
                    te.diag_agent_ce,
                    te.diag_model,
                    te.diag_warranty,
                    te.diag_problem,
                    te.diag_esc_approval,
                    te.diag_parts_request,
                    te.diagnose_note,
                    te.repair_note,
                    c.creator_name,
                    (
                        SELECT COUNT(DISTINCT u2.update_id) + COUNT(DISTINCT r2.reply_id)
                        FROM item_updates u2
                        LEFT JOIN item_update_replies r2 ON u2.update_id = r2.update_id
                        WHERE u2.monday_item_id = te.monday_item_id
                    ) AS disc_count,
                    te.has_wo,
                    CASE WHEN (te.wo_case_id IS NULL OR te.wo_case_id = '') AND te.has_wo = 1
                        THEN (
                            SELECT ws.work_order_id FROM wo_summary ws
                            WHERE LOWER(ws.serial_number) = LOWER(te.serial_number)
                            ORDER BY ws.created_on DESC LIMIT 1
                        )
                        ELSE NULL
                    END AS latest_wo_id
                FROM technical_escalation te
                LEFT JOIN creators c ON te.creator_id = c.creator_id
                {where_sql}
                ORDER BY te.item_created_at DESC
                LIMIT ? OFFSET ?
            """
            raw = edb.execute(data_sql, params + [per_page, offset]).fetchall()
            rows_list = [dict(r) for r in raw]
        except Exception as _e:
            current_app.logger.error("monday_data_api query failed: %s", _e)
        finally:
            edb.close()

    import math
    pages = max(1, math.ceil(total / per_page)) if total else 1
    return jsonify({
        "rows":  rows_list,
        "total": total,
        "page":  page,
        "pages": pages,
        "per_page": per_page,
    })


# ── Monday Data: single item detail API ───────────────────────────────────────

@admin_bp.route("/admin/api/monday-data/item/<item_id>", methods=["GET"])
def monday_data_item(item_id):
    """Return the full detail row for a single Monday item (used by the detail drawer)."""
    edb = _get_monday_sync_db()
    row = None
    if edb:
        try:
            raw = edb.execute("""
                SELECT
                    te.monday_item_id, te.board_id, te.asp_board, te.item_name,
                    te.item_created_at, te.item_updated_at, te.db_synced_at,
                    te.status, te.work_order_type, te.wo_case_id, te.serial_number,
                    te.location, te.ppsn_category, te.rrr_category,
                    te.diag_datetime, te.diag_agent_ce, te.diag_model,
                    te.diag_warranty, te.diag_problem, te.diag_esc_approval,
                    te.diag_parts_request, te.diagnose_note, te.repair_note,
                    c.creator_name
                FROM technical_escalation te
                LEFT JOIN creators c ON te.creator_id = c.creator_id
                WHERE te.monday_item_id = ?
            """, (item_id,)).fetchone()
            if raw:
                row = dict(raw)
        except Exception:
            pass
        finally:
            edb.close()
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(row)


# ── Monday Data: discussion JSON API ─────────────────────────────────────────

@admin_bp.route("/admin/monday-data/discussion/<item_id>", methods=["GET"])
def monday_data_discussion(item_id):
    edb = _get_monday_sync_db()
    result = {"updates": []}
    if not edb:
        return jsonify(result)

    try:
        # Load all updates for this item
        updates = edb.execute("""
            SELECT u.update_id, u.body_text, u.created_at, u.updated_at,
                   u.creator_id, c.creator_name
            FROM item_updates u
            LEFT JOIN creators c ON u.creator_id = c.creator_id
            WHERE u.monday_item_id = ?
            ORDER BY u.created_at ASC
        """, (item_id,)).fetchall()

        updates_out = []
        for upd in updates:
            upd_dict = dict(upd)
            # Load replies for this update
            replies = edb.execute("""
                SELECT r.reply_id, r.body_text, r.created_at,
                       r.creator_id, c.creator_name
                FROM item_update_replies r
                LEFT JOIN creators c ON r.creator_id = c.creator_id
                WHERE r.update_id = ?
                ORDER BY r.created_at ASC
            """, (upd_dict["update_id"],)).fetchall()
            upd_dict["replies"] = [dict(r) for r in replies]
            updates_out.append(upd_dict)

        result["updates"] = updates_out
    except Exception:
        pass
    finally:
        edb.close()

    return jsonify(result)
